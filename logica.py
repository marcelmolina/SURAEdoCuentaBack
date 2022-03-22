import datetime

import openpyxl as opyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import landscape
from reportlab.lib.units import mm,inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image, PageBreak
import string
from pytz import timezone

from reportlab.pdfbase import pdfmetrics, ttfonts
from io import BytesIO
import cx_Oracle
import os
from dotenv import load_dotenv
from apoyo import getTableNamesUDI
from apoyo import getHeadColumnsUDI
from apoyo import get_tablas_referencia
from apoyo import getheaderforcompressed
from apoyo import getcolumnstosum
from apoyo import getHeadColumnsBonos
from apoyo import getHeadColumnsComisones
from apoyo import getTipoSubBono
from apoyo import getTableNamesComisiones
from apoyo import getquery
from apoyo import getheaderpdf
from apoyo import getcolumnstosum_udi
import asyncio
import cx_Oracle_async

async def get_oracle_pool(app):
	try:
		load_dotenv()
		DB_HOST = os.getenv('DB_HOST')
		DB_PORT = os.getenv('DB_PORT')
		DB_SERVICE = os.getenv('DB_SERVICE')
		DB_USER = os.getenv('DB_USER')
		DB_PWD = os.getenv('DB_PWD')
		DB_SCHEMA = os.getenv('DB_SCHEMA')
		host = DB_HOST
		port = DB_PORT
		service_name = DB_SERVICE
		user = DB_USER
		password = DB_PWD
		schema = DB_SCHEMA
		sid = cx_Oracle_async.makedsn(host, port, service_name=service_name)
		try:
			pool = await cx_Oracle_async.create_pool(user,password,sid)
			return True,"",pool
		except Exception as ex:
			app.logger.error(ex)
			return False, 'Error en la conexion con la base de datos.', 0
	except Exception as ex:
		app.logger.error(ex)
		return False, 'Error en la configuracion de la base de datos.', 0

def getconexion(app):
	try:
		load_dotenv()
		DB_HOST = os.getenv('DB_HOST')
		DB_PORT = os.getenv('DB_PORT')
		DB_SERVICE = os.getenv('DB_SERVICE')
		DB_USER = os.getenv('DB_USER')
		DB_PWD = os.getenv('DB_PWD')
		DB_SCHEMA = os.getenv('DB_SCHEMA')
		host = DB_HOST
		port = DB_PORT
		service_name = DB_SERVICE
		user = DB_USER
		password = DB_PWD
		schema = DB_SCHEMA
		sid = cx_Oracle.makedsn(host, port, service_name=service_name)
		try:
			connection = cx_Oracle.connect(f"{user}/{password}@{host}:{port}/{service_name}")
			return True,"",connection
		except Exception as ex:
			app.logger.error(ex)
			return False, 'Error en la conexion con la base de datos.', 0
	except Exception as ex:
		app.logger.error(ex)
		return False, 'Error en la configuracion de la base de datos.', 0

def getperiodo(P_CLAVE,P_MES,P_ANIO,app):
	try:
		tabla = "TSIA_MOVAGEN"
		if P_CLAVE == 'P':
			tabla = "TSIA_MOVPROM"
		con_est, con_mssg, connection = getconexion(app)
		if not con_est:
			return False, con_mssg, 0, 0
		app.logger.info("Conectado a la base de datos.")
		entro = False
		P_Fefin = ""
		P_Feini = ""
		query1 = f"SELECT MAX(femovimi)-1 FROM {tabla} WHERE cdconc = 'SA' AND TO_CHAR(femovimi,'MMYYYY') = LPAD('{P_MES}',2,'0')||'{P_ANIO}'"
		cur1 = connection.cursor()
		cur1.execute(query1)
		for data in cur1:
			entro = True
			P_Fefin = data[0].strftime('%Y-%m-%d')
		if entro:
			entro = False
			query2 = f"SELECT MAX(femovimi) FROM {tabla} WHERE cdconc = 'SA' AND femovimi < TO_DATE('{P_Fefin}','YYYY-MM-DD')"
			cur2 = connection.cursor()
			cur2.execute(query2)
			for data in cur2:
				entro = True
				P_Feini = data[0].strftime('%Y-%m-%d')
		if not entro:
			return False,"Error en la busqueda de fechas.",0,0
		return True,"",P_Feini,P_Fefin
	except Exception as ex:
		app.logger.error(ex)
		return False, 'Error en la busqueda de fechas', 0,0

async def comisiones_xlsx(P_Clave,P_Feini,P_Fefin,P_COD,app):
	try:
		app.logger.info("Entrando a Estado de cuentas de Comisiones XLSX ("+P_COD+")")
		con_est, con_mssg, pool = await get_oracle_pool(app)
		if not con_est:
			return False, con_mssg, 0, 0, 0
		app.logger.info("Conectado a la base de datos.")
		has_agent = False
		app.logger.info("Iniciando carga de cursores")
		cursors = await tasks(app, 13, P_COD, 'COMISION', P_Clave, P_Feini, P_Fefin, pool)
		if 'Excepcion' in cursors:
			return False, 'Hubo un error obteniendo la data', 0, 0, 0
		app.logger.info("Cargados todos los cursores.")
		libro_nombre =P_Clave+"_" + P_Feini.replace("/", "")+"_"+ P_Fefin.replace("/", "")+'_comisiones.xlsx'
		plantilla = "plantilla_agentes.xlsx"
		if P_COD == 'P':
			plantilla = "plantilla_promotor.xlsx"
		wb = opyxl.load_workbook(plantilla)
		app.logger.info("Cargado archivo de plantilla xlsx.")
		ws = wb.worksheets[0]
		ws.cell(row=1, column=6).value = "ESTADOS DE CUENTA DE COMISIONES"
		ws.title = "Estado de Cuenta de Comisiones"
		has_agent = False
		app.logger.info("Leyendo cursor de cabecera")
		for row in cursors[0]:
			has_agent = True
			for i in range(0, len(row) - 4):
				ws.cell(row=4 + i, column=9).value = row[i]
			for i in range(len(row) - 4, len(row)):
				ws.cell(row=i, column=4).value = row[i]
		if not has_agent:
			app.logger.error("La cabecera volvio vacia.")
			return False,'Identificador no encontrado',0,0,0
		del cursors[0]
		f = 14  # principal gestor de filas del archivo
		greyFill = PatternFill(fill_type='solid', start_color='10b0c2', end_color='10b0c2')
		# NUEVO BLOQUE SECUENCIAL
		c_count = 1
		for cursor in cursors:
			app.logger.info(f"Leyendo cursor -> ({c_count})")
			fila_totales = [0, 0, 0, 0, 0]
			lista = getHeadColumnsComisones("excel", c_count)
			alphabet_string = string.ascii_uppercase
			alphabet_list = list(alphabet_string)
			ws.cell(row=f, column=1).value = getTableNamesComisiones(c_count)
			ws.cell(row=f, column=1).font = Font(name='Arial', size=9, bold=True)
			f += 1
			j = 0
			for item in lista:
				ws.cell(row=f, column=j + 1).value = item
				ws.cell(row=f, column=j + 1).border = Border(left=Side(style='thin'), right=Side(style='thin'),
															 top=Side(style='thin'), bottom=Side(style='thin'))
				ws.cell(row=f, column=j + 1).fill = greyFill
				ws.cell(row=f, column=j + 1).font = Font(name='Arial', size=9, bold=True,color='ffffff')
				ws.cell(row=f, column=j + 1).alignment = Alignment(horizontal="center", vertical="center")
				columna = alphabet_list[ws.cell(row=f, column=j + 1).column - 1]
				multiplicador = 2
				lista_columnas_esp = ['A', 'B', 'C', 'H']
				ancho = len(item)
				if len(item) <= 6:
					ancho *= 2
				if len(item) > 6 or columna == 'A':
					ancho *= 1.1
				if columna in lista_columnas_esp:
					ancho = 25
				if c_count == 5:
					ws.column_dimensions[columna].width = ancho
				j += 1
			j = 0
			k = 0
			f += 1
			has_data= False
			for row in cursor:
				for i in range(0, len(row)):
					if c_count not in [5, 6, 9, 10]:
						valor = row[i]
						if c_count in [1, 2]:
							if i != 0:
								valor = "{:,.2f}".format(valor)
						if c_count in [3, 4, 7, 8]:
							if i != 0:
								valor = "{:,.2f}".format(valor)
						if c_count in [11, 12]:
							if i in [1, 3, 6, 7, 8]:
								valor = "{:,.2f}".format(valor)
						ws.cell(row=f, column=i + 1).value = valor
						ws.cell(row=f, column=i + 1).border= Border(left=Side(style='thin'), right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
						ws.cell(row=f, column=i + 1).alignment = Alignment(horizontal="center", vertical="center")
						if len(str(valor)) > 17:
							ws.cell(row=f, column=i + 1).font = Font(name='Arial', size=8)
							if len(str(valor)) > 25:
								ws.cell(row=f, column=i + 1).font = Font(name='Arial', size=7)
					else:
						valor = row[i]
						if i in [10, 11, 14, 15]:
							valor = "{:,.2f}".format(valor)
						if i == 10:
							fila_totales[0] += abs(row[i])
						if i == 11:
							fila_totales[1] += abs(row[i])
						if i == 13:
							fila_totales[2] += abs(row[i])
						if i == 14:
							fila_totales[3] += abs(row[i])
						if i == 15:
							fila_totales[4] += abs(row[i])
						has_data = True
						ws.cell(row=f, column=i + 1).value = valor
						ws.cell(row=f, column=i + 1).border = Border(left=Side(style='thin'), right=Side(style='thin'),
																	 top=Side(style='thin'), bottom=Side(style='thin'))
						ws.cell(row=f, column=i + 1).alignment = Alignment(horizontal="center", vertical="center")
						if len(str(valor)) > 17:
							ws.cell(row=f, column=i + 1).font = Font(name='Arial', size=8)
							if len(str(valor)) > 25:
								ws.cell(row=f, column=i + 1).font = Font(name='Arial', size=7)
				f += 1
			#cursor.close()
			if c_count in [5, 6, 9, 10] and has_data:
				fila_totales[0] = "{:,.2f}".format(fila_totales[0])
				fila_totales[1] = "{:,.2f}".format(fila_totales[1])
				fila_totales[2] = "{:,.2f}".format(fila_totales[2])
				fila_totales[3] = "{:,.2f}".format(fila_totales[3])
				fila_totales[4] = "{:,.2f}".format(fila_totales[4])
				ws.cell(row=f, column=10).value = "TOTAL"
				ws.cell(row=f, column=10).alignment = Alignment(horizontal="center", vertical="center")
				ws.cell(row=f, column=11).value = fila_totales[0]
				ws.cell(row=f, column=11).alignment = Alignment(horizontal="center", vertical="center")
				ws.cell(row=f, column=12).value = fila_totales[1]
				ws.cell(row=f, column=12).alignment = Alignment(horizontal="center", vertical="center")
				ws.cell(row=f, column=14).value = fila_totales[2]
				ws.cell(row=f, column=14).alignment = Alignment(horizontal="center", vertical="center")
				ws.cell(row=f, column=15).value = fila_totales[3]
				ws.cell(row=f, column=15).alignment = Alignment(horizontal="center", vertical="center")
				ws.cell(row=f, column=16).value = fila_totales[4]
				ws.cell(row=f, column=16).alignment = Alignment(horizontal="center", vertical="center")
				f += 1
			f += 1
			c_count += 1
		# fin de bloque
		f += 2
		ramos, ramo_style, bonos, bono_style = get_tablas_referencia()
		t_apoyo = [ramos]
		for tabla in t_apoyo:
			c_linea = 0
			for linea in tabla:
				c_columna = 0
				for columna in linea:
					if c_linea in [0, 1]:
						if c_linea==0 and c_columna==0:
							top_1 = ws.cell(row=f, column=1)
							ws.merge_cells(start_row=f, start_column=1, end_row=f, end_column=8)
							top_1.value = columna
							top_1.fill = greyFill
							top_1.font = Font(name='Arial', size=9, bold=True,
																			 color='ffffff')
							top_1.alignment = Alignment(horizontal="center", vertical="center")
						if c_linea==1:
							ws.cell(row=f, column=c_columna + 1).value = columna
							ws.cell(row=f, column=c_columna + 1).fill = greyFill
							ws.cell(row=f, column=c_columna + 1).font = Font(name='Arial', size=9, bold=True,
																			 color='ffffff')
							ws.cell(row=f, column=c_columna + 1).alignment = Alignment(horizontal="center", vertical="center")
					else:
						ws.cell(row=f, column=c_columna + 1).value = columna
						ws.cell(row=f, column=c_columna + 1).alignment = Alignment(horizontal="center",
																				   vertical="center")
					c_columna += 1
				c_linea += 1
				f += 1
			f += 1
		virtual_wb = BytesIO()
		app.logger.info("Construyendo reporte xlsx.")
		wb.save(virtual_wb)
		return True,"",virtual_wb.getvalue(), wb.mime_type,libro_nombre
	except Exception as ex:
		app.logger.error(ex)
		return False, 'Error generando el reporte', 0, 0,0


async def dotask(app,i,connection,query):
	try:
		async with connection.cursor() as cursor:
			await cursor.execute(query)
			app.logger.info(f"Retorna data del cursor #{i}.")
			return await cursor.fetchall()
	except Exception as ex:
		app.logger.error(ex)
		return "Excepcion"

async def tasks(app,number,P_COD, tipo, P_Clave, P_Feini, P_Fefin,pool):
	try:
		async with pool.acquire() as connection:
			alltasks = []
			for i in range(number):
				alltasks.append(asyncio.create_task(dotask(app,i,connection,getquery(P_COD, tipo, i, P_Clave, P_Feini, P_Fefin))))
			cursors = await asyncio.gather(*alltasks)
			return cursors
	except Exception as ex:
		app.logger.error(ex)
		return "Excepcion"


async def comisiones_pdf(P_Clave,P_Feini,P_Fefin,P_COD,app):
	try:
		pdfmetrics.registerFont(ttfonts.TTFont("Arial", "Arial.ttf"))
		pdfmetrics.registerFont(ttfonts.TTFont("Arial_Bold", "Arial_Bold.ttf"))
		app.logger.info("Entrando a Estado de cuentas de Comisiones PDF (" + P_COD + ")")
		con_est, con_mssg, pool = await get_oracle_pool(app)
		if not con_est:
			return False, con_mssg, 0, 0, 0
		app.logger.info("Conectado a la base de datos.")
		has_agent = False
		app.logger.info("Iniciando carga de cursores")
		cursors = await tasks(app,13,P_COD, 'COMISION', P_Clave, P_Feini, P_Fefin,pool)
		if 'Excepcion' in cursors:
			return False, 'Hubo un error obteniendo la data', 0, 0, 0
		app.logger.info("Cargados todos los cursores.")
		libro_nombre =P_Clave+"_" + P_Feini.replace("/", "")+"_"+ P_Fefin.replace("/", "")+'_comisiones.pdf'

		virtual_wb = BytesIO()
		doc = SimpleDocTemplate(virtual_wb,pagesize=landscape((475*mm, 600*mm)),topMargin=45*mm)
		flowables = []
		app.logger.info("Leyendo cursor de cabecera")
		for row in cursors[0]:
			has_agent = True
			lista_aux = []
			for i in range(0, len(row)):
				lista_aux.append(row[i])
		if not has_agent:
			app.logger.error("La cabecera volvio vacia.")
			return False, 'Identificador no encontrado', 0, 0,0
		header_all = getheaderpdf(P_COD,lista_aux,'COMISIONES')
		grid = [('FONTNAME', (0, 0), (-1,-1), 'Arial_Bold'), ('ALIGN', (0, 0), (-1, -1), 'LEFT')]
		grid2 = [('ALIGN', (0, 0), (-1, -1), 'LEFT'),
				 ('FONTNAME',  (1, 0), (1, -1), 'Arial_Bold'),
				 ('FONTNAME',  (4, 0), (4, -1), 'Arial_Bold')]
		tbl_header = Table(header_all, hAlign='LEFT')
		tbl_header.setStyle(grid2)
		#flowables.append(tbl)
		del cursors[0]
		c_count = 1
		empty_cursors = []
		#,('FONTSIZE', (0, 0), (0, 0), 7)
		data_for_f_new_table=[]
		data_for_s_new_table = []
		data_for_t_new_table = []
		styl_new_count = 0

		style_newx2 = TableStyle(
			[('LINEABOVE', (0, 0), (-1, -1), 0.25, colors.white), ('ALIGN', (0, 0), (-1, -1), 'CENTER')])
		style_newx2.add('FONTNAME', (0, 0), (-1, 0), 'Arial_Bold')
		style_newx2.add('TEXTCOLOR', (0, 0), (-1, 0), colors.white)
		style_newx2.add('BACKGROUND', (0, 0), (-1, 0), '#10b0c2')
		style_newx2.add('BACKGROUND', (0, 1), (-1, 1), '#e8eaea')
		style_newx2.add('TEXTCOLOR', (0, 1), (-1, 1), colors.black)
		style_newx2.add('FONTNAME', (0, 1), (-1, 1), 'Arial')
		style_newx2.add('BACKGROUND', (0, 2), (-1, 2), '#e2e4e4')
		style_newx2.add('TEXTCOLOR', (0, 2), (-1, 2), colors.black)
		style_newx2.add('FONTNAME', (0, 2), (-1, 2), 'Arial')
		style_for_f_new_table=TableStyle(
				[('LINEABOVE', (0, 0), (-1, -1), 0.25, colors.white), ('ALIGN', (0, 0), (-1, -1), 'CENTER')])
		for cursor in cursors:
			tblstyle = TableStyle(
				[('LINEABOVE', (0, 0), (-1, -1), 0.25, colors.white), ('ALIGN', (0, 0), (-1, -1), 'CENTER')])
			app.logger.info(f"Leyendo cursor -> ({c_count})")
			lista = getHeadColumnsComisones("pdf", c_count)
			data_cursor = []
			theader=[]
			theader.append([getTableNamesComisiones(c_count)])
			taux = Table(theader,hAlign='LEFT')
			taux.setStyle(grid)
			data_cursor.append(lista)
			fila_totales = []
			if c_count  in [5,6,9,10]:
				fila_totales = ["", "", "", "", "", "", "", "TOTAL", 0, 0, "", 0, 0, 0, "", ""]
			if c_count in [1,2]:
				fila_totales= ["TOTAL",0,0,0,0,0,0,0]
			if c_count in [3,4,7,8]:
				fila_totales= ["TOTAL",0]
			if c_count in [11,12]:
				fila_totales= ["TOTAL",0,0,0," ","TOTAL PAGADO",0,0,0]
			numrow=1
			tblstyle.add('FONTNAME', (0, 0), (-1, 0), 'Arial_Bold')
			tblstyle.add('TEXTCOLOR', (0, 0), (-1, 0), colors.white)
			tblstyle.add('BACKGROUND', (0, 0), (-1, 0), '#10b0c2')
			for row in cursor:
				vcolor='#e2e4e4'
				if numrow % 2 == 0:
					vcolor='#e8eaea'
				tblstyle.add('BACKGROUND', (0, numrow), (-1, numrow), vcolor)
				tblstyle.add('TEXTCOLOR', (0, numrow), (-1, numrow), colors.black)
				tblstyle.add('FONTNAME', (0, numrow), (-1, numrow), 'Arial')
				numrow += 1
				lista_aux = []
				for i in range(0, len(row)):
					if c_count in [1,2,3,4,7,8,11,12]:
						valor = row[i]
						if c_count in [1,2]:
							if i != 0:
								if valor < 0:
									valor = "(" + "{:,.2f}".format(abs(valor)) + ")"
								else:
									valor = "{:,.2f}".format(valor)
								fila_totales[i] += abs(row[i])
						if c_count in [3,4,7,8]:
							if i != 0:
								if valor < 0:
									valor = "(" + "{:,.2f}".format(abs(valor)) + ")"
								else:
									valor = "{:,.2f}".format(valor)
								#fila_totales[i] += abs(row[i])
						if c_count in [11,12]:
							if i in [1,2,3,6,7,8]:
								if valor < 0:
									valor = "(" + "{:,.2f}".format(abs(valor)) + ")"
								else:
									valor = "{:,.2f}".format(valor)
								fila_totales[i] += abs(row[i])
						lista_aux.append(valor)
					else:
						if i not in [0,1,17]:
							valor = row[i]
							if i in [10, 11, 14, 15]:
								if valor < 0:
									valor = "(" + "{:,.2f}".format(abs(valor)) + ")"
								else:
									valor = "{:,.2f}".format(valor)
							lista_aux.append(valor)
							if i == 10:
								fila_totales[8] += abs(row[i])
							if i == 11:
								fila_totales[9] += abs(row[i])
							if i == 13:
								fila_totales[11] += abs(row[i])
							if i == 14:
								fila_totales[12] += abs(row[i])
							if i == 15:
								fila_totales[13] += abs(row[i])
				data_cursor.append(lista_aux)
			if c_count in [3,7]:
				data_for_s_new_table= list(data_cursor)
			if c_count in [4,8]:
				data_for_t_new_table= list(data_cursor)
			if c_count in [1,2,5, 6,9,10,11,12]:
				for i in range(14):
					if i in getcolumnstosum(c_count):
						fila_totales[i] = "{:,.2f}".format(fila_totales[i])
				data_cursor.append(fila_totales)
				vcolor = '#e2e4e4'
				if numrow % 2 == 0:
					vcolor = '#e8eaea'
				tblstyle.add('BACKGROUND', (0, numrow), (-1, numrow), vcolor)
				tblstyle.add('TEXTCOLOR', (0, numrow), (-1, numrow), colors.black)
				tblstyle.add('FONTNAME', (0, numrow), (-1, numrow), 'Arial')
			if len(data_cursor) <= 2:
				empty_cursors.append(c_count)
			tbl = Table(data_cursor,hAlign='LEFT',repeatRows=1)
			tbl.setStyle(tblstyle)
			if c_count in [1, 2]:
				#desarrollo de unir dos tablas
				style_for_f_new_table.add('SPAN', (0, styl_new_count), (7, styl_new_count))
				style_for_f_new_table.add('FONTNAME', (0, styl_new_count), (-1, styl_new_count), 'Arial_Bold')
				style_for_f_new_table.add('TEXTCOLOR', (0, styl_new_count), (-1, styl_new_count), colors.black)
				data_for_f_new_table.append([getTableNamesComisiones(c_count),' ',' ',' ',' ',' ',' ',' '])
				styl_new_count += 1
				first=True
				for row in data_cursor:
					if first:
						first=False
						style_for_f_new_table.add('FONTNAME', (0, styl_new_count), (-1, styl_new_count), 'Arial_Bold')
						style_for_f_new_table.add('TEXTCOLOR', (0, styl_new_count), (-1, styl_new_count), colors.white)
						style_for_f_new_table.add('BACKGROUND', (0, styl_new_count), (-1, styl_new_count), '#10b0c2')
					else:
						tcolor = '#e2e4e4'
						if styl_new_count % 2 == 0:
							tcolor = '#e8eaea'
						style_for_f_new_table.add('BACKGROUND', (0, styl_new_count), (-1, styl_new_count), tcolor)
						style_for_f_new_table.add('TEXTCOLOR', (0, styl_new_count), (-1, styl_new_count), colors.black)
						style_for_f_new_table.add('FONTNAME', (0, styl_new_count), (-1, styl_new_count), 'Arial')
					styl_new_count += 1
					data_for_f_new_table.append(row)
			if c_count ==2:
				tbl = Table(data_for_f_new_table)
				tbl.setStyle(style_for_f_new_table)
				data = [[tbl_header, tbl]]
				shell_table = Table(data,hAlign='LEFT',colWidths=[350*mm,200*mm])
				flowables.append(shell_table)
				flowables.append(Table([("", " ", ""), ("", "", "")]))
			if c_count in [4,8]:
				#necesito el header nuevo
				lista_new_data=[]
				list_header=getheaderforcompressed(data_for_s_new_table,data_for_t_new_table)
				lista_new_data.append(list_header)
				aux_list_1=[]
				aux_list_1.append('MXP')
				aux_list_2 = []
				aux_list_2.append('USD')
				for i in range(len(list_header)):
					if i>0:
						entro = False
						for couple in data_for_s_new_table:
							if couple[0] == list_header[i]:
								entro=True
								aux_list_1.append(couple[1])
						if not entro:
							aux_list_1.append("0.00")
						entro = False
						for couple in data_for_t_new_table:
							if couple[0] == list_header[i]:
								entro = True
								aux_list_2.append(couple[1])
						if not entro:
							aux_list_2.append("0.00")
				lista_new_data.append(aux_list_1)
				lista_new_data.append(aux_list_2)
				tbl = Table(lista_new_data, hAlign='LEFT')
				tbl.setStyle(style_newx2)
				if c_count-1 not in empty_cursors or c_count not in empty_cursors:
					flowables.append(taux)
					flowables.append(Table([("", " ", "")]))
					flowables.append(tbl)
					flowables.append(Table([("", " ", ""), ("", "", "")]))
			if c_count in [5,6,9,10,11,12]:
				if c_count not in empty_cursors:
					flowables.append(taux)
					flowables.append(Table([("", " ", "")]))
					flowables.append(tbl)
					flowables.append(Table([("", " ", ""), ("", "", "")]))
			c_count += 1
		flowables.append(PageBreak())
		ramos,ramo_style,bonos,bono_style=get_tablas_referencia()
		tbl = Table(ramos, hAlign='LEFT')
		tbl.setStyle(ramo_style)
		flowables.append(tbl)
		app.logger.info("Construyendo reporte pdf.")
		PageNumCanvas.setReporte(PageNumCanvas,'COMISIONES')
		doc.build(flowables,canvasmaker=PageNumCanvas)
		return True,"",virtual_wb.getvalue(),"application/pdf", libro_nombre
	except Exception as ex:
		app.logger.error(ex)
		return False, 'Error en la generacion del reporte.', 0, 0,0


async def bonos_pdf(P_Clave,P_Feini,P_Fefin,P_COD,app):
	try:
		app.logger.info("Entrando a Estado de cuentas de Bonos PDF (" + P_COD + ")")
		pdfmetrics.registerFont(ttfonts.TTFont("Arial", "Arial.ttf"))
		pdfmetrics.registerFont(ttfonts.TTFont("Arial_Bold", "Arial_Bold.ttf"))
		con_est, con_mssg, connection = getconexion(app)
		if not con_est:
			return False, con_mssg, 0, 0
		app.logger.info("Conectado a la base de datos.")
		app.logger.info("Iniciando carga de cursores")

		query1 = getquery(P_COD, 'BONO', 0, P_Clave, P_Feini, P_Fefin)
		cur1 = connection.cursor()
		cur1.execute(query1)
		statement = connection.cursor()
		c_det = connection.cursor()
		statement.execute("begin  PKG_MUI_ESTADOS_DE_CUENTA_1.P_DET_PADRE( :Pb_CODIGO,:Pb_AGENTE, :Pb_FEINI, :Pb_FEFIN, :c_det  ); end;",
			c_det = c_det,  Pb_CODIGO = str(P_COD) ,Pb_AGENTE = str(P_Clave) ,Pb_FEINI = P_Feini, Pb_FEFIN = P_Fefin)

		app.logger.info("Cargados todos los cursores.")
		virtual_wb = BytesIO()
		doc = SimpleDocTemplate(virtual_wb,pagesize=landscape((475*mm, 600*mm)),topMargin=45*mm)
		flowables = []
		libro_nombre = P_Clave+"_" + P_Feini.replace("/", "")+"_"+ P_Fefin.replace("/", "")+'_bonos.pdf'
		app.logger.info("Leyendo cursor de cabecera")
		has_agent = False
		for row in cur1:
			has_agent = True
			lista_aux = []
			for i in range(0, len(row)):
				lista_aux.append(row[i])
		cur1.close
		if not has_agent:
			app.logger.error("La cabecera volvio vacia.")
			return False, 'Identificador no encontrado', 0, 0,0
		header_all = getheaderpdf(P_COD,lista_aux,'BONOS')
		grid = [('FONTNAME', (0, 0), (-1, -1), 'Arial_Bold'), ('ALIGN', (0, 0), (-1, -1), 'LEFT')]
		grid2 = [('ALIGN', (0, 0), (-1, -1), 'LEFT'),
				 ('FONTNAME', (1, 0), (1, -1), 'Arial_Bold'),
				 ('FONTNAME', (4, 0), (4, -1), 'Arial_Bold')]
		tbl = Table(header_all)
		tbl.setStyle(grid2)
		flowables.append(tbl)
		j = 0
		lista = getHeadColumnsBonos("PDF",P_COD)
		data_body = []
		lista_aux = []
		has_data=False
		for item in lista:
			lista_aux.append(item)
		data_body.append(lista_aux)
		app.logger.info("Leyendo cursor -> (1)")
		numrow = 1
		tblstyle = TableStyle(
			[('LINEABOVE', (0, 0), (-1, -1), 0.25, colors.white), ('ALIGN', (0, 0), (-1, -1), 'CENTER')])
		tblstyle.add('FONTNAME', (0, 0), (-1, 0), 'Arial_Bold')
		tblstyle.add('TEXTCOLOR', (0, 0), (-1, 0), colors.white)
		tblstyle.add('BACKGROUND', (0, 0), (-1, 0), '#10b0c2')
		for row in c_det:
			vcolor = '#e2e4e4'
			if numrow % 2 == 0:
				vcolor = '#e8eaea'
			tblstyle.add('BACKGROUND', (0, numrow), (-1, numrow), vcolor)
			tblstyle.add('TEXTCOLOR', (0, numrow), (-1, numrow), colors.black)
			tblstyle.add('FONTNAME', (0, numrow), (-1, numrow), 'Arial')
			numrow += 1
			has_data=True
			lista_aux = []
			for i in range(0, len(row)):
				if i < 21 and i not in [3,8,9,19]:
					valor=row[i]
					if i == 10 and int(row[i]) == 1:
						valor=' '
					if i == 20:
						if row[i] is not None:
							auxnewdate = datetime.date.strftime(row[i], "%d/%m/%Y")
							valor = auxnewdate
						else:
							valor=' '
					if i == 2:
						valor=getTipoSubBono(int(row[i]))
					lista_aux.append(valor)
				if i==9:
					if row[21]== 'NO':
						lista_aux.append("Si")
					if row[21]== 'SI':
						lista_aux.append("No")
			data_body.append(lista_aux)
		c_det.close()
		if not has_data:
			app.logger.error("La tabla de detalle volvio vacia.")
			return False, 'Error generando el reporte.', 0, 0,0
		tbl = Table(data_body, repeatRows=1)
		tbl.setStyle(tblstyle)
		flowables.append(tbl)
		flowables.append(Table([("", " ", ""), ("", "", "")]))
		app.logger.info("Agregando tabla de soporte.")
		flowables.append(PageBreak())
		ramos, ramo_style, bonos, bono_style = get_tablas_referencia()
		tbl = Table(ramos, hAlign='LEFT')
		tbl.setStyle(ramo_style)
		flowables.append(tbl)
		flowables.append(Table([("", " ", ""), ("", "", "")]))
		tbl = Table(bonos, hAlign='LEFT')
		tbl.setStyle(bono_style)
		flowables.append(tbl)
		PageNumCanvas.setReporte(PageNumCanvas,'BONOS')
		app.logger.info("Construyendo reporte pdf.")
		doc.build(flowables,canvasmaker=PageNumCanvas)
		return True,"",virtual_wb.getvalue(),"application/pdf", libro_nombre
	except Exception as ex:
		return False, 'Error generando el reporte.', 0, 0,0


async def bonos_xlx(P_Clave,P_Feini,P_Fefin,P_COD,app):
	try:
		app.logger.info("Entrando a Estado de cuentas de Bonos XLSX (" + P_COD + ")")
		con_est, con_mssg, connection = getconexion(app)
		if not con_est:
			return False, con_mssg, 0, 0
		app.logger.info("Conectado a la base de datos.")
		app.logger.info("Iniciando carga de cursores")

		query1 = getquery(P_COD, 'BONO', 0, P_Clave, P_Feini, P_Fefin)
		cur1 = connection.cursor()
		cur1.execute(query1)
		statement = connection.cursor()
		c_det = connection.cursor()
		query1 = "begin  PKG_MUI_ESTADOS_DE_CUENTA_1.P_DET_PADRE ( :Pb_CODIGO,:Pb_AGENTE, :Pb_FEINI, :Pb_FEFIN, :c_det  ); end;"
		statement.execute(query1,
			c_det=c_det, Pb_CODIGO=str(P_COD), Pb_AGENTE=str(P_Clave), Pb_FEINI=P_Feini, Pb_FEFIN=P_Fefin)

		app.logger.info("Cargados todos los cursores.")
		plantilla = "plantilla_agentes.xlsx"
		if P_COD == 'P':
			plantilla = "plantilla_promotor.xlsx"
		wb = opyxl.load_workbook(plantilla)
		app.logger.info("Cargado archivo de plantilla xlsx.")
		ws = wb.worksheets[0]

		libro_nombre = P_Clave+"_" + P_Feini.replace("/", "")+"_"+ P_Fefin.replace("/", "")+'_bonos.xlsx'
		app.logger.info("Leyendo cursor de cabecera")
		for row in cur1:
			has_agent = True
			for i in range(0, len(row) - 4):
				ws.cell(row=4 + i, column=9).value = row[i]
			for i in range(len(row) - 4, len(row)):
				ws.cell(row=i, column=4).value = row[i]
		cur1.close()
		if not has_agent:
			app.logger.error("La cabecera volvio vacia.")
			return False, 'Identificador no encontrado.', 0, 0,0
		j = 0
		greyFill = PatternFill(fill_type='solid', start_color='10b0c2', end_color='10b0c2')
		lista = getHeadColumnsBonos("excel",P_COD)
		alphabet_string = string.ascii_uppercase
		alphabet_list = list(alphabet_string)
		for item in lista:
			ws.cell(row=13, column=j + 1).value = item
			ws.cell(row=13, column=j + 1).fill = greyFill
			ws.cell(row=13, column=j + 1).border = Border(left=Side(style='thin'), right=Side(style='thin'),
														 top=Side(style='thin'), bottom=Side(style='thin'))
			ws.cell(row=13, column=j + 1).font = Font(name='Arial', size=9, bold=True,color='ffffff')
			ws.cell(row=13, column=j + 1).alignment = Alignment(horizontal="center", vertical="center")
			columna = alphabet_list[ws.cell(row=13, column=j + 1).column - 1]
			multiplicador = 2
			lista_columnas_esp = ['B', 'G']
			ancho = len(item)
			if len(item) <= 6:
				ancho *= 2
			if len(item) > 6 or columna == 'A':
				ancho *= 1.1
			if columna in lista_columnas_esp:
				ancho = 25
			ws.column_dimensions[columna].width = ancho
			j += 1
		j = 0
		k = 0
		has_data = False
		app.logger.info("intentando fetch cursor -> (1)")
		data = c_det.fetchall()
		app.logger.info("logrado fetch cursor -> (1)")
		app.logger.info("Leyendo cursor -> (1)")
		for row in data:
			has_data = True
			lista_razones = []
			for i in range(0, len(row)):

				if i < 21:
					valor = row[i]
					aux = 0
					if i >= 9:
						aux = 1
					if i == 2:
						valor = getTipoSubBono(int(row[i]))
					if i == 10 and int(row[i]) == 1:
						valor = ' '
					if i == 20:
						if row[i] is not None:
							auxnewdate = datetime.date.strftime(row[i], "%d/%m/%Y")
							valor = auxnewdate
						else:
							valor= ' '
					ws.cell(row=14 + j, column=i + 1 + aux).value = valor
					ws.cell(row=14+j, column=i + 1+aux).border = Border(left=Side(style='thin'), right=Side(style='thin'),
																 top=Side(style='thin'), bottom=Side(style='thin'))
					ws.cell(row=14 + j, column=i + 1 + aux).alignment = Alignment(horizontal="center",
																				  vertical="center")
					if len(str(valor)) > 17:
						ws.cell(row=14 + j, column=i + 1 + aux).font = Font(name='Arial', size=8)
						if len(str(valor)) > 25:
							ws.cell(row=14 + j, column=i + 1 + aux).font = Font(name='Arial', size=7)
				else:
					if i == 21:
						if row[i] == 'NO':
							ws.cell(row=14 + j, column=10).value = 'Sí'
						else:
							ws.cell(row=14 + j, column=10).value = 'No'
						ws.cell(row=14 + j, column=10).border = Border(left=Side(style='thin'),
																	   right=Side(style='thin'),
																	   top=Side(style='thin'),
																	   bottom=Side(style='thin'))
						ws.cell(row=14 + j, column=10).alignment = Alignment(horizontal="center",
																					  vertical="center")
					if i > 21 and row[21] == 'SI':
						if row[i] is not None and len(row[i]) > 0:
							lista_razones.append(row[i])
			if row[21] == 'SI' and len(lista_razones) > 0:
				ws.cell(row=14 + j, column=11).value = ", ".join(lista_razones)
			if row[21] == 'NO':
				ws.cell(row=14 + j, column=11).value = ' '

			j += 1
		c_det.close()
		if not has_data:
			app.logger.error("La tabla de detalle volvio vacia.")
			return False, 'Error en la generacion del reporte.', 0, 0,0
		j += 16
		ramos, ramo_style, bonos, bono_style = get_tablas_referencia()
		t_apoyo=[ramos,bonos]
		for tabla in t_apoyo:
			c_linea=0
			for linea in tabla:
				c_columna=0
				for columna in linea:
					if c_linea in [0, 1]:
						if c_linea == 0 and t_apoyo.index(tabla) == 1 and c_columna == 0:
							top_1 = ws.cell(row=j, column=1)
							top_2 = ws.cell(row=j, column=3)
							top_3 = ws.cell(row=j, column=5)
							top_4 = ws.cell(row=j, column=7)
							ws.merge_cells(start_row=j, start_column=1, end_row=j, end_column=2)
							ws.merge_cells(start_row=j, start_column=3, end_row=j, end_column=4)
							ws.merge_cells(start_row=j, start_column=5, end_row=j, end_column=6)
							ws.merge_cells(start_row=j, start_column=7, end_row=j, end_column=8)
							top_1.value = linea[0]
							top_1.fill = greyFill
							top_1.font = Font(name='Arial', size=9, bold=True,
											  color='ffffff')
							top_1.alignment = Alignment(horizontal="center",
														vertical="center")
							top_2.value = linea[2]
							top_2.fill = greyFill
							top_2.font = Font(name='Arial', size=9, bold=True,
											  color='ffffff')
							top_2.alignment = Alignment(horizontal="center",
														vertical="center")
							top_3.value = linea[4]
							top_3.fill = greyFill
							top_3.font = Font(name='Arial', size=9, bold=True,
											  color='ffffff')
							top_3.alignment = Alignment(horizontal="center",
														vertical="center")
							top_4.value = linea[6]
							top_4.fill = greyFill
							top_4.font = Font(name='Arial', size=9, bold=True,
											  color='ffffff')
							top_4.alignment = Alignment(horizontal="center",
														vertical="center")
						if c_linea == 0 and t_apoyo.index(tabla) == 0 and c_columna == 0:
							top_1 = ws.cell(row=j, column=c_columna + 1)
							ws.merge_cells(start_row=j, start_column=1, end_row=j, end_column=8)
							top_1.value = columna
							top_1.fill = greyFill
							top_1.font = Font(name='Arial', size=9, bold=True,
											  color='ffffff')
							top_1.alignment = Alignment(horizontal="center",
																					   vertical="center")
						if c_linea==1:
							ws.cell(row=j, column=c_columna + 1).value = columna
							ws.cell(row=j, column=c_columna + 1).fill = greyFill
							ws.cell(row=j, column=c_columna + 1).font = Font(name='Arial', size=9, bold=True,
																			 color='ffffff')
							ws.cell(row=j, column=c_columna + 1).alignment = Alignment(horizontal="center",
																					   vertical="center")
					else:
						ws.cell(row=j, column=c_columna+1).value = columna
						ws.cell(row=j, column=c_columna + 1).alignment = Alignment(horizontal="center",
																				   vertical="center")

					c_columna+=1
				c_linea+=1
				j += 1
			j += 1


		virtual_wb = BytesIO()
		app.logger.info("Construyendo reporte xlsx.")
		wb.save(virtual_wb)
		return True,"",virtual_wb.getvalue(),wb.mime_type,libro_nombre
	except Exception as ex:
		app.logger.error(ex)
		return False, 'Error en la generacion del reporte.', 0, 0,0


async def udi_pdf(P_Clave,P_Feini,P_Fefin,P_COD,app):
	try:
		pdfmetrics.registerFont(ttfonts.TTFont("Arial", "Arial.ttf"))
		pdfmetrics.registerFont(ttfonts.TTFont("Arial_Bold", "Arial_Bold.ttf"))
		app.logger.info("Entrando a Estado de cuentas PDF (" + P_COD + ")")
		con_est, con_mssg, pool = await get_oracle_pool(app)
		if not con_est:
			return False, con_mssg, 0, 0, 0
		app.logger.info("Conectado a la base de datos.")
		has_agent = False
		app.logger.info("Iniciando carga de cursores")
		cursors = await tasks(app, 8, P_COD, ' ', P_Clave, P_Feini, P_Fefin, pool)
		if 'Excepcion' in cursors:
			return False, 'Hubo un error obteniendo la data', 0, 0, 0
		app.logger.info("Cargados todos los cursores.")
		libro_nombre = P_Clave + "_" + P_Feini.replace("/", "") + "_" + P_Fefin.replace("/", "") + '_udi.pdf'

		virtual_wb = BytesIO()
		doc = SimpleDocTemplate(virtual_wb, pagesize=landscape((475 * mm, 600 * mm)), topMargin=45 * mm)
		flowables = []
		app.logger.info("Leyendo cursor de cabecera")
		for row in cursors[0]:
			has_agent = True
			lista_aux = []
			for i in range(0, len(row)):
				lista_aux.append(row[i])
		if not has_agent:
			app.logger.error("La cabecera volvio vacia.")
			return False, 'Identificador no encontrado', 0, 0, 0
		header_all = getheaderpdf(P_COD, lista_aux, 'UDI')
		grid = [('FONTNAME', (0, 0), (-1, -1), 'Arial_Bold'), ('ALIGN', (0, 0), (-1, -1), 'LEFT')]
		grid2 = [('ALIGN', (0, 0), (-1, -1), 'LEFT'),
				 ('FONTNAME', (1, 0), (1, -1), 'Arial_Bold'),
				 ('FONTNAME', (4, 0), (4, -1), 'Arial_Bold')]
		tbl_header = Table(header_all, hAlign='LEFT')
		tbl_header.setStyle(grid2)
		flowables.append(tbl_header)
		del cursors[0]
		c_count = 1
		empty_cursors = []
		data_for_s_new_table = []
		data_for_t_new_table = []

		style_newx2 = TableStyle(
			[('LINEABOVE', (0, 0), (-1, -1), 0.25, colors.white), ('ALIGN', (0, 0), (-1, -1), 'CENTER')])
		style_newx2.add('FONTNAME', (0, 0), (-1, 0), 'Arial_Bold')
		style_newx2.add('TEXTCOLOR', (0, 0), (-1, 0), colors.white)
		style_newx2.add('BACKGROUND', (0, 0), (-1, 0), '#10b0c2')
		style_newx2.add('BACKGROUND', (0, 1), (-1, 1), '#e8eaea')
		style_newx2.add('TEXTCOLOR', (0, 1), (-1, 1), colors.black)
		style_newx2.add('FONTNAME', (0, 1), (-1, 1), 'Arial')
		style_newx2.add('BACKGROUND', (0, 2), (-1, 2), '#e2e4e4')
		style_newx2.add('TEXTCOLOR', (0, 2), (-1, 2), colors.black)
		style_newx2.add('FONTNAME', (0, 2), (-1, 2), 'Arial')
		for cursor in cursors:
			tblstyle = TableStyle(
				[('LINEABOVE', (0, 0), (-1, -1), 0.25, colors.white), ('ALIGN', (0, 0), (-1, -1), 'CENTER')])
			app.logger.info(f"Leyendo cursor -> ({c_count})")
			lista = getHeadColumnsUDI("pdf", c_count)
			data_cursor = []
			theader = []
			theader.append([getTableNamesUDI(c_count)])
			taux = Table(theader, hAlign='LEFT')
			taux.setStyle(grid)
			data_cursor.append(lista)
			fila_totales = []
			if c_count in [4,6]:
				fila_totales = ["", "", "", "", "", "", "", "TOTAL", 0, 0, ' ', 0,0,'','','']
			if c_count in [1, 2]:
				fila_totales = ["TOTAL", 0, 0, 0, 0, 0, 0, 0]
			if c_count in [3, 5]:
				fila_totales = ["TOTAL", 0]
			if c_count in [7]:
				fila_totales = ["TOTAL", 0, 0, 0]
			numrow = 1
			tblstyle.add('FONTNAME', (0, 0), (-1, 0), 'Arial_Bold')
			tblstyle.add('TEXTCOLOR', (0, 0), (-1, 0), colors.white)
			tblstyle.add('BACKGROUND', (0, 0), (-1, 0), '#10b0c2')
			for row in cursor:
				vcolor = '#e2e4e4'
				if numrow % 2 == 0:
					vcolor = '#e8eaea'
				tblstyle.add('BACKGROUND', (0, numrow), (-1, numrow), vcolor)
				tblstyle.add('TEXTCOLOR', (0, numrow), (-1, numrow), colors.black)
				tblstyle.add('FONTNAME', (0, numrow), (-1, numrow), 'Arial')
				numrow += 1
				lista_aux = []
				for i in range(0, len(row)):
					if c_count in [1, 2, 3, 5, 7]:
						valor = row[i]
						if c_count in [1, 2]:
							if i != 0:
								if valor < 0:
									valor = "(" + "{:,.2f}".format(abs(valor)) + ")"
								else:
									valor = "{:,.2f}".format(valor)
								fila_totales[i] += abs(row[i])
						if c_count in [3,5]:
							if i != 0:
								if valor < 0:
									valor = "(" + "{:,.2f}".format(abs(valor)) + ")"
								else:
									valor = "{:,.2f}".format(valor)
						# fila_totales[i] += abs(row[i])
						if c_count in [7]:
							if i != 0:
								if valor < 0:
									valor = "(" + "{:,.2f}".format(abs(valor)) + ")"
								else:
									valor = "{:,.2f}".format(valor)
								fila_totales[i] += abs(row[i])
						lista_aux.append(valor)
					else:
						valor = row[i]
						if i == 15:
							if row[i] is not None:
								auxnewdate = datetime.date.strftime(row[i], "%d/%m/%Y")
								valor = auxnewdate
							else:
								valor = ' '
						if i in [8,9, 11, 12]:
							if valor < 0:
								valor = "(" + "{:,.2f}".format(abs(valor)) + ")"
							else:
								valor = "{:,.2f}".format(valor)
							fila_totales[i] += abs(row[i])
						lista_aux.append(valor)
				data_cursor.append(lista_aux)
			if c_count in [1]:
				data_for_s_new_table = list(data_cursor)
			if c_count in [2]:
				data_for_t_new_table = list(data_cursor)
			if c_count in [1, 2, 4, 6, 7]:
				for i in range(17):
					if i in getcolumnstosum_udi(c_count):
						fila_totales[i] = "{:,.2f}".format(fila_totales[i])
				data_cursor.append(fila_totales)
				vcolor = '#e2e4e4'
				if numrow % 2 == 0:
					vcolor = '#e8eaea'
				tblstyle.add('BACKGROUND', (0, numrow), (-1, numrow), vcolor)
				tblstyle.add('TEXTCOLOR', (0, numrow), (-1, numrow), colors.black)
				tblstyle.add('FONTNAME', (0, numrow), (-1, numrow), 'Arial')
			if len(data_cursor) <= 2:
				empty_cursors.append(c_count)
			tbl = Table(data_cursor, hAlign='LEFT', repeatRows=1)
			tbl.setStyle(tblstyle)

			if c_count in [2]:
				# necesito el header nuevo
				lista_new_data = []
				lista_new_data.append(getHeadColumnsUDI("pdf", c_count))
				if len(data_for_s_new_table)>1:
					lista_new_data.append(data_for_s_new_table[1])
				else:
					lista_new_data.append(['DAÑO','0','0','0','0','0','0','0'])
				if len(data_for_t_new_table)>1:
					lista_new_data.append(data_for_t_new_table[1])
				else:
					lista_new_data.append(['VIDA','0','0','0','0','0','0','0'])

				tbl = Table(lista_new_data, hAlign='LEFT')
				tbl.setStyle(style_newx2)
				if c_count - 1 not in empty_cursors or c_count not in empty_cursors:
					flowables.append(taux)
					flowables.append(Table([("", " ", "")]))
					flowables.append(tbl)
					flowables.append(Table([("", " ", ""), ("", "", "")]))
			if c_count in [3,4,5,6,7]:
				if c_count not in empty_cursors:
					flowables.append(taux)
					flowables.append(Table([("", " ", "")]))
					flowables.append(tbl)
					flowables.append(Table([("", " ", ""), ("", "", "")]))
			c_count += 1
		flowables.append(PageBreak())
		ramos, ramo_style, bonos, bono_style = get_tablas_referencia()
		tbl = Table(ramos, hAlign='LEFT')
		tbl.setStyle(ramo_style)
		flowables.append(tbl)
		app.logger.info("Construyendo reporte pdf.")
		PageNumCanvas.setReporte(PageNumCanvas, 'COMISIONES')
		doc.build(flowables, canvasmaker=PageNumCanvas)
		return True, "", virtual_wb.getvalue(), "application/pdf", libro_nombre
	except Exception as ex:
		app.logger.error(ex)
		return False, 'Error en la generacion del reporte.', 0, 0, 0


async def udi_xlsx(P_Clave,P_Feini,P_Fefin,P_COD,app):
	try:
		app.logger.info("Entrando a Estado de cuentas XLSX ("+P_COD+")")
		con_est, con_mssg, pool = await get_oracle_pool(app)
		if not con_est:
			return False, con_mssg, 0, 0, 0
		app.logger.info("Conectado a la base de datos.")
		has_agent = False
		app.logger.info("Iniciando carga de cursores")
		cursors = await tasks(app, 8, P_COD, ' ', P_Clave, P_Feini, P_Fefin, pool)
		if 'Excepcion' in cursors:
			return False, 'Hubo un error obteniendo la data', 0, 0, 0
		app.logger.info("Cargados todos los cursores.")
		libro_nombre =P_Clave+"_" + P_Feini.replace("/", "")+"_"+ P_Fefin.replace("/", "")+'_udi.xlsx'
		plantilla = "plantilla_udi.xlsx"
		wb = opyxl.load_workbook(plantilla)
		app.logger.info("Cargado archivo de plantilla xlsx.")
		ws = wb.worksheets[0]
		ws.cell(row=1, column=6).value = "ESTADOS DE CUENTA DE BONOS"
		ws.title = "Estado de Cuenta de Bonos"
		has_agent = False
		app.logger.info("Leyendo cursor de cabecera")
		for row in cursors[0]:
			has_agent = True
			for i in range(0, len(row) - 4):
				ws.cell(row=4 + i, column=9).value = row[i]
			for i in range(len(row) - 4, len(row)):
				ws.cell(row=i, column=4).value = row[i]
		if not has_agent:
			app.logger.error("La cabecera volvio vacia.")
			return False,'Identificador no encontrado',0,0,0
		del cursors[0]
		f = 14  # principal gestor de filas del archivo
		greyFill = PatternFill(fill_type='solid', start_color='10b0c2', end_color='10b0c2')
		# NUEVO BLOQUE SECUENCIAL
		c_count = 1
		for cursor in cursors:

			app.logger.info(f"Leyendo cursor -> ({c_count})")
			fila_totales = [0, 0, 0, 0]
			lista = getHeadColumnsUDI("excel", c_count)
			alphabet_string = string.ascii_uppercase
			alphabet_list = list(alphabet_string)


			if len(cursor) > 0:
				ws.cell(row=f, column=1).value = getTableNamesComisiones(c_count)
				ws.cell(row=f, column=1).font = Font(name='Arial', size=9, bold=True)
				f += 1
				j = 0
				for item in lista:
					ws.cell(row=f, column=j + 1).value = item
					ws.cell(row=f, column=j + 1).border = Border(left=Side(style='thin'), right=Side(style='thin'),
																 top=Side(style='thin'), bottom=Side(style='thin'))
					ws.cell(row=f, column=j + 1).fill = greyFill
					ws.cell(row=f, column=j + 1).font = Font(name='Arial', size=9, bold=True,color='ffffff')
					ws.cell(row=f, column=j + 1).alignment = Alignment(horizontal="center", vertical="center")
					columna = alphabet_list[ws.cell(row=f, column=j + 1).column - 1]
					multiplicador = 2
					lista_columnas_esp = ['A', 'B', 'C', 'H']
					ancho = len(item)
					if len(item) <= 6:
						ancho *= 2
					if len(item) > 6 or columna == 'A':
						ancho *= 1.1
					if columna in lista_columnas_esp:
						ancho = 25
					if c_count == 4:
						ws.column_dimensions[columna].width = ancho
					j += 1
			j = 0
			k = 0
			if len(cursor) > 0:
				f += 1
			has_data= False
			for row in cursor:
				for i in range(0, len(row)):
					if c_count not in [4,6]:
						valor = row[i]
						if c_count in [1, 2]:
							if i != 0:
								valor = "{:,.2f}".format(valor)
						if c_count in [3, 5, 7]:
							if i != 0:
								valor = "{:,.2f}".format(valor)
						ws.cell(row=f, column=i + 1).value = valor
						ws.cell(row=f, column=i + 1).border= Border(left=Side(style='thin'), right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
						ws.cell(row=f, column=i + 1).alignment = Alignment(horizontal="center", vertical="center")
						if len(str(valor)) > 17:
							ws.cell(row=f, column=i + 1).font = Font(name='Arial', size=8)
							if len(str(valor)) > 25:
								ws.cell(row=f, column=i + 1).font = Font(name='Arial', size=7)
					else:
						valor = row[i]
						if i in [8, 9, 11, 12]:
							valor = "{:,.2f}".format(valor)
						if i == 8:
							fila_totales[0] += abs(row[i])
						if i == 9:
							fila_totales[1] += abs(row[i])
						if i == 11:
							fila_totales[2] += abs(row[i])
						if i == 12:
							fila_totales[3] += abs(row[i])
						has_data = True
						ws.cell(row=f, column=i + 1).value = valor
						ws.cell(row=f, column=i + 1).border = Border(left=Side(style='thin'), right=Side(style='thin'),
																	 top=Side(style='thin'), bottom=Side(style='thin'))
						ws.cell(row=f, column=i + 1).alignment = Alignment(horizontal="center", vertical="center")
						if len(str(valor)) > 17:
							ws.cell(row=f, column=i + 1).font = Font(name='Arial', size=8)
							if len(str(valor)) > 25:
								ws.cell(row=f, column=i + 1).font = Font(name='Arial', size=7)
				f += 1
			#cursor.close()
			if c_count in [4,6] and has_data:
				fila_totales[0] = "{:,.2f}".format(fila_totales[0])
				fila_totales[1] = "{:,.2f}".format(fila_totales[1])
				fila_totales[2] = "{:,.2f}".format(fila_totales[2])
				fila_totales[3] = "{:,.2f}".format(fila_totales[3])
				ws.cell(row=f, column=8).value = "TOTAL"
				ws.cell(row=f, column=8).alignment = Alignment(horizontal="center", vertical="center")
				ws.cell(row=f, column=9).value = fila_totales[0]
				ws.cell(row=f, column=9).alignment = Alignment(horizontal="center", vertical="center")
				ws.cell(row=f, column=10).value = fila_totales[1]
				ws.cell(row=f, column=10).alignment = Alignment(horizontal="center", vertical="center")
				ws.cell(row=f, column=12).value = fila_totales[2]
				ws.cell(row=f, column=12).alignment = Alignment(horizontal="center", vertical="center")
				ws.cell(row=f, column=13).value = fila_totales[3]
				ws.cell(row=f, column=3).alignment = Alignment(horizontal="center", vertical="center")
				f += 1
			if len(cursor) > 0:
				f += 1
			c_count += 1
		# fin de bloque
		f += 2
		ramos, ramo_style, bonos, bono_style = get_tablas_referencia()
		t_apoyo = [ramos]
		for tabla in t_apoyo:
			c_linea = 0
			for linea in tabla:
				c_columna = 0
				for columna in linea:
					if c_linea in [0, 1]:
						if c_linea==0 and c_columna==0:
							top_1 = ws.cell(row=f, column=1)
							ws.merge_cells(start_row=f, start_column=1, end_row=f, end_column=8)
							top_1.value = columna
							top_1.fill = greyFill
							top_1.font = Font(name='Arial', size=9, bold=True,
																			 color='ffffff')
							top_1.alignment = Alignment(horizontal="center", vertical="center")
						if c_linea==1:
							ws.cell(row=f, column=c_columna + 1).value = columna
							ws.cell(row=f, column=c_columna + 1).fill = greyFill
							ws.cell(row=f, column=c_columna + 1).font = Font(name='Arial', size=9, bold=True,
																			 color='ffffff')
							ws.cell(row=f, column=c_columna + 1).alignment = Alignment(horizontal="center", vertical="center")
					else:
						ws.cell(row=f, column=c_columna + 1).value = columna
						ws.cell(row=f, column=c_columna + 1).alignment = Alignment(horizontal="center",
																				   vertical="center")
					c_columna += 1
				c_linea += 1
				f += 1
			f += 1
		virtual_wb = BytesIO()
		app.logger.info("Construyendo reporte xlsx.")
		wb.save(virtual_wb)
		return True,"",virtual_wb.getvalue(), wb.mime_type,libro_nombre
	except Exception as ex:
		app.logger.error(ex)
		return False, 'Error generando el reporte', 0, 0,0


class PageNumCanvas(canvas.Canvas):
	reporte = ""
	def setReporte(self,val):
		self.reporte = val
	def __init__(self, *args, **kwargs):
		"""Constructor"""
		canvas.Canvas.__init__(self, *args, **kwargs)
		self.pages = []

	def showPage(self):
		self.pages.append(dict(self.__dict__))
		self._startPage()

	def save(self):
		page_count = len(self.pages)

		for page in self.pages:
			self.__dict__.update(page)
			self.draw_page_number(page_count)
			canvas.Canvas.showPage(self)

		canvas.Canvas.save(self)

	def draw_page_number(self, page_count):
		tahora= datetime.datetime.now(timezone('UTC'))
		now_mexico = tahora.astimezone(timezone('America/Mexico_City'))
		page = "Pagina %s de %s" % (self._pageNumber, page_count)
		self.setFont("Arial", 11)
		self.drawRightString((600 - 25) * mm, (475 - 23) * mm, "Blvd. Adolfo López Mateos No. 2448 / Col. Alta Vista Deleg. Alvaro Obregón / C. P. 01060 México, D. F.")
		self.drawRightString((600 - 25) * mm, (475 - 31) * mm, "STel. 57-23-79-99, 01-800-723-79-00")
		self.setFont("Arial_Bold", 11)
		self.drawRightString((600 - 25) * mm, (475 - 15) * mm, "Seguros SURA, S. A. de C. V.")
		self.setFont("Arial", 11)
		self.setFillColor(colors.gray)
		self.drawRightString((600 - 25) * mm, (475 - 455) * mm, page)
		self.drawRightString((600 - 515) * mm, (475-455) * mm, f"Generado el {now_mexico.strftime('%m/%d/%Y, %H:%M:%S')}")
		im = Image('logo_sura.png')
		im.drawOn(self,30*mm,(475-40)*mm)
		self.setFillColor('#10b0c2')
		self.setFont("Arial_Bold", 25)
		self.drawString((600 - 480) * mm, (475 - 18) * mm, "ESTADO DE CUENTA")
		self.setFillColor('#2e3a8f')
		self.drawString((600 - 480) * mm, (475 - 29) * mm, "DE "+self.reporte)
