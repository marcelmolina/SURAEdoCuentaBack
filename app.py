# -*- coding: utf-8 -*-
import os
import string
import time
import cx_Oracle
import openpyxl as opyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from flask import Flask, jsonify, request, Response, make_response
from io import BytesIO
import datetime
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
import os
from dotenv import load_dotenv
from flask_cors import CORS

load_dotenv()
SERVER_PORT = os.getenv('SERVER_PORT')
DB_HOST = os.getenv('DB_HOST')
DB_PORT = os.getenv('DB_PORT')
DB_SERVICE = os.getenv('DB_SERVICE')
DB_USER = os.getenv('DB_USER')
DB_PWD = os.getenv('DB_PWD')
DB_SCHEMA = os.getenv('DB_SCHEMA')


app = Flask(__name__)
CORS(app)
app.config['JSON_SORT_KEYS'] = False
context_path = "/api/estados-cuenta/"

# Testing Route
@app.route(context_path+'/ping', methods=['GET'])
def ping():

    return jsonify({'response': 'pong!'})


@app.route(context_path+'/periodo', methods=['GET'])
def periodo():
	P_MES = request.args['mes']
	P_ANIO = request.args['anio']
	P_CLAVE = request.args['clave']
	host = DB_HOST
	port = DB_PORT
	service_name = DB_SERVICE
	user = DB_USER
	password = DB_PWD
	schema = DB_SCHEMA
	sid = cx_Oracle.makedsn(host, port, service_name=service_name)
	# Declaracion de cursores a utilizar
	try:
		connection = cx_Oracle.connect(f"{user}/{password}@{host}:{port}/{service_name}")
		try:
			statement = connection.cursor()
			cur1 = connection.cursor()
			cur2 = connection.cursor()
			v_desde = cur1.var(cx_Oracle.Date)
			v_hasta = cur2.var(cx_Oracle.Date)
			statement.execute(
				"begin " + schema + ".PKG_MUI_ESTADOS_DE_CUENTA_1.P_DESDEHASTA( :Pb_CLAVE, :Pb_MES, :Pb_ANIO, :desde,:hasta); end;",
				desde=v_desde,hasta=v_hasta, Pb_CLAVE=str(P_CLAVE), Pb_MES=P_MES, Pb_ANIO=P_ANIO)
			if not v_desde:
				return make_response(jsonify(succes=False, message="Error en la busqueda de fechas."), 400)
			P_Feini = str(v_desde.values[0])
			P_Fefin = str(v_hasta.values[0])
			return make_response(jsonify(succes=True, desde=P_Feini, hasta=P_Fefin), 200)
		except Exception as ex:
			app.logger.error(ex)
			statement.close()
			return make_response(jsonify(succes=False, message="Error en la busqueda de fechas."), 400)
	except Exception as ex:
		app.logger.error(ex)
		return make_response(jsonify(succes=False, message="Error en la busqueda de fechas."), 400)


@app.route(context_path + '/agentes/bonos/excel', methods=['GET'])
def bono_agente_xlsx():
	P_Clave = request.args['codigo']
	P_Feini = request.args['desde']
	P_Fefin = request.args['hasta']
	P_Feini = datetime.datetime.strptime(P_Feini, "%Y-%m-%d").strftime("%d/%m/%Y")
	P_Fefin = datetime.datetime.strptime(P_Fefin, "%Y-%m-%d").strftime("%d/%m/%Y")
	try:
		file, filemime, filename = bonos_xlx(P_Clave, P_Feini, P_Fefin, 'A')
		return Response(file, mimetype=filemime,
						headers={"Content-Disposition": "attachment;filename=" + filename})
	except Exception as ex:
		app.logger.error(ex)
		return make_response(jsonify(succes=False, message=ex), 400)


@app.route(context_path + '/agentes/bonos/pdf', methods=['GET'])
def bono_agente_pdf():
	P_Clave = request.args['codigo']
	P_Feini = request.args['desde']
	P_Fefin = request.args['hasta']
	P_Feini = datetime.datetime.strptime(P_Feini, "%Y-%m-%d").strftime("%d/%m/%Y")
	P_Fefin = datetime.datetime.strptime(P_Fefin, "%Y-%m-%d").strftime("%d/%m/%Y")
	try:
		file, filename = bonos_pdf(P_Clave, P_Feini, P_Fefin, 'A')
		return Response(file, mimetype="application/pdf",
						headers={"Content-Disposition": "attachment;filename=" + filename})
	except Exception as ex:
		app.logger.error(ex)
		return make_response(jsonify(succes=False, message=ex), 400)


@app.route(context_path + '/agentes/comisiones/pdf', methods=['GET'])
def comisiones_agente_pdf():
	P_Clave = request.args['codigo']
	P_Feini = request.args['desde']
	P_Fefin = request.args['hasta']
	P_Feini = datetime.datetime.strptime(P_Feini, "%Y-%m-%d").strftime("%d/%m/%Y")
	P_Fefin = datetime.datetime.strptime(P_Fefin, "%Y-%m-%d").strftime("%d/%m/%Y")
	try:
		file, filename = comisiones_pdf(P_Clave, P_Feini, P_Fefin, 'A')
		return Response(file, mimetype="application/pdf",
						headers={"Content-Disposition": "attachment;filename=" + filename})
	except Exception as ex:
		app.logger.error(ex)
		return make_response(jsonify(succes=False, message=ex), 400)


@app.route(context_path + '/agentes/comisiones/excel', methods=['GET'])
def comisiones_agente_xlsx():
	P_Clave = request.args['codigo']
	P_Feini = request.args['desde']
	P_Fefin = request.args['hasta']
	P_Feini = datetime.datetime.strptime(P_Feini, "%Y-%m-%d").strftime("%d/%m/%Y")
	P_Fefin = datetime.datetime.strptime(P_Fefin, "%Y-%m-%d").strftime("%d/%m/%Y")
	try:
		file, filemime, filename = comisiones_xlsx(P_Clave, P_Feini, P_Fefin, 'A')
		return Response(file, mimetype=filemime,
						headers={"Content-Disposition": "attachment;filename=" + filename})
	except Exception as ex:
		app.logger.error(ex)
		return make_response(jsonify(succes=False, message=ex), 400)


@app.route(context_path + '/promotores/bonos/excel', methods=['GET'])
def bono_promotores_xlsx():
	P_Clave = request.args['codigo']
	P_Feini = request.args['desde']
	P_Fefin = request.args['hasta']
	P_Feini = datetime.datetime.strptime(P_Feini, "%Y-%m-%d").strftime("%d/%m/%Y")
	P_Fefin = datetime.datetime.strptime(P_Fefin, "%Y-%m-%d").strftime("%d/%m/%Y")
	try:
		file, filemime, filename = bonos_xlx(P_Clave, P_Feini, P_Fefin, 'P')
		return Response(file, mimetype=filemime,
						headers={"Content-Disposition": "attachment;filename=" + filename})
	except Exception as ex:
		app.logger.error(ex)
		return make_response(jsonify(succes=False, message=ex), 400)

@app.route(context_path + '/promotores/bonos/pdf', methods=['GET'])
def bono_promotores_pdf():
	P_Clave = request.args['codigo']
	P_Feini = request.args['desde']
	P_Fefin = request.args['hasta']
	P_Feini = datetime.datetime.strptime(P_Feini, "%Y-%m-%d").strftime("%d/%m/%Y")
	P_Fefin = datetime.datetime.strptime(P_Fefin, "%Y-%m-%d").strftime("%d/%m/%Y")
	try:
		file, filename = bonos_pdf(P_Clave, P_Feini, P_Fefin, 'P')
		return Response(file, mimetype="application/pdf",
						headers={"Content-Disposition": "attachment;filename=" + filename})
	except Exception as ex:
		app.logger.error(ex)
		return make_response(jsonify(succes=False, message=ex), 400)


@app.route(context_path + '/promotores/comisiones/excel', methods=['GET'])
def comisiones_promotor_xlsx():
	P_Clave = request.args['codigo']
	P_Feini = request.args['desde']
	P_Fefin = request.args['hasta']
	P_Feini = datetime.datetime.strptime(P_Feini, "%Y-%m-%d").strftime("%d/%m/%Y")
	P_Fefin = datetime.datetime.strptime(P_Fefin, "%Y-%m-%d").strftime("%d/%m/%Y")
	try:
		file, filemime, filename = comisiones_xlsx(P_Clave, P_Feini, P_Fefin, 'P')
		return Response(file, mimetype=filemime,
						headers={"Content-Disposition": "attachment;filename=" + filename})
	except Exception as ex:
		app.logger.error(ex)
		return make_response(jsonify(succes=False, message=ex), 400)


@app.route(context_path + '/promotores/comisiones/pdf', methods=['GET'])
def comisiones_promotores_pdf():
	P_Clave = request.args['codigo']
	P_Feini = request.args['desde']
	P_Fefin = request.args['hasta']
	P_Feini = datetime.datetime.strptime(P_Feini, "%Y-%m-%d").strftime("%d/%m/%Y")
	P_Fefin = datetime.datetime.strptime(P_Fefin, "%Y-%m-%d").strftime("%d/%m/%Y")
	try:
		file,filename=comisiones_pdf(P_Clave,P_Feini,P_Fefin,'P')
		return Response(file, mimetype="application/pdf",
						headers={"Content-Disposition": "attachment;filename=" + filename})
	except Exception as ex:
		app.logger.error(ex)
		return make_response(jsonify(succes=False, message=ex), 400)


def comisiones_xlsx(P_Clave,P_Feini,P_Fefin,P_COD):
	try:
		print("Estado de cuentas de Comisiones")
		host = DB_HOST
		port = DB_PORT
		service_name = DB_SERVICE
		user = DB_USER
		password = DB_PWD
		schema = DB_SCHEMA
		sid = cx_Oracle.makedsn(host, port, service_name=service_name)
		# Declaracion de cursores a utilizar
		try:
			connection = cx_Oracle.connect(f"{user}/{password}@{host}:{port}/{service_name}")
			try:
				libro_nombre =P_Clave+"_" + P_Feini.replace("/", "")+"_"+ P_Fefin.replace("/", "")+'_comisiones.xlsx'
				plantilla = "plantilla_agentes.xlsx"
				if P_COD == 'P':
					plantilla = "plantilla_promotor.xlsx"
				wb = opyxl.load_workbook(plantilla)
				ws = wb.worksheets[0]
				ws.cell(row=1, column=6).value = "ESTADOS DE CUENTA DE COMISIONES"
				ws.title = "Estado de Cuenta de Comisiones"
				cursors = []
				for times in range(13):
					cursor = connection.cursor()
					query = getquery(P_COD, 'COMISION', times, P_Clave, P_Feini, P_Fefin)
					cursor.execute(query)
					cursors.append(cursor)
				has_agent = False
				for row in cursors[0]:
					has_agent = True
					for i in range(0, len(row) - 4):
						ws.cell(row=4 + i, column=9).value = row[i]
					for i in range(len(row) - 4, len(row)):
						ws.cell(row=i, column=4).value = row[i]
				if not has_agent:
					raise ValueError('Identificador no encontrado')
				del cursors[0]
				f = 13  # principal gestor de filas del archivo
				greyFill = PatternFill(fill_type='solid', start_color='d9d9d9', end_color='d9d9d9')
				# NUEVO BLOQUE SECUENCIAL
				c_count = 1
				for cursor in cursors:
					fila_totales = [0, 0, 0, 0, 0]
					lista = getHeadColumnsComisones("excel", c_count)
					alphabet_string = string.ascii_uppercase
					alphabet_list = list(alphabet_string)
					ws.cell(row=f, column=1).value = getTableNamesComisiones(c_count)
					ws.cell(row=f, column=1).font = Font(name='Arial', size=9, bold=True)
					f += 1
					j = 0
					for item in lista:
						ws.cell(row=f, column=j + 1).value = item
						ws.cell(row=f, column=j + 1).fill = greyFill
						ws.cell(row=f, column=j + 1).font = Font(name='Arial', size=9, bold=True)
						ws.cell(row=f, column=j + 1).alignment = Alignment(horizontal="center", vertical="center")
						columna = alphabet_list[ws.cell(row=f, column=j + 1).column - 1]
						multiplicador = 2
						lista_columnas_esp = ['A', 'B', 'C', 'H']
						ancho = len(item)
						if len(item) <= 6:
							ancho *= 2
						if len(item) > 6 or columna == 'A':
							ancho *= 1.1
						if columna in lista_columnas_esp:
							ancho = 25
						if c_count == 5:
							ws.column_dimensions[columna].width = ancho
						j += 1
					j = 0
					k = 0
					f += 1
					has_data= False
					for row in cursor:
						for i in range(0, len(row)):
							if c_count not in [5, 6, 7, 8]:
								valor = row[i]
								if c_count in [1, 2]:
									if i != 0:
										valor = "{:,.2f}".format(valor)
								if c_count in [3, 4, 9, 10]:
									if i != 0:
										valor = "{:,.2f}".format(valor)
								if c_count in [11, 12]:
									if i in [1, 3, 6, 7, 8]:
										valor = "{:,.2f}".format(valor)
								ws.cell(row=f, column=i + 1).value = valor
								ws.cell(row=f, column=i + 1).alignment = Alignment(horizontal="center", vertical="center")
								if len(str(valor)) > 17:
									ws.cell(row=f, column=i + 1).font = Font(name='Arial', size=8)
									if len(str(valor)) > 25:
										ws.cell(row=f, column=i + 1).font = Font(name='Arial', size=7)
							else:
								valor = row[i]
								if i in [10, 11, 14, 15]:
									valor = "{:,.2f}".format(valor)
								if i == 10:
									fila_totales[0] += abs(row[i])
								if i == 11:
									fila_totales[1] += abs(row[i])
								if i == 13:
									fila_totales[2] += abs(row[i])
								if i == 14:
									fila_totales[3] += abs(row[i])
								if i == 15:
									fila_totales[4] += abs(row[i])
								has_data = True
								ws.cell(row=f, column=i + 1).value = valor
								ws.cell(row=f, column=i + 1).alignment = Alignment(horizontal="center", vertical="center")
								if len(str(valor)) > 17:
									ws.cell(row=f, column=i + 1).font = Font(name='Arial', size=8)
									if len(str(valor)) > 25:
										ws.cell(row=f, column=i + 1).font = Font(name='Arial', size=7)
						f += 1
					cursor.close()
					if c_count in [5, 6, 7, 8] and has_data:
						fila_totales[0] = "{:,.2f}".format(fila_totales[0])
						fila_totales[1] = "{:,.2f}".format(fila_totales[1])
						fila_totales[2] = "{:,.2f}".format(fila_totales[2])
						fila_totales[3] = "{:,.2f}".format(fila_totales[3])
						fila_totales[4] = "{:,.2f}".format(fila_totales[4])
						ws.cell(row=f, column=10).value = "TOTAL"
						ws.cell(row=f, column=10).alignment = Alignment(horizontal="center", vertical="center")
						ws.cell(row=f, column=11).value = fila_totales[0]
						ws.cell(row=f, column=11).alignment = Alignment(horizontal="center", vertical="center")
						ws.cell(row=f, column=12).value = fila_totales[1]
						ws.cell(row=f, column=12).alignment = Alignment(horizontal="center", vertical="center")
						ws.cell(row=f, column=14).value = fila_totales[2]
						ws.cell(row=f, column=14).alignment = Alignment(horizontal="center", vertical="center")
						ws.cell(row=f, column=15).value = fila_totales[3]
						ws.cell(row=f, column=15).alignment = Alignment(horizontal="center", vertical="center")
						ws.cell(row=f, column=16).value = fila_totales[4]
						ws.cell(row=f, column=16).alignment = Alignment(horizontal="center", vertical="center")
						f += 1
					c_count += 1
				# fin de bloque
				virtual_wb = BytesIO()
				wb.save(virtual_wb)
				return virtual_wb.getvalue(), wb.mime_type,libro_nombre
			except Exception as ex:
				app.logger.error(ex)
				raise ValueError('Error generando el reporte')
		except Exception as ex:
			app.logger.error(ex)
			raise ValueError('Error en la conexion con la base de datos.')
	except Exception as ex:
		app.logger.error(ex)
		raise ValueError('Error en la configuracion de la base de datos.')


def comisiones_pdf(P_Clave,P_Feini,P_Fefin,P_COD):
	try:
		print("Estado de cuentas comisiones")
		host = DB_HOST
		port = DB_PORT
		service_name = DB_SERVICE
		user = DB_USER
		password = DB_PWD
		schema = DB_SCHEMA
		sid = cx_Oracle.makedsn(host, port, service_name=service_name)
		# Declaracion de cursores a utilizar
		try:
			connection = cx_Oracle.connect(f"{user}/{password}@{host}:{port}/{service_name}")
			try:
				has_agent = False
				cursors = []
				for times in range(13):
					cursor = connection.cursor()
					query = getquery(P_COD, 'COMISION', times, P_Clave, P_Feini, P_Fefin)
					cursor.execute(query)
					cursors.append(cursor)
				libro_nombre =P_Clave+"_" + P_Feini.replace("/", "")+"_"+ P_Fefin.replace("/", "")+'_comisiones.pdf'

				virtual_wb = BytesIO()
				doc = SimpleDocTemplate(virtual_wb,pagesize=landscape((432*mm, 546*mm)))
				flowables = []
				for row in cursors[0]:
					has_agent = True
					lista_aux = []
					for i in range(0, len(row)):
						lista_aux.append(row[i])
				if not has_agent:
					raise ValueError('Identificador no encontrado')
				header_all = getheaderpdf(P_COD,lista_aux,'COMISIONES')
				grid = [('FONTNAME', (0, 0), (-1,0), 'Courier-Bold')]
				tbl = Table(header_all)
				tbl.setStyle(grid)
				flowables.append(tbl)
				del cursors[0]
				c_count = 1
				tblstyle = TableStyle(
					[('GRID', (0, 0), (0, 0), 0.25, colors.gray), ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
					 ('FONTSIZE', (0, 0), (0, 0), 7)])
				for cursor in cursors:
					lista = getHeadColumnsComisones("pdf", c_count)
					data_cursor = []
					taux = Table([("", getTableNamesComisiones(c_count), ""), ("", "", "")])
					taux.setStyle(grid)
					flowables.append(taux)
					data_cursor.append(lista)
					fila_totales = ["", "", "", "", "", "", "", "TOTAL", 0, 0, "", 0, 0, 0, "", ""]
					for row in cursor:
						lista_aux = []
						for i in range(0, len(row)):
							if c_count not in [5,6,7,8]:
								valor = row[i]
								if c_count in [1,2]:
									if i != 0:
										if valor < 0:
											valor = "(" + "{:,.2f}".format(abs(valor)) + ")"
										else:
											valor = "{:,.2f}".format(valor)
								if c_count in [3,4,9,10]:
									if i != 0:
										if valor < 0:
											valor = "(" + "{:,.2f}".format(abs(valor)) + ")"
										else:
											valor = "{:,.2f}".format(valor)
								if c_count in [11,12]:
									if i in [1,3,6,7,8]:
										if valor < 0:
											valor = "(" + "{:,.2f}".format(abs(valor)) + ")"
										else:
											valor = "{:,.2f}".format(valor)
								lista_aux.append(valor)
							else:
								if i not in [1,6,17]:
									valor = row[i]
									if i in [10, 11, 14, 15]:
										if valor < 0:
											valor = "(" + "{:,.2f}".format(abs(valor)) + ")"
										else:
											valor = "{:,.2f}".format(valor)
									lista_aux.append(valor)
									if i == 10:
										fila_totales[8] += abs(row[i])
									if i == 11:
										fila_totales[9] += abs(row[i])
									if i == 13:
										fila_totales[11] += abs(row[i])
									if i == 14:
										fila_totales[12] += abs(row[i])
									if i == 15:
										fila_totales[13] += abs(row[i])
						data_cursor.append(lista_aux)
					if c_count in [5,6,7,8] and len(data_cursor)>0:
						fila_totales[8] = "{:,.2f}".format(fila_totales[8])
						fila_totales[9] = "{:,.2f}".format(fila_totales[9])
						fila_totales[11] = "{:,.2f}".format(fila_totales[11])
						fila_totales[12] = "{:,.2f}".format(fila_totales[12])
						fila_totales[13] = "{:,.2f}".format(fila_totales[13])
						data_cursor.append(fila_totales)
					tbl = Table(data_cursor)
					tbl.setStyle(tblstyle)
					flowables.append(tbl)
					flowables.append(Table([("", " ", ""), ("", "", "")]))
					c_count += 1
					cursor.close()
				# fin de bloque
				doc.build(flowables)
				return virtual_wb.getvalue(), libro_nombre
			except Exception as ex:
				app.logger.error(ex)
				raise ValueError('Error en la generacion del reporte.')
		except Exception as ex:
			app.logger.error(ex)
			raise ValueError('Error en la conexion con la base de datos.')
	except Exception as ex:
		app.logger.error(ex)
		raise ValueError('Error en la configuracion de la base de datos.')


def bonos_pdf(P_Clave,P_Feini,P_Fefin,P_COD):
	try:
		print("""

					\  \      __             __             _                  /  /
					 >  >    |_  _| _       /  _|_ _       |_) _ __  _  _     <  < 
					/  /     |__(_|(_) o    \__ |_(_| o    |_)(_)| |(_)_>      \  \ 

					""")

		host = DB_HOST
		port = DB_PORT
		service_name = DB_SERVICE
		user = DB_USER
		password = DB_PWD
		schema = DB_SCHEMA
		sid = cx_Oracle.makedsn(host, port, service_name=service_name)
		# Declaracion de cursores a utilizar
		try:
			connection = cx_Oracle.connect(f"{user}/{password}@{host}:{port}/{service_name}")
			try:
				has_agent = False
				cursors = []
				for times in range(2):
					cursor = connection.cursor()
					query = getquery(P_COD, 'BONO', times, P_Clave, P_Feini, P_Fefin)
					cursor.execute(query)
					cursors.append(cursor)

				virtual_wb = BytesIO()
				doc = SimpleDocTemplate(virtual_wb,pagesize=landscape((432*mm, 546*mm)))
				flowables = []
				libro_nombre = P_Clave+"_" + P_Feini.replace("/", "")+"_"+ P_Fefin.replace("/", "")+'_bonos.pdf'

				for row in cursors[0]:
					has_agent = True
					lista_aux = []
					for i in range(0, len(row)):
						lista_aux.append(row[i])
				cursors[0].close
				if not has_agent:
					raise ValueError('Identificador no encontrado')
				header_all = getheaderpdf(P_COD,lista_aux,'BONOS')
				grid = [('FONTNAME', (0, 0), (-1,0), 'Courier-Bold')]
				tbl = Table(header_all)
				tbl.setStyle(grid)
				flowables.append(tbl)

				j = 0
				lista = getHeadColumns("pdf")
				data_body = []
				lista_aux = []
				has_data=False
				for item in lista:
					lista_aux.append(item)
				data_body.append(lista_aux)
				for row in cursors[1]:
					has_data=True
					lista_aux = []
					for i in range(0, len(row)):
						if i < 21 and i not in [3,8,9,19]:
							if i == 2:
								lista_aux.append(getTipoSubBono(row[i]))
							else:
								lista_aux.append(row[i])
					data_body.append(lista_aux)
				cursors[1].close()
				if not has_data:
					raise ValueError('Error generando el reporte.')
				tbl = Table(data_body)
				tblstyle = TableStyle([('GRID',(0,0),(0,0),0.25,colors.gray),('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),('FONTSIZE', (0, 0), (0, 0), 7)])
				tbl.setStyle(tblstyle)
				flowables.append(tbl)
				doc.build(flowables)
				return virtual_wb.getvalue(), libro_nombre
			except Exception as ex:
				raise ValueError('Error generando el reporte.')
		except Exception as ex:
			raise ValueError('Error de conexion con la base de datos.')
	except Exception as ex:
		app.logger.error(ex)
		raise ValueError('Error de configuracion de base de datos.')


def bonos_xlx(P_Clave,P_Feini,P_Fefin,P_COD):
	try:
		print("""

					\  \      __             __             _                  /  /
					 >  >    |_  _| _       /  _|_ _       |_) _ __  _  _     <  < 
					/  /     |__(_|(_) o    \__ |_(_| o    |_)(_)| |(_)_>      \  \ 

					""")
		host = DB_HOST
		port = DB_PORT
		service_name = DB_SERVICE
		user = DB_USER
		password = DB_PWD
		schema = DB_SCHEMA
		sid = cx_Oracle.makedsn(host, port, service_name=service_name)
		# Declaracion de cursores a utilizar
		try:
			connection = cx_Oracle.connect(f"{user}/{password}@{host}:{port}/{service_name}")
			try:
				plantilla = "plantilla_agentes.xlsx"
				if P_COD == 'P':
					plantilla = "plantilla_promotor.xlsx"
				wb = opyxl.load_workbook(plantilla)
				ws = wb.worksheets[0]
				cursors = []
				for times in range(2):
					cursor = connection.cursor()
					query = getquery(P_COD,'BONO',times,P_Clave,P_Feini,P_Fefin)
					cursor.execute(query)
					cursors.append(cursor)

				libro_nombre = P_Clave+"_" + P_Feini.replace("/", "")+"_"+ P_Fefin.replace("/", "")+'_bonos.xlsx'
				for row in cursors[0]:
					has_agent = True
					for i in range(0, len(row) - 4):
						ws.cell(row=4 + i, column=9).value = row[i]
					for i in range(len(row) - 4, len(row)):
						ws.cell(row=i, column=4).value = row[i]
				cursors[0].close()
				if not has_agent:
					raise ValueError('Identificador no encontrado')
				j = 0
				greyFill = PatternFill(fill_type='solid', start_color='d9d9d9', end_color='d9d9d9')
				lista = getHeadColumns("excel")
				alphabet_string = string.ascii_uppercase
				alphabet_list = list(alphabet_string)
				for item in lista:
					ws.cell(row=13, column=j + 1).value = item
					ws.cell(row=13, column=j + 1).fill = greyFill
					ws.cell(row=13, column=j + 1).font = Font(name='Arial', size=9, bold=True)
					ws.cell(row=13, column=j + 1).alignment = Alignment(horizontal="center", vertical="center")
					columna = alphabet_list[ws.cell(row=13, column=j + 1).column - 1]
					multiplicador = 2
					lista_columnas_esp = ['B', 'G']
					ancho = len(item)
					if len(item) <= 6:
						ancho *= 2
					if len(item) > 6 or columna == 'A':
						ancho *= 1.1
					if columna in lista_columnas_esp:
						ancho = 25
					ws.column_dimensions[columna].width = ancho
					j += 1
				j = 0
				k = 0
				has_data = False
				for row in cursors[1]:
					has_data = True
					lista_razones = []
					for i in range(0, len(row)):

						if i < 21:
							aux = 0
							if i >=9:
								aux=1
							valor = row[i]
							if i == 2:
								valor = getTipoSubBono(row[i])
							ws.cell(row=14 + j, column=i + 1+aux).value = valor
							ws.cell(row=14 + j, column=i + 1+aux).alignment = Alignment(horizontal="center",vertical="center")
							if len(str(valor)) > 17:
								ws.cell(row=14 + j, column=i + 1+aux).font = Font(name='Arial', size=8)
								if len(str(valor)) > 25:
									ws.cell(row=14 + j, column=i + 1+aux).font = Font(name='Arial', size=7)
						else:
							if i == 21:
								if row[i] == 'NO':
									ws.cell(row=14 + j, column=10).value = 'Sí'
								else:
									ws.cell(row=14 + j, column=10).value = 'No'
							if i > 21 and row[21] == 'SI':
								if row[i] is not None and len(row[i]) > 0:
									lista_razones.append(row[i])
					if row[21] == 'SI' and len(lista_razones) > 0:
						ws.cell(row=14 + j, column=11).value = ", ".join(lista_razones)
					if row[21] == 'NO':
						ws.cell(row=14 + j, column=11).value = ' '
					j += 1
				cursors[1].close()
				if not has_data:
					raise ValueError('Error en la generacion del reporte.')
				virtual_wb = BytesIO()
				wb.save(virtual_wb)
				return virtual_wb.getvalue(),wb.mime_type,libro_nombre
			except Exception as ex:
				app.logger.error(ex)
				raise ValueError('Error en la generacion del reporte.')
		except Exception as ex:
			app.logger.error(ex)
			raise ValueError('Error en la conexion con la base de datos.')
		print("\nTermina Proceso " + time.strftime("%X"))
	except Exception as ex:
		app.logger.error(ex)
		raise ValueError('Error en la configuracion de la base de datos.')



def getHeadColumns(extension):
    lista = []
    lista.append('# Bono')
    lista.append('Tipo Bono')
    lista.append('Subtipo Bono')
    if extension == "excel":
        lista.append('Grupo')
    lista.append('Oficina')
    lista.append('Ramo')
    lista.append('Poliza')
    lista.append('Contratante')
    if extension == "excel":
        lista.append('Agentes')
        lista.append('Computabilidad')
        lista.append('Causa de No Computabilidad')
    lista.append('Tipo Cambio')
    lista.append('# Recibo')
    lista.append('Serie')
    lista.append('Prima Total')
    lista.append('Prima Neta')
    lista.append('% Bono Pagado')
    lista.append('Monto comisión neta')
    lista.append('Total comisión pagada')
    lista.append('# Liquidación')
    if extension == "excel":
        lista.append('# Comprobante')
    lista.append('Fecha Movimiento')
    return lista


def getHeadColumnsComisones(extension,cursor):
	lista = []
	if cursor in [1,2]:
		lista.append("TIPO")
		lista.append("BASE")
		lista.append("IVA")
		lista.append("SUBTOTAL")
		lista.append("IVARETENIDO")
		lista.append("ISR")
		lista.append("IMP. CEDULAR")
		lista.append("TOTAL")
	if cursor in [3,4,9,10]:
		lista.append("CONCEPTO")
		lista.append("IMPORTE")

	if cursor in [5,6,7,8]:
		lista.append("Daños/Vida")
		if extension =="excel":
			lista.append("Grupo")
		lista.append("Oficina")
		lista.append("Ramo")
		lista.append("Poliza")
		lista.append("Contratante")
		if extension == "excel":
			lista.append("Clave Agente")
		lista.append("Tipo de Cambio")
		lista.append("# Recibo")
		lista.append("Serie de Recibo")
		lista.append("Prima Total")
		lista.append("Prima Neta")
		lista.append("% Comisión pagado")
		lista.append("% Comisión de derecho")
		lista.append("Monto Comisión Neta")
		lista.append("Total Comisión pagado")
		lista.append("# Liquidación")
		if extension == "excel":
			lista.append("# Comprobante")
		lista.append("Fecha aplicación de la póliza")
	if cursor in [11,12]:
		lista.append("FECHA DE LIQUIDACIÓN")
		lista.append("DAÑOS")
		lista.append("VIDA")
		lista.append("TOTAL")
		lista.append("FECHA DE PAGO")
		lista.append("NUMERO DE COMPROBANTE")
		lista.append("IMPORTEPAGADO DAÑOS")
		lista.append("IMPORTE PAGADO VIDA")
		lista.append("TOTAL")
	return lista


def getTipoSubBono(id):
	codigo=""
	if id == 20:
		codigo = "BS"
	if id == 30:
		codigo = "BC"
	if id == 40:
		codigo = "BP"
	if id == 50:
		codigo = "BSE"
	if id == 60:
		codigo = "BCS"
	if id == 70:
		codigo = "BPS"
	if id == 80:
		codigo = "BPF"
	if id == 82:
		codigo = "BD"
	return codigo


def getTableNamesComisiones(tabla):
	nombre=""
	if tabla == 1:
		nombre = "TOTAL DE PERCEPCIONES MENSUALES"
	if tabla == 2:
		nombre = "TOTAL DE PERCEPCIONES ACUMULADO ANUAL"
	if tabla in [3,5]:
		nombre = "DAÑOS MONEDA MXP"
	if tabla in [4,6]:
		nombre = "DAÑOS MONEDA USD"
	if tabla in [9,7]:
		nombre = "VIDA MONEDA MXP"
	if tabla in [10,8]:
		nombre = "VIDA MONEDA USD"
	if tabla == 11:
		nombre = "RESUMEN DE DEPOSITOS EN MXP"
	if tabla == 12:
		nombre = "RESUMEN DE DEPOSITOS EN USD"
	return nombre


def getquery(clave,tipo,tabla,codigo,desde,hasta):
	query = ""
	if clave == 'A':
		if tipo == 'BONO':
			if tabla == 0:
				query = getheaderquery('A',codigo,desde,hasta)
			if tabla == 1:
				query = f"SELECT MAIN.NMBONO,MAIN.DSBONO,MAIN.CDTIPBON, MAIN.CDGRUPO,MAIN.CDUNIECO, MAIN.CDRAMO,MAIN.NMPOLIZA, nomtomador(MAIN.CDUNIECO, MAIN.CDRAMO, 'M', MAIN.NMPOLIZA, 9999999999999999999) CONTRATANTE, MAIN.cdagente,(SELECT OTVALOR07 FROM TVALOPOL Z WHERE Z.CDUNIECO=MAIN.CDUNIECO AND Z.CDRAMO=MAIN.CDRAMO AND Z.NMPOLIZA=MAIN.NMPOLIZA AND Z.ESTADO='M' AND Z.NMSUPLEM=(SELECT MAX(NMSUPLEM) FROM TVALOPOL Y WHERE Y.CDUNIECO=Z.CDUNIECO AND Z.CDRAMO=Y.CDRAMO AND Z.NMPOLIZA=Y.NMPOLIZA AND Z.ESTADO=Y.ESTADO)) COMPUTABILIDAD, MAIN.TIPOCAMB, MAIN.NMRECIBO,MAIN.SERIERECIBO,NVL(MAIN.PMATOTAL,0) PRIMA_TOTAL,NVL(MAIN.PMANETAP,0) PRIMA_NETA,MAIN.PORCEPAG, ABS(MAIN.IMPORAGTN),(ABS(MAIN.IMPORAGTN)+ABS(MAIN.IMPORAGTI)) TOTAL_PAGADO ,MAIN.NUMPRELIQ, DECODE(NVL(MAIN.CDCOMPRO,0),0,'',DECODE(MAIN.CDMETPAG,'EFT',MAIN.NMTRANSF,'CHK',MAIN.NMCHEQUE,MAIN.CDCOMPRO)) NUM_COMPROBANTE, TO_CHAR(MAIN.FEMOVIMI,'DD/MM/YYYY') FECHA, (DECODE ((SELECT DISTINCT 'SI' FROM bon_trecexc WHERE nmbono = MAIN.nmbono AND cdrecpag = MAIN.cdrecpag " \
						f"AND cdagente = MAIN.cdagente AND cdtipmov = MAIN.cdtipmov),'SI', 'SI', 'NO' )) EXCLUDO, (SELECT DECODE(swexcep, '1', 'Por agente para Rank', '2', 'Por agente para Pago', '3', 'Por agente para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 1) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov AND ROWNUM < 2) POR_AGENTE, (SELECT DECODE(swexcep, '1', 'Por promotor para Rank', '2', 'Por promotor para Pago', '3', 'Por promotor para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 2) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov AND ROWNUM < 2) POR_PROMOTOR, (SELECT DECODE(swexcep, '1', 'Por ramo para Rank', '2', 'Por ramo para Pago', '3', 'Por ramo para Todo') FROM bon_trecexc aa, " \
						f"(SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 4) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov AND ROWNUM < 2) POR_RAMO, (SELECT DECODE(swexcep, '1', 'Por oficina para Rank', '2', 'Por oficina para Pago', '3', 'Por oficina para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 5) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov) POR_OFICINA, (SELECT DECODE(swexcep,'1', 'Por poliza para Rank', '2', 'Por poliza para Pago', '3', 'Por poliza para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 6) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov AND ROWNUM < 2) " \
						f"POR_POLIZA, (SELECT DECODE(swexcep, '1', 'Por grupo para Rank', '2', 'Por grupo para Pago', '3', 'Por grupo para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 7) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov AND ROWNUM < 2) POR_GRUPO, (SELECT DECODE(swexcep, '1', 'Por subgrupo para Rank', '2', 'Por subgrupo para Pago', '3', 'Por subgrupo para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 8) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov) POR_SUBGRUPO, (SELECT DECODE(swexcep, '1', 'Por tipo de persona para Rank', '2', 'Por tipo de persona para Pago', '3', 'Por tipo de persona para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 9) bb " \
						f"WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov AND ROWNUM < 2) POR_TIP_PERS, (SELECT DECODE(swexcep, '1', 'Por Reaseguro para Rank', '2', 'Por Reaseguro para Pago', '3', 'Por Reaseguro para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 10) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov AND ROWNUM < 2) POR_REASEGURO, (SELECT DECODE(swexcep, '1', 'Por Coaseguro para Rank', '2', 'Por Coaseguro para Pago', '3', 'Por Coaseguro para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 11) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov AND ROWNUM < 2) POR_COASEGURO, (SELECT DECODE(swexcep, '1', " \
						f"'Por multianual para Rank', '2', 'Por multianual para Pago', '3', 'Por multianual para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 12) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov AND ROWNUM < 2) POR_MULTIANUAL, (SELECT DECODE(swexcep, '1', 'Por dividendos para Rank', '2', 'Por dividendos para Pago', '3', 'Por dividendos para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 13) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov AND ROWNUM < 2) POR_DIVIDENDOS, (SELECT DECODE(swexcep, '1', 'Por no computable para Rank', '2', 'Por no computable para Pago', '3', 'Por no computable para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 14) bb WHERE " \
						f"aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov AND ROWNUM < 2) POR_NO_COMPU, (SELECT DECODE(swexcep, '1', 'Por portafolio para Rank', '2', 'Por portafolio para Pago', '3', 'Por portafolio para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 15) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov AND ROWNUM < 2) POR_PORTAFOLIO, (SELECT DECODE(swexcep, '1', 'Por prestador para Rank', '2', 'Por prestador para Pago', '3', 'Por prestador para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 17) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov AND ROWNUM < 2) POR_PRESTADOR, (SELECT DECODE(swexcep, '1', " \
						f"'Por negocio para Rank', '2', 'Por negocio para Pago', '3', 'Por negocio para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 18) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov AND ROWNUM < 2) POR_NEGOCIO FROM (SELECT E.TIPOCAMB, C.NMBONO,D.DSBONO,C.CDTIPBON, E.CDGRUPO,E.CDUNIECO, E.CDRAMO,E.NMPOLIZA,a.cdagente,E.NMRECIBO,F.SERIERECIBO,G.NMTRANSF,G.CDMETPAG,G.NMCHEQUE,G.CDCOMPRO, F.PMATOTAL,F.PMANETAP,C.PORCEPAG,F.IMPORAGTN,F.IMPORAGTI,A.NUMPRELIQ,A.FEMOVIMI,E.CDTIPMOV,E.CDRECPAG FROM TSIA_MOVAGEN A , BON_TPAGLIQ B, BON_TCONFBON C, BON_TCATBONO D, VBON_DESGLOSE E, tsia_detcom F, MINTPSOFT G WHERE A.CDAGENTE ='{codigo}' AND A.FEMOVIMI BETWEEN TO_DATE('{desde}','dd/mm/yyyy') AND TO_DATE('{hasta}','dd/mm/yyyy') AND A.CDCONC IN ('B','BC') AND B.NUMPRELIQ =A.NUMPRELIQ AND B.CDAGPROU = A.CDAGENTE AND B.CDCVEMOV = A.CDCVEMOV AND B.CDCONC = A.CDCONC AND C.NMBONO = B.NMBONO AND " \
						f"D.CDBONO = C.CDBONO AND E.NMBONO= C.NMBONO AND A.CDAGENTE= E.CDAGENTE AND E.NMPOLIZA= F.NMPOLIZA AND E.CDUNIECO= F.CDUNIECO AND E.CDRAMO= F.CDRAMO AND E.NMRECIBO= F.NMRECIBO AND A.CDAGENTE= F.CDAGENTE AND B.NUMPRELIQ =G.NUMPRELI (+) AND A.CDCVEMOV = B.CDCVEMOV(+) AND A.CDCONC = B.CDCONC (+) GROUP BY E.TIPOCAMB,C.NMBONO,D.DSBONO, E.CDGRUPO, E.CDUNIECO, E.CDRAMO,E.NMPOLIZA,a.cdagente, F.SERIERECIBO,G.NMTRANSF,C.CDTIPBON, F.PMATOTAL,F.PMANETAP,G.CDMETPAG,G.NMCHEQUE,G.CDCOMPRO, C.PORCEPAG,F.IMPORAGTN, F.IMPORAGTI,E.CDTIPMOV,E.CDRECPAG, A.NUMPRELIQ,E.NMRECIBO,A.FEMOVIMI ORDER BY FEMOVIMI ASC) MAIN"
		if tipo == 'COMISION':
			if tabla == 0:
				query = getheaderquery('A',codigo,desde,hasta)
			if tabla == 1:
				query = f"select a.tipo, round (nvl(a.HONORARIOS,0),2) base, round (nvl(a.IVA,0),2) iva , round (nvl(a.HONORARIOS,0) + nvl(a.IVA,0),2) Subtotal, round (nvl(a.IVA_RETENIDO,0),2) IVA_RETENIDO, round (nvl(a.ISR,0),2)ISR, round (nvl(a.IMPUESTO_CEDULAR,0),2) IMPUESTO_CEDULAR, round (nvl(a.HONORARIOS,0) + nvl(a.IVA,0)+ nvl(a.IVA_RETENIDO,0) + nvl(a.ISR,0) + nvl(a.IMPUESTO_CEDULAR,0),2) total from ( SELECT a.cdagente, 'DAÑOS' tipo, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'V',NVL(a.ptimport,0),0),0) * (SELECT ptcambio FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) = (SELECT MAX(TRUNC(fevalor)) FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) BETWEEN TO_DATE('01'||LPAD(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),2,0)||to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY'),'DDMMYYYY') AND TO_DATE('01'||LPAD(DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM') + 1),2,0)||DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12," \
						f"to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY') + 1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY')),'DDMMYYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'V',NVL(a.ptimport,0),0),0)),0)) IVA, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'R',NVL(a.ptimport,0),0),0) * (SELECT ptcambio FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) = (SELECT MAX(TRUNC(fevalor)) FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) BETWEEN TO_DATE('01'||LPAD(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),2,0)||to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY'),'DDMMYYYY') AND TO_DATE('01'||LPAD(DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM') + 1),2,0)||DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY') + 1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY')),'DDMMYYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'R',NVL(a.ptimport,0),0),0)),0)) IVA_RETENIDO, " \
						f"(NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'I',NVL(a.ptimport,0),0),0) * (SELECT ptcambio FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) = (SELECT MAX(TRUNC(fevalor)) FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) BETWEEN TO_DATE('01'||LPAD(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),2,0)||to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY'),'DDMMYYYY') AND TO_DATE('01'||LPAD(DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM') + 1),2,0)||DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY') + 1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY')),'DDMMYYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'I',NVL(a.ptimport,0),0),0)),0)) ISR, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'C',NVL(a.ptimport,0),0),0) * (SELECT ptcambio FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) = (SELECT MAX(TRUNC(fevalor)) FROM tcambios WHERE " \
						f"cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) BETWEEN TO_DATE('01'||LPAD(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),2,0)||to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY'),'DDMMYYYY') AND TO_DATE('01'||LPAD(DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM') + 1),2,0)||DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY') + 1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY')),'DDMMYYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'C',NVL(a.ptimport,0),0),0)),0)) IMPUESTO_CEDULAR, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'T',NVL(a.ptimport,0),0),0) * (SELECT ptcambio FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) = (SELECT MAX(TRUNC(fevalor)) FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) BETWEEN TO_DATE('01'||LPAD(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),2,0)||to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY'),'DDMMYYYY') AND " \
						f"TO_DATE('01'||LPAD(DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM') + 1),2,0)||DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY') + 1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY')),'DDMMYYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'T',NVL(a.ptimport,0),0),0)),0)) HONORARIOS FROM tsia_movagen a, tsia_tipmovi b WHERE a.cdcvemov < 1000 AND a.cdconc = b.cdconc AND a.cdcvemov = b.cdcvemov AND b.swconinm IN ('C','I','R','V','T') AND TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') and A.CDAGENTE='{codigo}' GROUP BY a.cdagente UNION ALL SELECT a.cdagente, 'VIDA' tipo, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'V',NVL(a.ptimport,0),0),0) * (SELECT ptcambio FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) = (SELECT MAX(TRUNC(fevalor)) FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) BETWEEN " \
						f"TO_DATE('01'||LPAD(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),2,0)||to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY'),'DDMMYYYY') AND TO_DATE('01'||LPAD(DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM') + 1),2,0)||DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY') + 1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY')),'DDMMYYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'V',NVL(a.ptimport,0),0),0)),0)) IVA, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'R',NVL(a.ptimport,0),0),0) * (SELECT ptcambio FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) = (SELECT MAX(TRUNC(fevalor)) FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) BETWEEN TO_DATE('01'||LPAD(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),2,0)||to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY'),'DDMMYYYY') AND TO_DATE('01'||LPAD(DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,1," \
						f"to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM') + 1),2,0)||DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY') + 1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY')),'DDMMYYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'R',NVL(a.ptimport,0),0),0)),0)) IVA_RETENIDO, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'I',NVL(a.ptimport,0),0),0) * (SELECT ptcambio FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) = (SELECT MAX(TRUNC(fevalor)) FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) BETWEEN TO_DATE('01'||LPAD(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),2,0)||to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY'),'DDMMYYYY') AND TO_DATE('01'||LPAD(DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM') + 1),2,0)||DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY') + 1,to_char(TO_DATE('{hasta}','DD/MM/YYYY')," \
						f"'YYYY')),'DDMMYYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'I',NVL(a.ptimport,0),0),0)),0)) ISR, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'C',NVL(a.ptimport,0),0),0) * (SELECT ptcambio FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) = (SELECT MAX(TRUNC(fevalor)) FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) BETWEEN TO_DATE('01'||LPAD(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),2,0)||to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY'),'DDMMYYYY') AND TO_DATE('01'||LPAD(DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM') + 1),2,0)||DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY') + 1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY')),'DDMMYYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'C',NVL(a.ptimport,0),0),0)),0)) IMPUESTO_CEDULAR, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'T',NVL(a.ptimport,0),0),0) * (SELECT ptcambio " \
						f"FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) = (SELECT MAX(TRUNC(fevalor)) FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) BETWEEN TO_DATE('01'||LPAD(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),2,0)||to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY'),'DDMMYYYY') AND TO_DATE('01'||LPAD(DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM') + 1),2,0)||DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY') + 1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY')),'DDMMYYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'T',NVL(a.ptimport,0),0),0)),0)) HONORARIOS FROM tsia_movagen a, tsia_tipmovi b WHERE a.cdcvemov >= 1000 AND a.cdconc = b.cdconc AND a.cdcvemov = b.cdcvemov AND b.swconinm IN ('C','I','R','V','T') AND TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') and A.CDAGENTE='{codigo}' GROUP BY a.cdagente ) a"
			if tabla == 2:
				query = f"SELECT a.tipo, round (nvl(a.HONORARIOS,0),2) base, round (nvl(a.IVA,0),2) iva , round (nvl(a.HONORARIOS,0) + nvl(a.IVA,0),2) Subtotal, round (nvl(a.IVA_RETENIDO,0),2) IVA_RETENIDO, round (nvl(a.ISR,0),2)ISR, round (nvl(a.IMPUESTO_CEDULAR,0),2) IMPUESTO_CEDULAR, round (nvl(a.HONORARIOS,0) + nvl(a.IVA,0)+ nvl(a.IVA_RETENIDO,0) + nvl(a.ISR,0) + nvl(a.IMPUESTO_CEDULAR,0),2) total FROM ( SELECT a.cdagente, 'DAÑOS' tipo, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'V',NVL(a.ptimport,0),0),0) * F_TIPO_CAMBIO(A.FEMOVIMI, A.CDAGENTE, 'A', (SELECT MAX(femovimi) FROM tsia_movagen WHERE cdconc = 'SA' AND femovimi < TRUNC(TO_DATE('{hasta}','DD-MM-YYYY'),'YYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'V',NVL(a.ptimport,0),0),0)),0)) IVA, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'R',NVL(a.ptimport,0),0),0) * F_TIPO_CAMBIO(A.FEMOVIMI, A.CDAGENTE, 'A', (SELECT MAX(femovimi) FROM tsia_movagen WHERE cdconc = 'SA' AND femovimi < TRUNC(TO_DATE('{hasta}','DD-MM-YYYY'),'YYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP'," \
						f"DECODE(b.swconinm,'R',NVL(a.ptimport,0),0),0)),0)) IVA_RETENIDO, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'I',NVL(a.ptimport,0),0),0) * F_TIPO_CAMBIO(A.FEMOVIMI, A.CDAGENTE, 'A', (SELECT MAX(femovimi) FROM tsia_movagen WHERE cdconc = 'SA' AND femovimi < TRUNC(TO_DATE('{hasta}','DD-MM-YYYY'),'YYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'I',NVL(a.ptimport,0),0),0)),0)) ISR, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'C',NVL(a.ptimport,0),0),0) * F_TIPO_CAMBIO(A.FEMOVIMI, A.CDAGENTE, 'A', (SELECT MAX(femovimi) FROM tsia_movagen WHERE cdconc = 'SA' AND femovimi < TRUNC(TO_DATE('{hasta}','DD-MM-YYYY'),'YYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'C',NVL(a.ptimport,0),0),0)),0)) IMPUESTO_CEDULAR, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'T',NVL(a.ptimport,0),0),0) * F_TIPO_CAMBIO(A.FEMOVIMI, A.CDAGENTE, 'A', (SELECT MAX(femovimi) FROM tsia_movagen WHERE cdconc = 'SA' AND femovimi < TRUNC(TO_DATE('{hasta}','DD-MM-YYYY'),'YYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP'," \
						f"DECODE(b.swconinm,'T',NVL(a.ptimport,0),0),0)),0)) HONORARIOS FROM tsia_movagen a, tsia_tipmovi b WHERE a.cdcvemov < 1000 AND a.cdconc = b.cdconc AND a.cdcvemov = b.cdcvemov AND b.swconinm IN ('C','I','R','V','T') AND TRUNC(a.femovimi) BETWEEN (SELECT MAX(femovimi) FROM tsia_movagen WHERE cdconc = 'SA' AND femovimi < TRUNC(TO_DATE('{hasta}','DD-MM-YYYY'),'YYYY')) AND to_date('{hasta}','dd/mm/yyyy') and A.CDAGENTE='{codigo}' GROUP BY a.cdagente UNION ALL SELECT a.cdagente, 'VIDA' tipo, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'V',NVL(a.ptimport,0),0),0) * F_TIPO_CAMBIO(A.FEMOVIMI, A.CDAGENTE, 'A', (SELECT MAX(femovimi) FROM tsia_movagen WHERE cdconc = 'SA' AND femovimi < TRUNC(TO_DATE('{hasta}','DD-MM-YYYY'),'YYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'V',NVL(a.ptimport,0),0),0)),0)) IVA, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'R',NVL(a.ptimport,0),0),0) * F_TIPO_CAMBIO(A.FEMOVIMI, A.CDAGENTE, 'A', (SELECT MAX(femovimi) FROM tsia_movagen WHERE cdconc = 'SA' " \
						f"AND femovimi < TRUNC(TO_DATE('{hasta}','DD-MM-YYYY'),'YYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'R',NVL(a.ptimport,0),0),0)),0)) IVA_RETENIDO, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'I',NVL(a.ptimport,0),0),0) * F_TIPO_CAMBIO(A.FEMOVIMI, A.CDAGENTE, 'A', (SELECT MAX(femovimi) FROM tsia_movagen WHERE cdconc = 'SA' AND femovimi < TRUNC(TO_DATE('{hasta}','DD-MM-YYYY'),'YYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'I',NVL(a.ptimport,0),0),0)),0)) ISR, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'C',NVL(a.ptimport,0),0),0) * F_TIPO_CAMBIO(A.FEMOVIMI, A.CDAGENTE, 'A', (SELECT MAX(femovimi) FROM tsia_movagen WHERE cdconc = 'SA' AND femovimi < TRUNC(TO_DATE('{hasta}','DD-MM-YYYY'),'YYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'C',NVL(a.ptimport,0),0),0)),0)) IMPUESTO_CEDULAR, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'T',NVL(a.ptimport,0),0),0) * F_TIPO_CAMBIO(A.FEMOVIMI, A.CDAGENTE, 'A', (SELECT MAX(femovimi) FROM tsia_movagen WHERE cdconc = 'SA' AND femovimi < " \
						f"TRUNC(TO_DATE('{hasta}','DD-MM-YYYY'),'YYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'T',NVL(a.ptimport,0),0),0)),0)) HONORARIOS FROM tsia_movagen a, tsia_tipmovi b WHERE a.cdcvemov >= 1000 AND a.cdconc = b.cdconc AND a.cdcvemov = b.cdcvemov AND b.swconinm IN ('C','I','R','V','T') AND a.femovimi BETWEEN (SELECT MAX(femovimi) FROM tsia_movagen WHERE cdconc = 'SA' AND femovimi < TRUNC(TO_DATE('{hasta}','DD-MM-YYYY'),'YYYY')) AND to_date('{hasta}','dd/mm/yyyy') and A.CDAGENTE='{codigo}' GROUP BY a.cdagente )A"
			if tabla == 3:
				query = f"SELECT CONCEPTO,SUM(ROUND(IMPORTE,2)) IMPORTE FROM ( SELECT A.FEMOVIMI, a.cdagente, a.cdmoneda moneda, decode(a.cdconc,'QD',22, 'V',2 , 'Q',22, 'VD',2, decode(a.cdconc, 'I',2, 'ID',2 , 'PC',1999, 'B',2, 'IC',607, decode(b.dsconc,'Imp Cedular Derecho',607,a.cdcvemov) ) ) movimiento, decode(a.cdconc,'QD','IVA Retenido' , 'V','IVA' , 'Q','IVA Retenido', 'VD','IVA', decode(a.cdconc, 'I','ISR', 'ID','ISR' , 'PC','Pago de Comisión', 'IC','Impuesto Cedular', decode(b.dsconc,'Imp Cedular Derecho','Impuesto Cedular', b.dsconc) ) ) concepto, decode(a.cdconc,'QD',2.1 , 'V',2 , 'Q',2.1, 'VD',2, decode(a.cdconc, 'I',4, 'ID',4 , 'PC',1999, 'B',1.2, 'ND',1,decode(b.dsconc,'Imp Cedular Derecho',6, b.numprior) ) ) numprior , decode(a.cdconc,'QD','Q' , 'V','V' , 'Q','Q', 'VD','V', decode(a.cdconc, 'I','I', 'ID','I' ,decode(b.dsconc,'Imp Cedular Derecho','IC', a.cdconc) )) cdconc, SUM(a.ptimport) importe, SUM(a.ptimport) importe2 FROM tsia_movagen a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') " \
						f"AND a.cdcvemov = b.cdcvemov AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND a.cdcvemov < 1000  AND A.CDCONC NOT IN ('PC') and A.CDAGENTE='{codigo}' AND a.cdmoneda='MXP' GROUP BY A.FEMOVIMI, a.cdagente, a.cdmoneda, decode(a.cdconc,'QD',22, 'V',2 , 'Q',22, 'VD',2, decode(a.cdconc, 'I',2, 'ID',2 , 'PC',1999, 'B',2, 'IC',607, decode(b.dsconc,'Imp Cedular Derecho',607,a.cdcvemov) ) ) , decode(a.cdconc,'QD','IVA Retenido' , 'V','IVA' , 'Q','IVA Retenido', 'VD','IVA', decode(a.cdconc, 'I','ISR', 'ID','ISR' , 'PC','Pago de Comisión', 'IC','Impuesto Cedular', decode(b.dsconc,'Imp Cedular Derecho','Impuesto Cedular', b.dsconc) ) ) , decode(a.cdconc,'QD',2.1 , 'V',2 , 'Q',2.1, 'VD',2, decode(a.cdconc, 'I',4, 'ID',4 , 'PC',1999, 'B',1.2, 'ND',1,decode(b.dsconc,'Imp Cedular Derecho',6, b.numprior) ) ) , decode(a.cdconc,'QD','Q' , 'V','V' , 'Q','Q', 'VD','V', decode(a.cdconc, 'I','I', 'ID','I' ,decode(b.dsconc,'Imp Cedular Derecho','IC', a.cdconc) )) UNION SELECT A.FEMOVIMI, a.cdagente, a.cdmoneda moneda, 19 movimiento, 'Subtotal' concepto, 1.9 numprior , " \
						f"'SB' cdconc, SUM ( DECODE (a.cdconc, 'N', a.ptimport, 'ND', a.ptimport, 'GA', a.ptimport, 'B', a.ptimport, 'CD', a.ptimport, 'DC', a.ptimport, 'BC', a.ptimport, 'PI', a.ptimport, decode(b.dsconc,'Descuento de Comisión', a.ptimport, 0))) Importe,0 importe2 FROM tsia_movagen a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND a.cdcvemov < 1000 and A.CDAGENTE='{codigo}' AND a.cdmoneda='MXP' HAVING SUM ( DECODE (a.cdconc, 'N', a.ptimport, 'ND', a.ptimport, 'GA', a.ptimport, 'B', a.ptimport, 'DC', a.ptimport, 'BC', a.ptimport, 'PI', a.ptimport,decode(b.dsconc,'Descuento de Comisión', a.ptimport, 0)) )<> 0 GROUP BY A.FEMOVIMI, a.cdagente, a.cdmoneda UNION SELECT T.FEMOVIMI, T.CDAGENTE, T.MONEDA, T.MOVIMIENTO, T.CONCEPTO, T.NUMPRIOR, T.CDCONC, SUM(T.IMPORTE), SUM(T.SBIMPORTE) FROM (SELECT A.FEMOVIMI, A.CDAGENTE, A.CDMONEDA MONEDA, 1999 MOVIMIENTO, 'Pago de Comisión' CONCEPTO, 1999 numprior, 'PC' CDCONC, A.PTIMPORT IMPORTE, " \
						f"A.PTIMPORT SBIMPorte FROM TSIA_MOVAGEN a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND a.cdcvemov = 600 AND A.CDAGENTE = '{codigo}' AND a.CDMONEDA = 'MXP') T GROUP BY T.FEMOVIMI, T.CDAGENTE, T.MONEDA, T.MOVIMIENTO, T.CONCEPTO, T.NUMPRIOR, T.CDCONC UNION SELECT A.FEMOVIMI, A.CDAGENTE, A.CDMONEDA MONEDA, DECODE(a.cdconc, 'I',2, 'V',2, 'Q',22, 'IC',607, 'PC',1999) MOVIMIENTO, DECODE(a.cdconc, 'I','ISR', 'V','IVA', 'Q','IVA Retenido', 'IC','Impuesto Cedular', 'PC','Pago de Comisión' ) CONCEPTO, DECODE(a.cdconc, 'I',4, 'V',2, 'Q',2.1, 'IC',6, 'PC',1999) numprior, A.CDCONC CDCONC, A.PTIMPORT*-1 IMPORTE, A.PTIMPORT*-1 SBIMPorte FROM TSIA_MOVAGEN a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND a.cdcvemov = 600 AND a.CDCONC IN ('I','V','Q','IC') AND A.CDAGENTE = '{codigo}' " \
						f"AND a.CDMONEDA='MXP' ) GROUP BY NUMPRIOR,MOVIMIENTO,CDCONC,CONCEPTO order by NUMPRIOR,MOVIMIENTO,CDCONC"
			if tabla == 4:
				query = f"SELECT CONCEPTO,SUM(ROUND(IMPORTE,2)) IMPORTE FROM ( SELECT a.cdagente, a.cdmoneda moneda, decode(a.cdconc,'QD',22, 'V',2 , 'Q',22, 'VD',2, decode(a.cdconc, 'I',2, 'ID',2 , 'PC',1999, 'B',2, 'IC',607, decode(b.dsconc,'Imp Cedular Derecho',607,a.cdcvemov) ) ) movimiento, decode(a.cdconc,'QD','IVA Retenido' , 'V','IVA' , 'Q','IVA Retenido', 'VD','IVA', decode(a.cdconc, 'I','ISR', 'ID','ISR' , 'PC','Pago de Comisión', 'IC','Impuesto Cedular', decode(b.dsconc,'Imp Cedular Derecho','Impuesto Cedular', b.dsconc) ) ) concepto, decode(a.cdconc,'QD',2.1 , 'V',2 , 'Q',2.1, 'VD',2, decode(a.cdconc, 'I',4, 'ID',4 , 'PC',1999, 'B',1.2, 'ND',1,decode(b.dsconc,'Imp Cedular Derecho',6, b.numprior) ) ) numprior , decode(a.cdconc,'QD','Q' , 'V','V' , 'Q','Q', 'VD','V', decode(a.cdconc, 'I','I', 'ID','I' ,decode(b.dsconc,'Imp Cedular Derecho','IC', a.cdconc) )) cdconc, SUM(a.ptimport) importe, SUM(a.ptimport) importe2 FROM tsia_movagen a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov " \
						f"AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND a.cdcvemov < 1000 AND A.CDCONC NOT IN ('PC') and A.CDAGENTE='{codigo}' AND a.cdmoneda='USD' GROUP BY a.cdagente, a.cdmoneda, decode(a.cdconc,'QD',22, 'V',2 , 'Q',22, 'VD',2, decode(a.cdconc, 'I',2, 'ID',2 , 'PC',1999, 'B',2, 'IC',607, decode(b.dsconc,'Imp Cedular Derecho',607,a.cdcvemov) ) ) , decode(a.cdconc,'QD','IVA Retenido' , 'V','IVA' , 'Q','IVA Retenido', 'VD','IVA', decode(a.cdconc, 'I','ISR', 'ID','ISR' , 'PC','Pago de Comisión', 'IC','Impuesto Cedular', decode(b.dsconc,'Imp Cedular Derecho','Impuesto Cedular', b.dsconc) ) ) , decode(a.cdconc,'QD',2.1 , 'V',2 , 'Q',2.1, 'VD',2, decode(a.cdconc, 'I',4, 'ID',4 , 'PC',1999, 'B',1.2, 'ND',1,decode(b.dsconc,'Imp Cedular Derecho',6, b.numprior) ) ) , decode(a.cdconc,'QD','Q' , 'V','V' , 'Q','Q', 'VD','V', decode(a.cdconc, 'I','I', 'ID','I' ,decode(b.dsconc,'Imp Cedular Derecho','IC', a.cdconc) )) UNION SELECT a.cdagente, a.cdmoneda moneda, 19 movimiento, 'Subtotal' concepto, 1.9 numprior , 'SB' cdconc, SUM ( DECODE (a.cdconc, 'N', a.ptimport, 'ND', a.ptimport, 'GA', " \
						f"a.ptimport, 'B', a.ptimport, 'CD', a.ptimport, 'DC', a.ptimport, 'BC', a.ptimport, 'PI', a.ptimport, decode(b.dsconc,'Descuento de Comisión', a.ptimport, 0))) Importe,0 importe2 FROM tsia_movagen a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND a.cdcvemov < 1000 and A.CDAGENTE='{codigo}' AND a.cdmoneda='USD' HAVING SUM ( DECODE (a.cdconc, 'N', a.ptimport, 'ND', a.ptimport, 'GA', a.ptimport, 'B', a.ptimport, 'DC', a.ptimport, 'BC', a.ptimport, 'PI', a.ptimport,decode(b.dsconc,'Descuento de Comisión', a.ptimport, 0)) )<> 0 GROUP BY a.cdagente, a.cdmoneda UNION SELECT T.CDAGENTE, T.MONEDA, T.MOVIMIENTO, T.CONCEPTO, T.NUMPRIOR, T.CDCONC, SUM(T.IMPORTE), SUM(T.SBIMPORTE) FROM (SELECT A.CDAGENTE, A.CDMONEDA MONEDA, 1999 MOVIMIENTO, 'Pago de Comisión' CONCEPTO, 1999 numprior, 'PC' CDCONC, A.PTIMPORT IMPORTE, A.PTIMPORT SBIMPorte FROM TSIA_MOVAGEN a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') " \
						f"AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND a.cdcvemov = 600 AND A.CDAGENTE = '{codigo}' AND a.CDMONEDA = 'USD') T GROUP BY T.CDAGENTE, T.MONEDA, T.MOVIMIENTO, T.CONCEPTO, T.NUMPRIOR, T.CDCONC UNION SELECT A.CDAGENTE, A.CDMONEDA MONEDA, DECODE(a.cdconc, 'I',2, 'V',2, 'Q',22, 'IC',607, 'PC',1999) MOVIMIENTO, DECODE(a.cdconc, 'I','ISR', 'V','IVA', 'Q','IVA Retenido', 'IC','Impuesto Cedular', 'PC','Pago de Comisión' ) CONCEPTO, DECODE(a.cdconc, 'I',4, 'V',2, 'Q',2.1, 'IC',6, 'PC',1999) numprior, A.CDCONC CDCONC, A.PTIMPORT*-1 IMPORTE, A.PTIMPORT*-1 SBIMPorte FROM TSIA_MOVAGEN a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND a.cdcvemov = 600 AND a.CDCONC IN ('I','V','Q','IC') AND A.CDAGENTE = '{codigo}' AND a.CDMONEDA='USD' ) GROUP BY NUMPRIOR,MOVIMIENTO,CDCONC,CONCEPTO order by NUMPRIOR,MOVIMIENTO,CDCONC"
			if tabla == 5:
				query = f"SELECT decode(cdtipram,3,'VIDA','DAÑO') TIPO, (SELECT OTVALOR05 FROM TVALOPOL A WHERE A.CDUNIECO=MAIN.CDUNIECO AND A.CDRAMO=MAIN.CDRAMO AND A.ESTADO='M' AND A.NMPOLIZA=MAIN.NMPOLIZA AND A.NMSUPLEM=(SELECT MAX(NMSUPLEM) FROM TVALOPOL B WHERE A.CDUNIECO=B.CDUNIECO AND A.CDRAMO=B.CDRAMO AND A.ESTADO=B.ESTADO AND A.NMPOLIZA=B.NMPOLIZA)) grupo, MAIN.CDUNIECO OFICINA, MAIN.CDRAMO RAMO, MAIN.NMPOLIZA POLIZA, nomtomador(MAIN.CDUNIECO, MAIN.CDRAMO, 'M', MAIN.NMPOLIZA, 9999999999999999999) CONTRATANTE, MAIN.CDAGENTE, MAIN.CDMONEDA, MAIN.NMRECIBO RECIBO, MAIN.SERIERECIBO SERIE_REC, nvl(MAIN.PRIMA_TOTAL,0) PRIMA_TOTAL, nvl(MAIN.PrimaNeta,0) PRIMA_NETA, nvl(MAIN.PORAGEN,0) PORC_COMISION, nvl(MAIN.COM_DERECHO,0) COMISION_DERECHO, nvl(MAIN.COM_NETA,0) COMISION_NETA, (nvl(MAIN.ComisionNeta,0) + nvl(MAIN.IvaComision,0) + nvl(MAIN.IvaRetenido,0)+ nvl(MAIN.IsrComision,0) + nvl(MAIN.ComisionDerecho,0) + nvl(MAIN.IsrComDerecho,0) + nvl(MAIN.IvaRetDerecho,0) + nvl(MAIN.IvaDerecho,0)) TOTAL_COMISION, MAIN.NUMPRELI,MAIN.NMTRANSF, TO_CHAR(MAIN.FEMOVIMI,'DD/MM/YYYY') FECHACORTE " \
						f"FROM ( SELECT G.NUMPRELI,G.NMTRANSF,A.CDMONEDA,A.CDAGENTE,c.cdtipram,A.FEMOVIMI, A.CDUNIECO,A.CDRAMO,A.NMPOLIZA,A.NMRECIBO,NVL(A.SERIERECIBO,0) SERIERECIBO, A.FEINICIOREC, A.FEMOVIMI FEAPLIC,A.DSNOMASEG, NVL(A.PMATOTAL,0) PRIMA_TOTAL,a.PMANETAP PrimaNeta, a.IMPORAGTN COM_NETA,a.PORAGEN, NVL(decode(a.SWAPDER,'A',IMPORAGTND,NULL),0) COM_DERECHO, /*SUMATORIAS*/ ROUND(NVL(a.imporagtn,0),2) ComisionNeta, ROUND(NVL(a.imporagtv,0),2) IvaComision, ROUND(NVL(a.imporagtq,0),2) IvaRetenido, ROUND(NVL(a.imporagti,0),2) IsrComision, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtnd,0),0),2) ComisionDerecho, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtid,0),0),2) IsrComDerecho, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtqd,0),0),2) IvaRetDerecho, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtvd,0),0),2) IvaDerecho from tsia_detcom a, TRAMOS C,MINTPSOFT G where TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdramo = c.cdramo AND c.cdtipram IN (1,2) AND A.CDMONEDA='MXP' AND A.CDAGENTE='{codigo}' AND A.NMPOLIZA=G.NMPOLIZA (+) " \
						f"AND A.NMRECIBO=G.NMRECIBO (+) ) MAIN ORDER BY FEMOVIMI, CDUNIECO,CDRAMO,NMPOLIZA,NMRECIBO"
			if tabla == 6:
				query = f"SELECT decode(cdtipram,3,'VIDA','DAÑO') TIPO, (SELECT OTVALOR05 FROM TVALOPOL A WHERE A.CDUNIECO=MAIN.CDUNIECO AND A.CDRAMO=MAIN.CDRAMO AND A.ESTADO='M' AND A.NMPOLIZA=MAIN.NMPOLIZA AND A.NMSUPLEM=(SELECT MAX(NMSUPLEM) FROM TVALOPOL B WHERE A.CDUNIECO=B.CDUNIECO AND A.CDRAMO=B.CDRAMO AND A.ESTADO=B.ESTADO AND A.NMPOLIZA=B.NMPOLIZA)) grupo, MAIN.CDUNIECO OFICINA, MAIN.CDRAMO RAMO, MAIN.NMPOLIZA POLIZA, nomtomador(MAIN.CDUNIECO, MAIN.CDRAMO, 'M', MAIN.NMPOLIZA, 9999999999999999999) CONTRATANTE, MAIN.CDAGENTE, MAIN.CDMONEDA, MAIN.NMRECIBO RECIBO, MAIN.SERIERECIBO SERIE_REC, nvl(MAIN.PRIMA_TOTAL,0) PRIMA_TOTAL, nvl(MAIN.PrimaNeta,0) PRIMA_NETA, nvl(MAIN.PORAGEN,0) PORC_COMISION, nvl(MAIN.COM_DERECHO,0) COMISION_DERECHO, nvl(MAIN.COM_NETA,0) COMISION_NETA, (nvl(MAIN.ComisionNeta,0) + nvl(MAIN.IvaComision,0) + nvl(MAIN.IvaRetenido,0)+ nvl(MAIN.IsrComision,0) + nvl(MAIN.ComisionDerecho,0) + nvl(MAIN.IsrComDerecho,0) + nvl(MAIN.IvaRetDerecho,0) + nvl(MAIN.IvaDerecho,0)) TOTAL_COMISION, MAIN.NUMPRELI,MAIN.NMTRANSF, TO_CHAR(MAIN.FEMOVIMI,'DD/MM/YYYY') FECHACORTE " \
						f"FROM ( SELECT G.NUMPRELI,G.NMTRANSF,A.CDMONEDA,A.CDAGENTE,c.cdtipram,A.FEMOVIMI, A.CDUNIECO,A.CDRAMO,A.NMPOLIZA,A.NMRECIBO,NVL(A.SERIERECIBO,0) SERIERECIBO, A.FEINICIOREC, A.FEMOVIMI FEAPLIC,A.DSNOMASEG, NVL(A.PMATOTAL,0) PRIMA_TOTAL,a.PMANETAP PrimaNeta, a.IMPORAGTN COM_NETA,a.PORAGEN, NVL(decode(a.SWAPDER,'A',IMPORAGTND,NULL),0) COM_DERECHO, /*SUMATORIAS*/ ROUND(NVL(a.imporagtn,0),2) ComisionNeta, ROUND(NVL(a.imporagtv,0),2) IvaComision, ROUND(NVL(a.imporagtq,0),2) IvaRetenido, ROUND(NVL(a.imporagti,0),2) IsrComision, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtnd,0),0),2) ComisionDerecho, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtid,0),0),2) IsrComDerecho, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtqd,0),0),2) IvaRetDerecho, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtvd,0),0),2) IvaDerecho from tsia_detcom a, TRAMOS C,MINTPSOFT G where TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdramo = c.cdramo AND c.cdtipram IN (1,2) AND A.CDMONEDA='USD' AND A.CDAGENTE='{codigo}' AND A.NMPOLIZA=G.NMPOLIZA (+) " \
						f"AND A.NMRECIBO=G.NMRECIBO (+) ) MAIN ORDER BY FEMOVIMI, CDUNIECO,CDRAMO,NMPOLIZA,NMRECIBO"
			if tabla == 7:
				query = f"SELECT decode(cdtipram,3,'VIDA','DAÑO') TIPO, (SELECT OTVALOR05 FROM TVALOPOL A WHERE A.CDUNIECO=MAIN.CDUNIECO AND A.CDRAMO=MAIN.CDRAMO AND A.ESTADO='M' AND A.NMPOLIZA=MAIN.NMPOLIZA AND A.NMSUPLEM=(SELECT MAX(NMSUPLEM) FROM TVALOPOL B WHERE A.CDUNIECO=B.CDUNIECO AND A.CDRAMO=B.CDRAMO AND A.ESTADO=B.ESTADO AND A.NMPOLIZA=B.NMPOLIZA)) grupo, MAIN.CDUNIECO OFICINA, MAIN.CDRAMO RAMO, MAIN.NMPOLIZA POLIZA, nomtomador(MAIN.CDUNIECO, MAIN.CDRAMO, 'M', MAIN.NMPOLIZA, 9999999999999999999) CONTRATANTE, MAIN.CDAGENTE, MAIN.CDMONEDA, MAIN.NMRECIBO RECIBO, MAIN.SERIERECIBO SERIE_REC, nvl(MAIN.PRIMA_TOTAL,0) PRIMA_TOTAL, nvl(MAIN.PrimaNeta,0) PRIMA_NETA, nvl(MAIN.PORAGEN,0) PORC_COMISION, nvl(MAIN.COM_DERECHO,0) COMISION_DERECHO, nvl(MAIN.COM_NETA,0) COMISION_NETA, (nvl(MAIN.ComisionNeta,0) + nvl(MAIN.IvaComision,0) + nvl(MAIN.IvaRetenido,0)+ nvl(MAIN.IsrComision,0) + nvl(MAIN.ComisionDerecho,0) + nvl(MAIN.IsrComDerecho,0) + nvl(MAIN.IvaRetDerecho,0) + nvl(MAIN.IvaDerecho,0)) TOTAL_COMISION, MAIN.NUMPRELI,MAIN.NMTRANSF, TO_CHAR(MAIN.FEMOVIMI,'DD/MM/YYYY') FECHACORTE " \
						f"FROM ( SELECT G.NUMPRELI,G.NMTRANSF,A.CDMONEDA,A.CDAGENTE,c.cdtipram,A.FEMOVIMI, A.CDUNIECO,A.CDRAMO,A.NMPOLIZA,A.NMRECIBO,NVL(A.SERIERECIBO,0) SERIERECIBO, A.FEINICIOREC, A.FEMOVIMI FEAPLIC,A.DSNOMASEG, NVL(A.PMATOTAL,0) PRIMA_TOTAL,a.PMANETAP PrimaNeta, a.IMPORAGTN COM_NETA,a.PORAGEN, NVL(decode(a.SWAPDER,'A',IMPORAGTND,NULL),0) COM_DERECHO, /*SUMATORIAS*/ ROUND(NVL(a.imporagtn,0),2) ComisionNeta, ROUND(NVL(a.imporagtv,0),2) IvaComision, ROUND(NVL(a.imporagtq,0),2) IvaRetenido, ROUND(NVL(a.imporagti,0),2) IsrComision, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtnd,0),0),2) ComisionDerecho, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtid,0),0),2) IsrComDerecho, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtqd,0),0),2) IvaRetDerecho, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtvd,0),0),2) IvaDerecho from tsia_detcom a, TRAMOS C,MINTPSOFT G where TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdramo = c.cdramo AND c.cdtipram IN (3) AND A.CDMONEDA='MXP' AND A.CDAGENTE='{codigo}' AND A.NMPOLIZA=G.NMPOLIZA (+) " \
						f"AND A.NMRECIBO=G.NMRECIBO (+) ) MAIN ORDER BY FEMOVIMI, CDUNIECO,CDRAMO,NMPOLIZA,NMRECIBO"
			if tabla == 8:
				query = f"SELECT decode(cdtipram,3,'VIDA','DAÑO') TIPO, (SELECT OTVALOR05 FROM TVALOPOL A WHERE A.CDUNIECO=MAIN.CDUNIECO AND A.CDRAMO=MAIN.CDRAMO AND A.ESTADO='M' AND A.NMPOLIZA=MAIN.NMPOLIZA AND A.NMSUPLEM=(SELECT MAX(NMSUPLEM) FROM TVALOPOL B WHERE A.CDUNIECO=B.CDUNIECO AND A.CDRAMO=B.CDRAMO AND A.ESTADO=B.ESTADO AND A.NMPOLIZA=B.NMPOLIZA)) grupo, MAIN.CDUNIECO OFICINA, MAIN.CDRAMO RAMO, MAIN.NMPOLIZA POLIZA, nomtomador(MAIN.CDUNIECO, MAIN.CDRAMO, 'M', MAIN.NMPOLIZA, 9999999999999999999) CONTRATANTE, MAIN.CDAGENTE, MAIN.CDMONEDA, MAIN.NMRECIBO RECIBO, MAIN.SERIERECIBO SERIE_REC, nvl(MAIN.PRIMA_TOTAL,0) PRIMA_TOTAL, nvl(MAIN.PrimaNeta,0) PRIMA_NETA, nvl(MAIN.PORAGEN,0) PORC_COMISION, nvl(MAIN.COM_DERECHO,0) COMISION_DERECHO, nvl(MAIN.COM_NETA,0) COMISION_NETA, (nvl(MAIN.ComisionNeta,0) + nvl(MAIN.IvaComision,0) + nvl(MAIN.IvaRetenido,0)+ nvl(MAIN.IsrComision,0) + nvl(MAIN.ComisionDerecho,0) + nvl(MAIN.IsrComDerecho,0) + nvl(MAIN.IvaRetDerecho,0) + nvl(MAIN.IvaDerecho,0)) TOTAL_COMISION, MAIN.NUMPRELI,MAIN.NMTRANSF, TO_CHAR(MAIN.FEMOVIMI,'DD/MM/YYYY') FECHACORTE " \
						f"FROM ( SELECT G.NUMPRELI,G.NMTRANSF,A.CDMONEDA,A.CDAGENTE,c.cdtipram,A.FEMOVIMI, A.CDUNIECO,A.CDRAMO,A.NMPOLIZA,A.NMRECIBO,NVL(A.SERIERECIBO,0) SERIERECIBO, A.FEINICIOREC, A.FEMOVIMI FEAPLIC,A.DSNOMASEG, NVL(A.PMATOTAL,0) PRIMA_TOTAL,a.PMANETAP PrimaNeta, a.IMPORAGTN COM_NETA,a.PORAGEN, NVL(decode(a.SWAPDER,'A',IMPORAGTND,NULL),0) COM_DERECHO, /*SUMATORIAS*/ ROUND(NVL(a.imporagtn,0),2) ComisionNeta, ROUND(NVL(a.imporagtv,0),2) IvaComision, ROUND(NVL(a.imporagtq,0),2) IvaRetenido, ROUND(NVL(a.imporagti,0),2) IsrComision, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtnd,0),0),2) ComisionDerecho, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtid,0),0),2) IsrComDerecho, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtqd,0),0),2) IvaRetDerecho, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtvd,0),0),2) IvaDerecho from tsia_detcom a, TRAMOS C,MINTPSOFT G where TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdramo = c.cdramo AND c.cdtipram IN (3) AND A.CDMONEDA='USD' AND A.CDAGENTE='{codigo}' AND A.NMPOLIZA=G.NMPOLIZA (+) " \
						f"AND A.NMRECIBO=G.NMRECIBO (+) ) MAIN ORDER BY FEMOVIMI, CDUNIECO,CDRAMO,NMPOLIZA,NMRECIBO"
			if tabla == 9:
				query = f"SELECT CONCEPTO,SUM(ROUND(nvl(IMPORTE,0),2)) IMPORTE FROM ( SELECT a.cdagente, a.cdmoneda moneda, decode(a.cdconc,'QD',22, 'V',2 , 'Q',22, 'VD',2, decode(a.cdconc, 'I',2, 'ID',2 , 'PC',1999, 'B',2, 'IC',607, decode(b.dsconc,'Imp Cedular Derecho',607,a.cdcvemov) ) ) movimiento, decode(a.cdconc,'QD','IVA Retenido' , 'V','IVA' , 'Q','IVA Retenido', 'VD','IVA', decode(a.cdconc, 'I','ISR', 'ID','ISR' , 'PC','Pago de Comisión', 'IC','Impuesto Cedular', decode(b.dsconc,'Imp Cedular Derecho','Impuesto Cedular', b.dsconc) ) ) concepto, decode(a.cdconc,'QD',2.1 , 'V',2 , 'Q',2.1, 'VD',2, decode(a.cdconc, 'I',4, 'ID',4 , 'PC',1999, 'B',1.2, 'ND',1,decode(b.dsconc,'Imp Cedular Derecho',6, b.numprior) ) ) numprior , decode(a.cdconc,'QD','Q' , 'V','V' , 'Q','Q', 'VD','V', decode(a.cdconc, 'I','I', 'ID','I' ,decode(b.dsconc,'Imp Cedular Derecho','IC', a.cdconc) )) cdconc, SUM(a.ptimport) importe, SUM(a.ptimport) importe2 FROM tsia_movagen a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov " \
						f"AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND A.CDCONC NOT IN ('PC') AND a.cdcvemov > 1000 and A.CDAGENTE='{codigo}' AND a.cdmoneda='MXP' GROUP BY a.cdagente, a.cdmoneda, decode(a.cdconc,'QD',22, 'V',2 , 'Q',22, 'VD',2, decode(a.cdconc, 'I',2, 'ID',2 , 'PC',1999, 'B',2, 'IC',607, decode(b.dsconc,'Imp Cedular Derecho',607,a.cdcvemov) ) ) , decode(a.cdconc,'QD','IVA Retenido' , 'V','IVA' , 'Q','IVA Retenido', 'VD','IVA', decode(a.cdconc, 'I','ISR', 'ID','ISR' , 'PC','Pago de Comisión', 'IC','Impuesto Cedular', decode(b.dsconc,'Imp Cedular Derecho','Impuesto Cedular', b.dsconc) ) ) , decode(a.cdconc,'QD',2.1 , 'V',2 , 'Q',2.1, 'VD',2, decode(a.cdconc, 'I',4, 'ID',4 , 'PC',1999, 'B',1.2, 'ND',1,decode(b.dsconc,'Imp Cedular Derecho',6, b.numprior) ) ) , decode(a.cdconc,'QD','Q' , 'V','V' , 'Q','Q', 'VD','V', decode(a.cdconc, 'I','I', 'ID','I' ,decode(b.dsconc,'Imp Cedular Derecho','IC', a.cdconc) )) UNION SELECT a.cdagente, a.cdmoneda moneda, 19 movimiento, 'Subtotal' concepto, 1.9 numprior , 'SB' cdconc, SUM ( DECODE (a.cdconc, 'N', a.ptimport, 'ND', a.ptimport, 'GA', a.ptimport, " \
						f"'B', a.ptimport, 'CD', a.ptimport, 'DC', a.ptimport, 'BC', a.ptimport, 'PI', a.ptimport, decode(b.dsconc,'Descuento de Comisión', a.ptimport, 0))) Importe,0 importe2 FROM tsia_movagen a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND a.cdcvemov > 1000 and A.CDAGENTE='{codigo}' AND a.cdmoneda='MXP' HAVING SUM ( DECODE (a.cdconc, 'N', a.ptimport, 'ND', a.ptimport, 'GA', a.ptimport, 'B', a.ptimport, 'DC', a.ptimport, 'BC', a.ptimport, 'PI', a.ptimport,decode(b.dsconc,'Descuento de Comisión', a.ptimport, 0)) )<> 0 GROUP BY a.cdagente, a.cdmoneda UNION SELECT T.CDAGENTE, T.MONEDA, T.MOVIMIENTO, T.CONCEPTO, T.NUMPRIOR, T.CDCONC, SUM(T.IMPORTE), SUM(T.SBIMPORTE) FROM (SELECT A.CDAGENTE, A.CDMONEDA MONEDA, 1999 MOVIMIENTO, 'Pago de Comisión' CONCEPTO, 1999 numprior, 'PC' CDCONC, A.PTIMPORT IMPORTE, A.PTIMPORT SBIMPorte FROM TSIA_MOVAGEN a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') " \
						f"AND a.cdcvemov = b.cdcvemov AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND a.cdcvemov = 1600 AND A.CDAGENTE = '{codigo}' AND a.CDMONEDA = 'MXP') T GROUP BY T.CDAGENTE, T.MONEDA, T.MOVIMIENTO, T.CONCEPTO, T.NUMPRIOR, T.CDCONC UNION SELECT A.CDAGENTE, A.CDMONEDA MONEDA, DECODE(a.cdconc, 'I',2, 'V',2, 'Q',22, 'IC',607, 'PC',1999) MOVIMIENTO, DECODE(a.cdconc, 'I','ISR', 'V','IVA', 'Q','IVA Retenido', 'IC','Impuesto Cedular', 'PC','Pago de Comisión' ) CONCEPTO, DECODE(a.cdconc, 'I',4, 'V',2, 'Q',2.1, 'IC',6, 'PC',1999) numprior, A.CDCONC CDCONC, A.PTIMPORT*-1 IMPORTE, A.PTIMPORT*-1 SBIMPorte FROM TSIA_MOVAGEN a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND a.cdcvemov = 1600 AND a.CDCONC IN ('I','V','Q','IC') AND A.CDAGENTE = '{codigo}' AND a.CDMONEDA='MXP' ) GROUP BY NUMPRIOR,MOVIMIENTO,CDCONC,CONCEPTO order by NUMPRIOR,MOVIMIENTO,CDCONC"
			if tabla == 10:
				query = f"SELECT CONCEPTO,SUM(ROUND(nvl(IMPORTE,0),2)) IMPORTE FROM ( SELECT a.cdagente, a.cdmoneda moneda, decode(a.cdconc,'QD',22, 'V',2 , 'Q',22, 'VD',2, decode(a.cdconc, 'I',2, 'ID',2 , 'PC',1999, 'B',2, 'IC',607, decode(b.dsconc,'Imp Cedular Derecho',607,a.cdcvemov) ) ) movimiento, decode(a.cdconc,'QD','IVA Retenido' , 'V','IVA' , 'Q','IVA Retenido', 'VD','IVA', decode(a.cdconc, 'I','ISR', 'ID','ISR' , 'PC','Pago de Comisión', 'IC','Impuesto Cedular', decode(b.dsconc,'Imp Cedular Derecho','Impuesto Cedular', b.dsconc) ) ) concepto, decode(a.cdconc,'QD',2.1 , 'V',2 , 'Q',2.1, 'VD',2, decode(a.cdconc, 'I',4, 'ID',4 , 'PC',1999, 'B',1.2, 'ND',1,decode(b.dsconc,'Imp Cedular Derecho',6, b.numprior) ) ) numprior , decode(a.cdconc,'QD','Q' , 'V','V' , 'Q','Q', 'VD','V', decode(a.cdconc, 'I','I', 'ID','I' ,decode(b.dsconc,'Imp Cedular Derecho','IC', a.cdconc) )) cdconc, SUM(a.ptimport) importe, SUM(a.ptimport) importe2 FROM tsia_movagen a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov " \
						f"AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND A.CDCONC NOT IN ('PC') AND a.cdcvemov > 1000 and A.CDAGENTE='{codigo}' AND a.cdmoneda='USD' GROUP BY a.cdagente, a.cdmoneda, decode(a.cdconc,'QD',22, 'V',2 , 'Q',22, 'VD',2, decode(a.cdconc, 'I',2, 'ID',2 , 'PC',1999, 'B',2, 'IC',607, decode(b.dsconc,'Imp Cedular Derecho',607,a.cdcvemov) ) ) , decode(a.cdconc,'QD','IVA Retenido' , 'V','IVA' , 'Q','IVA Retenido', 'VD','IVA', decode(a.cdconc, 'I','ISR', 'ID','ISR' , 'PC','Pago de Comisión', 'IC','Impuesto Cedular', decode(b.dsconc,'Imp Cedular Derecho','Impuesto Cedular', b.dsconc) ) ) , decode(a.cdconc,'QD',2.1 , 'V',2 , 'Q',2.1, 'VD',2, decode(a.cdconc, 'I',4, 'ID',4 , 'PC',1999, 'B',1.2, 'ND',1,decode(b.dsconc,'Imp Cedular Derecho',6, b.numprior) ) ) , decode(a.cdconc,'QD','Q' , 'V','V' , 'Q','Q', 'VD','V', decode(a.cdconc, 'I','I', 'ID','I' ,decode(b.dsconc,'Imp Cedular Derecho','IC', a.cdconc) )) UNION SELECT a.cdagente, a.cdmoneda moneda, 19 movimiento, 'Subtotal' concepto, 1.9 numprior , 'SB' cdconc, SUM ( DECODE (a.cdconc, 'N', a.ptimport, 'ND', a.ptimport, 'GA'," \
						f" a.ptimport, 'B', a.ptimport, 'CD', a.ptimport, 'DC', a.ptimport, 'BC', a.ptimport, 'PI', a.ptimport, decode(b.dsconc,'Descuento de Comisión', a.ptimport, 0))) Importe,0 importe2 FROM tsia_movagen a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND a.cdcvemov > 1000 and A.CDAGENTE='{codigo}' AND a.cdmoneda='USD' HAVING SUM ( DECODE (a.cdconc, 'N', a.ptimport, 'ND', a.ptimport, 'GA', a.ptimport, 'B', a.ptimport, 'DC', a.ptimport, 'BC', a.ptimport, 'PI', a.ptimport,decode(b.dsconc,'Descuento de Comisión', a.ptimport, 0)) )<> 0 GROUP BY a.cdagente, a.cdmoneda UNION SELECT T.CDAGENTE, T.MONEDA, T.MOVIMIENTO, T.CONCEPTO, T.NUMPRIOR, T.CDCONC, SUM(T.IMPORTE), SUM(T.SBIMPORTE) FROM (SELECT A.CDAGENTE, A.CDMONEDA MONEDA, 1999 MOVIMIENTO, 'Pago de Comisión' CONCEPTO, 1999 numprior, 'PC' CDCONC, A.PTIMPORT IMPORTE, A.PTIMPORT SBIMPorte FROM TSIA_MOVAGEN a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') " \
						f"AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND a.cdcvemov = 1600 AND A.CDAGENTE = '{codigo}' AND a.CDMONEDA = 'USD') T GROUP BY T.CDAGENTE, T.MONEDA, T.MOVIMIENTO, T.CONCEPTO, T.NUMPRIOR, T.CDCONC UNION SELECT A.CDAGENTE, A.CDMONEDA MONEDA, DECODE(a.cdconc, 'I',2, 'V',2, 'Q',22, 'IC',607, 'PC',1999) MOVIMIENTO, DECODE(a.cdconc, 'I','ISR', 'V','IVA', 'Q','IVA Retenido', 'IC','Impuesto Cedular', 'PC','Pago de Comisión' ) CONCEPTO, DECODE(a.cdconc, 'I',4, 'V',2, 'Q',2.1, 'IC',6, 'PC',1999) numprior, A.CDCONC CDCONC, A.PTIMPORT*-1 IMPORTE, A.PTIMPORT*-1 SBIMPorte FROM TSIA_MOVAGEN a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND a.cdcvemov = 1600 AND a.CDCONC IN ('I','V','Q','IC') AND A.CDAGENTE = '{codigo}' AND a.CDMONEDA='USD' ) GROUP BY NUMPRIOR,MOVIMIENTO,CDCONC,CONCEPTO order by NUMPRIOR,MOVIMIENTO,CDCONC"
			if tabla == 11:
				query = f"SELECT T.FECHA_MOVIMIENTO FECHA_MOVIMIENTO , SUM(T.IMPORTE_DANOS) IMPORTE_DANOS , SUM(T.IMPORTE_VIDA) IMPORTE_VIDA , SUM(T.TOTAL) TOTAL , T.FECHA_PAGO FECHA_PAGO , T.NUM_COMPROBANTE NUM_COMPROBANTE , SUM(T.IMPORTE_PAG_DANOS) IMPORTE_PAG_DANOS, SUM(T.IMPORTE_PAG_VIDA) IMPORTE_PAG_VIDA , SUM(T.TOTAL_PAGADO) TOTAL_PAGADO FROM ( SELECT TO_CHAR(A.FEMOVIMI,'DD/MM/YYYY') FECHA_MOVIMIENTO , DECODE(A.CDCVEMOV,600,A.PTIMPORT,0) IMPORTE_DANOS , DECODE(A.CDCVEMOV,1600,A.PTIMPORT,0) IMPORTE_VIDA , A.PTIMPORT TOTAL , DECODE(NVL(B.CDCOMPRO,0),0,'',TO_CHAR(B.FECHAENV,'DD/MM/YYYY')) FECHA_PAGO , DECODE(NVL(B.CDCOMPRO,0),0,'',DECODE(B.CDMETPAG,'EFT',B.NMTRANSF,'CHK',B.NMCHEQUE,B.CDCOMPRO)) NUM_COMPROBANTE , DECODE(NVL(B.CDCOMPRO,0),0,0,DECODE(B.CDCVEMOV,600,B.PTIMPPAG,0)) IMPORTE_PAG_DANOS, DECODE(NVL(B.CDCOMPRO,0),0,0,DECODE(B.CDCVEMOV,1600,B.PTIMPPAG,0)) IMPORTE_PAG_VIDA , DECODE(NVL(B.CDCOMPRO,0),0,0,B.PTIMPPAG) TOTAL_PAGADO , A.CDAGENTE AGENTE , B.NUMPRELI NUMPRELI FROM TSIA_MOVAGEN A, MINTPSOFT B WHERE A.CDCVEMOV IN (600,1600) AND TRUNC(A.FEMOVIMI) " \
						f"BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND A.NUMPRELIQ = B.NUMPRELI(+) AND A.CDCVEMOV = B.CDCVEMOV(+) AND A.CDCONC = B.CDCONC (+) AND A.CDAGENTE = '{codigo}' AND A.CDMONEDA = 'MXP') T GROUP BY T.AGENTE,T.NUMPRELI,T.FECHA_MOVIMIENTO,T.FECHA_PAGO,T.NUM_COMPROBANTE ORDER BY T.FECHA_MOVIMIENTO ASC"
			if tabla == 12:
				query = f"SELECT T.FECHA_MOVIMIENTO FECHA_MOVIMIENTO , SUM(T.IMPORTE_DANOS) IMPORTE_DANOS , SUM(T.IMPORTE_VIDA) IMPORTE_VIDA , SUM(T.TOTAL) TOTAL , T.FECHA_PAGO FECHA_PAGO , T.NUM_COMPROBANTE NUM_COMPROBANTE , SUM(T.IMPORTE_PAG_DANOS) IMPORTE_PAG_DANOS, SUM(T.IMPORTE_PAG_VIDA) IMPORTE_PAG_VIDA , SUM(T.TOTAL_PAGADO) TOTAL_PAGADO FROM ( SELECT TO_CHAR(A.FEMOVIMI,'DD/MM/YYYY') FECHA_MOVIMIENTO , DECODE(A.CDCVEMOV,600,A.PTIMPORT,0) IMPORTE_DANOS , DECODE(A.CDCVEMOV,1600,A.PTIMPORT,0) IMPORTE_VIDA , A.PTIMPORT TOTAL , DECODE(NVL(B.CDCOMPRO,0),0,'',TO_CHAR(B.FECHAENV,'DD/MM/YYYY')) FECHA_PAGO , DECODE(NVL(B.CDCOMPRO,0),0,'',DECODE(B.CDMETPAG,'EFT',B.NMTRANSF,'CHK',B.NMCHEQUE,B.CDCOMPRO)) NUM_COMPROBANTE , DECODE(NVL(B.CDCOMPRO,0),0,0,DECODE(B.CDCVEMOV,600,B.PTIMPPAG,0)) IMPORTE_PAG_DANOS, DECODE(NVL(B.CDCOMPRO,0),0,0,DECODE(B.CDCVEMOV,1600,B.PTIMPPAG,0)) IMPORTE_PAG_VIDA , DECODE(NVL(B.CDCOMPRO,0),0,0,B.PTIMPPAG) TOTAL_PAGADO , A.CDAGENTE AGENTE , B.NUMPRELI NUMPRELI FROM TSIA_MOVAGEN A, MINTPSOFT B WHERE A.CDCVEMOV IN (600,1600) AND TRUNC(A.FEMOVIMI) " \
						f"BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND A.NUMPRELIQ = B.NUMPRELI(+) AND A.CDCVEMOV = B.CDCVEMOV(+) AND A.CDCONC = B.CDCONC (+) AND A.CDAGENTE = '{codigo}' AND A.CDMONEDA = 'USD') T GROUP BY T.AGENTE,T.NUMPRELI,T.FECHA_MOVIMIENTO,T.FECHA_PAGO,T.NUM_COMPROBANTE ORDER BY T.FECHA_MOVIMIENTO ASC"
	if clave == 'P':
		if tipo == 'BONO':
			if tabla == 0:
				query = getheaderquery('P',codigo,desde,hasta)
			if tabla == 1:
				query = f"SELECT MAIN.NMBONO,MAIN.DSBONO,MAIN.CDTIPBON, MAIN.CDGRUPO,MAIN.CDUNIECO, MAIN.CDRAMO,MAIN.NMPOLIZA, nomtomador(MAIN.CDUNIECO, MAIN.CDRAMO, 'M', MAIN.NMPOLIZA, 9999999999999999999) CONTRATANTE, MAIN.cdagente,(SELECT OTVALOR07 FROM TVALOPOL Z WHERE Z.CDUNIECO=MAIN.CDUNIECO AND Z.CDRAMO=MAIN.CDRAMO AND Z.NMPOLIZA=MAIN.NMPOLIZA AND Z.ESTADO='M' AND Z.NMSUPLEM=(SELECT MAX(NMSUPLEM) FROM TVALOPOL Y WHERE Y.CDUNIECO=Z.CDUNIECO AND Z.CDRAMO=Y.CDRAMO AND Z.NMPOLIZA=Y.NMPOLIZA AND Z.ESTADO=Y.ESTADO)) COMPUTABILIDAD, MAIN.TIPOCAMB, MAIN.NMRECIBO,MAIN.SERIERECIBO,NVL(MAIN.PMATOTAL,0) PRIMA_TOTAL,NVL(MAIN.PMANETAP,0) PRIMA_NETA,MAIN.PORCEPAG, ABS(MAIN.imporpron),(ABS(MAIN.imporpron)+ABS(MAIN.imporproi)) TOTAL_PAGADO ,MAIN.NUMPRELIQ, DECODE(NVL(MAIN.CDCOMPRO,0),0,'',DECODE(MAIN.CDMETPAG,'EFT',MAIN.NMTRANSF,'CHK',MAIN.NMCHEQUE,MAIN.CDCOMPRO)) NUM_COMPROBANTE, TO_CHAR(MAIN.FEMOVIMI,'DD/MM/YYYY') FECHA, (DECODE ((SELECT DISTINCT 'SI' " \
						f"FROM bon_trecexc WHERE nmbono = MAIN.nmbono AND cdrecpag = MAIN.cdrecpag AND cdagente = MAIN.cdagente AND cdtipmov = MAIN.cdtipmov),'SI', 'SI', 'NO' )) EXCLUDO, (SELECT DECODE(swexcep, '1', 'Por agente para Rank', '2', 'Por agente para Pago', '3', 'Por agente para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 1) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov AND ROWNUM < 2) POR_AGENTE, (SELECT DECODE(swexcep, '1', 'Por promotor para Rank', '2', 'Por promotor para Pago', '3', 'Por promotor para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 2) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov " \
						f"AND ROWNUM < 2) POR_PROMOTOR, (SELECT DECODE(swexcep, '1', 'Por ramo para Rank', '2', 'Por ramo para Pago', '3', 'Por ramo para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 4) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov AND ROWNUM < 2) POR_RAMO, (SELECT DECODE(swexcep, '1', 'Por oficina para Rank', '2', 'Por oficina para Pago', '3', 'Por oficina para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 5) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov) POR_OFICINA, (SELECT DECODE(swexcep,'1', 'Por poliza para Rank', '2', 'Por poliza para Pago', '3', 'Por poliza para Todo') FROM bon_trecexc aa, (SELECT codigo, " \
						f"descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 6) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov AND ROWNUM < 2) POR_POLIZA, (SELECT DECODE(swexcep, '1', 'Por grupo para Rank', '2', 'Por grupo para Pago', '3', 'Por grupo para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 7) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov AND ROWNUM < 2) POR_GRUPO, (SELECT DECODE(swexcep, '1', 'Por subgrupo para Rank', '2', 'Por subgrupo para Pago', '3', 'Por subgrupo para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 8) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND " \
						f"aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov) POR_SUBGRUPO, (SELECT DECODE(swexcep, '1', 'Por tipo de persona para Rank', '2', 'Por tipo de persona para Pago', '3', 'Por tipo de persona para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 9) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov AND ROWNUM < 2) POR_TIP_PERS, (SELECT DECODE(swexcep, '1', 'Por Reaseguro para Rank', '2', 'Por Reaseguro para Pago', '3', 'Por Reaseguro para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 10) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov AND ROWNUM < 2) POR_REASEGURO, (SELECT DECODE(swexcep, '1', 'Por Coaseguro para Rank', '2', " \
						f"'Por Coaseguro para Pago', '3', 'Por Coaseguro para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 11) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov AND ROWNUM < 2) POR_COASEGURO, (SELECT DECODE(swexcep, '1', 'Por multianual para Rank', '2', 'Por multianual para Pago', '3', 'Por multianual para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 12) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov AND ROWNUM < 2) POR_MULTIANUAL, (SELECT DECODE(swexcep, '1', 'Por dividendos para Rank', '2', 'Por dividendos para Pago', '3', 'Por dividendos para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' " \
						f"AND codigo = 13) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov AND ROWNUM < 2) POR_DIVIDENDOS, (SELECT DECODE(swexcep, '1', 'Por no computable para Rank', '2', 'Por no computable para Pago', '3', 'Por no computable para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 14) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov AND ROWNUM < 2) POR_NO_COMPU, (SELECT DECODE(swexcep, '1', 'Por portafolio para Rank', '2', 'Por portafolio para Pago', '3', 'Por portafolio para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 15) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente " \
						f"AND aa.cdtipmov = MAIN.cdtipmov AND ROWNUM < 2) POR_PORTAFOLIO, (SELECT DECODE(swexcep, '1', 'Por prestador para Rank', '2', 'Por prestador para Pago', '3', 'Por prestador para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 17) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov AND ROWNUM < 2) POR_PRESTADOR, (SELECT DECODE(swexcep, '1', 'Por negocio para Rank', '2', 'Por negocio para Pago', '3', 'Por negocio para Todo') FROM bon_trecexc aa, (SELECT codigo, descripl FROM tmanteni WHERE cdtabla = 'CATEXCBON' AND codigo = 18) bb WHERE aa.nmbono = MAIN.nmbono AND aa.cdexcep = bb.codigo AND aa.cdrecpag = MAIN.cdrecpag AND aa.cdagente = MAIN.cdagente AND aa.cdtipmov = MAIN.cdtipmov AND ROWNUM < 2) POR_NEGOCIO FROM (SELECT E.TIPOCAMB, C.NMBONO,D.DSBONO,C.CDTIPBON, E.CDGRUPO,E.CDUNIECO, E.CDRAMO,E.NMPOLIZA,F.CDAGENTE," \
						f"E.NMRECIBO,F.SERIERECIBO,G.NMTRANSF,G.CDMETPAG,G.NMCHEQUE,G.CDCOMPRO, F.PMATOTAL,F.PMANETAP,C.PORCEPAG,F.imporpron,F.imporproi,A.NUMPRELIQ,A.FEMOVIMI,E.CDTIPMOV,E.CDRECPAG FROM TSIA_MOVPROM A , BON_TPAGLIQ B, BON_TCONFBON C, BON_TCATBONO D, VBON_DESGLOSE E, tsia_detcom F, MINTPSOFT G WHERE A.cdpromot ='{codigo}' AND A.FEMOVIMI BETWEEN TO_DATE('{desde}','dd/mm/yyyy') AND TO_DATE('{hasta}','dd/mm/yyyy') AND A.CDCONC IN ('B','BC') AND B.NUMPRELIQ =A.NUMPRELIQ AND B.CDAGPROU = A.cdpromot AND B.CDCVEMOV = A.CDCVEMOV AND B.CDCONC = A.CDCONC AND C.NMBONO = B.NMBONO AND D.CDBONO = C.CDBONO AND E.NMBONO= C.NMBONO AND A.cdpromot= E.cdpromot AND E.NMPOLIZA= F.NMPOLIZA AND E.CDUNIECO= F.CDUNIECO AND E.CDRAMO= F.CDRAMO AND E.NMRECIBO= F.NMRECIBO AND A.cdpromot= F.cdpromot AND B.NUMPRELIQ =G.NUMPRELI (+) AND A.CDCVEMOV = B.CDCVEMOV(+) AND A.CDCONC = B.CDCONC (+) GROUP BY E.TIPOCAMB,C.NMBONO,D.DSBONO, E.CDGRUPO, E.CDUNIECO, E.CDRAMO,E.NMPOLIZA,F.CDAGENTE, F.SERIERECIBO,G.NMTRANSF,C.CDTIPBON, " \
						f"F.PMATOTAL,F.PMANETAP,G.CDMETPAG,G.NMCHEQUE,G.CDCOMPRO, C.PORCEPAG,F.imporpron, F.imporproi,E.CDTIPMOV,E.CDRECPAG, A.NUMPRELIQ,E.NMRECIBO,A.FEMOVIMI ORDER BY FEMOVIMI ASC) MAIN"
		if tipo == 'COMISION':
			if tabla == 0:
				query = getheaderquery('P',codigo,desde,hasta)
			if tabla == 1:
				query = f"SELECT TIPO, ROUND(HONORARIOS,2) base, round (IVA,2) iva , round (HONORARIOS + IVA,2) Subtotal, round (IVA_RETENIDO,2) IVA_RETENIDO, round (ISR,2) ISR, round (IMPUESTO_CEDULAR,2) IMPUESTO_CEDULAR, round (HONORARIOS + IVA+ IVA_RETENIDO + ISR + IMPUESTO_CEDULAR,2) total FROM( SELECT a.cdpromot, 'DAÑOS' tipo, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'V',NVL(a.ptimport,0),0),0) * (SELECT ptcambio FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) = (SELECT MAX(TRUNC(fevalor)) FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) BETWEEN TO_DATE('01'||LPAD(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),2,0)||to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY'),'DDMMYYYY') AND TO_DATE('01'||LPAD(DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM') + 1),2,0)||DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY') + 1,to_char(TO_DATE('{hasta}'," \
						f"'DD/MM/YYYY'),'YYYY')),'DDMMYYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'V',NVL(a.ptimport,0),0),0)),0)) IVA, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'R',NVL(a.ptimport,0),0),0) * (SELECT ptcambio FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) = (SELECT MAX(TRUNC(fevalor)) FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) BETWEEN TO_DATE('01'||LPAD(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),2,0)||to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY'),'DDMMYYYY') AND TO_DATE('01'||LPAD(DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM') + 1),2,0)||DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY') + 1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY')),'DDMMYYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'R',NVL(a.ptimport,0),0),0)),0)) IVA_RETENIDO, (NVL(SUM(DECODE(a.cdmoneda,'USD'," \
						f"DECODE(b.swconinm,'I',NVL(a.ptimport,0),0),0) * (SELECT ptcambio FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) = (SELECT MAX(TRUNC(fevalor)) FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) BETWEEN TO_DATE('01'||LPAD(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),2,0)||to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY'),'DDMMYYYY') AND TO_DATE('01'||LPAD(DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM') + 1),2,0)||DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY') + 1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY')),'DDMMYYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'I',NVL(a.ptimport,0),0),0)),0)) ISR, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'C',NVL(a.ptimport,0),0),0) * (SELECT ptcambio FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) = (SELECT MAX(TRUNC(fevalor)) FROM " \
						f"tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) BETWEEN TO_DATE('01'||LPAD(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),2,0)||to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY'),'DDMMYYYY') AND TO_DATE('01'||LPAD(DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM') + 1),2,0)||DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY') + 1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY')),'DDMMYYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'C',NVL(a.ptimport,0),0),0)),0)) IMPUESTO_CEDULAR, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'T',NVL(a.ptimport,0),0),0) * (SELECT ptcambio FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) = (SELECT MAX(TRUNC(fevalor)) FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) BETWEEN TO_DATE('01'||LPAD(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),2,0)||to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY')," \
						f"'DDMMYYYY') AND TO_DATE('01'||LPAD(DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM') + 1),2,0)||DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY') + 1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY')),'DDMMYYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'T',NVL(a.ptimport,0),0),0)),0)) HONORARIOS FROM tsia_movprom a, tsia_tipmovi b WHERE a.cdcvemov < 1000 AND a.cdconc = b.cdconc AND a.cdcvemov = b.cdcvemov AND b.swconinm IN ('C','I','R','V','T') AND TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND A.CDPROMOT='{codigo}' GROUP BY a.cdpromot UNION ALL SELECT a.cdpromot, 'VIDA' tipo, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'V',NVL(a.ptimport,0),0),0) * (SELECT ptcambio FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) = (SELECT MAX(TRUNC(fevalor)) FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) " \
						f"BETWEEN TO_DATE('01'||LPAD(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),2,0)||to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY'),'DDMMYYYY') AND TO_DATE('01'||LPAD(DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM') + 1),2,0)||DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY') + 1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY')),'DDMMYYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'V',NVL(a.ptimport,0),0),0)),0)) IVA, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'R',NVL(a.ptimport,0),0),0) * (SELECT ptcambio FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) = (SELECT MAX(TRUNC(fevalor)) FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) BETWEEN TO_DATE('01'||LPAD(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),2,0)||to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY'),'DDMMYYYY') AND TO_DATE('01'||LPAD(DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,1," \
						f"to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM') + 1),2,0)||DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY') + 1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY')),'DDMMYYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'R',NVL(a.ptimport,0),0),0)),0)) IVA_RETENIDO, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'I',NVL(a.ptimport,0),0),0) * (SELECT ptcambio FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) = (SELECT MAX(TRUNC(fevalor)) FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) BETWEEN TO_DATE('01'||LPAD(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),2,0)||to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY'),'DDMMYYYY') AND TO_DATE('01'||LPAD(DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM') + 1),2,0)||DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY') + 1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY'))," \
						f"'DDMMYYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'I',NVL(a.ptimport,0),0),0)),0)) ISR, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'C',NVL(a.ptimport,0),0),0) * (SELECT ptcambio FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) = (SELECT MAX(TRUNC(fevalor)) FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) BETWEEN TO_DATE('01'||LPAD(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),2,0)||to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY'),'DDMMYYYY') AND TO_DATE('01'||LPAD(DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM') + 1),2,0)||DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY') + 1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY')),'DDMMYYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'C',NVL(a.ptimport,0),0),0)),0)) IMPUESTO_CEDULAR, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'T',NVL(a.ptimport,0),0),0) * (SELECT ptcambio FROM tcambios " \
						f"WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) = (SELECT MAX(TRUNC(fevalor)) FROM tcambios WHERE cdmoneda = 'MXP' AND cdmonbas = 'USD' AND TRUNC(fevalor) BETWEEN TO_DATE('01'||LPAD(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),2,0)||to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY'),'DDMMYYYY') AND TO_DATE('01'||LPAD(DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM') + 1),2,0)||DECODE(to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'MM'),12,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY') + 1,to_char(TO_DATE('{hasta}','DD/MM/YYYY'),'YYYY')),'DDMMYYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'T',NVL(a.ptimport,0),0),0)),0)) HONORARIOS FROM tsia_movprom a, tsia_tipmovi b WHERE a.cdcvemov >= 1000 AND a.cdconc = b.cdconc AND a.cdcvemov = b.cdcvemov AND b.swconinm IN ('C','I','R','V','T') AND TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND A.CDPROMOT='{codigo}' GROUP BY a.cdpromot)"
			if tabla == 2:
				query = f"SELECT TIPO, ROUND(HONORARIOS,2) base, round (IVA,2) iva , round (HONORARIOS + IVA,2) Subtotal, round (IVA_RETENIDO,2) IVA_RETENIDO, round (ISR,2) ISR, round (IMPUESTO_CEDULAR,2) IMPUESTO_CEDULAR, round (HONORARIOS + IVA+ IVA_RETENIDO + ISR + IMPUESTO_CEDULAR,2) total FROM(SELECT a.cdpromot, 'DAÑOS' tipo, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'V',NVL(a.ptimport,0),0),0) * F_TIPO_CAMBIO(A.FEMOVIMI, A.CDPROMOT, 'P', (SELECT MAX(femovimi) FROM tsia_movprom WHERE cdconc = 'SA' AND femovimi < TRUNC(TO_DATE('{hasta}','DD-MM-YYYY'),'YYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'V',NVL(a.ptimport,0),0),0)),0)) IVA, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'R',NVL(a.ptimport,0),0),0) * F_TIPO_CAMBIO(A.FEMOVIMI, A.CDPROMOT, 'P', (SELECT MAX(femovimi) FROM tsia_movprom WHERE cdconc = 'SA' AND femovimi < TRUNC(TO_DATE('{hasta}','DD-MM-YYYY'),'YYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'R',NVL(a.ptimport,0),0),0)),0)) IVA_RETENIDO, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'I'," \
						f"NVL(a.ptimport,0),0),0) * F_TIPO_CAMBIO(A.FEMOVIMI, A.CDPROMOT, 'P', (SELECT MAX(femovimi) FROM tsia_movprom WHERE cdconc = 'SA' AND femovimi < TRUNC(TO_DATE('{hasta}','DD-MM-YYYY'),'YYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'I',NVL(a.ptimport,0),0),0)),0)) ISR, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'C',NVL(a.ptimport,0),0),0) * F_TIPO_CAMBIO(A.FEMOVIMI, A.CDPROMOT, 'P', (SELECT MAX(femovimi) FROM tsia_movprom WHERE cdconc = 'SA' AND femovimi < TRUNC(TO_DATE('{hasta}','DD-MM-YYYY'),'YYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'C',NVL(a.ptimport,0),0),0)),0)) IMPUESTO_CEDULAR, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'T',NVL(a.ptimport,0),0),0) * F_TIPO_CAMBIO(A.FEMOVIMI, A.CDPROMOT, 'P', (SELECT MAX(femovimi) FROM tsia_movprom WHERE cdconc = 'SA' AND femovimi < TRUNC(TO_DATE('{hasta}','DD-MM-YYYY'),'YYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'T',NVL(a.ptimport,0),0),0)),0)) HONORARIOS FROM tsia_movprom a, tsia_tipmovi b WHERE a.cdcvemov < 1000 AND a.cdconc = b.cdconc " \
						f"AND a.cdcvemov = b.cdcvemov AND b.swconinm IN ('C','I','R','V','T') AND TRUNC(a.femovimi) BETWEEN (SELECT MAX(femovimi) FROM tsia_movprom WHERE cdconc = 'SA' AND femovimi < TRUNC(TO_DATE('{hasta}','DD-MM-YYYY'),'YYYY')) AND to_date('{hasta}','dd/mm/yyyy') AND A.CDPROMOT='{codigo}' GROUP BY a.cdpromot UNION ALL SELECT a.cdpromot, 'VIDA' tipo, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'V',NVL(a.ptimport,0),0),0) * F_TIPO_CAMBIO(A.FEMOVIMI, A.CDPROMOT, 'P', (SELECT MAX(femovimi) FROM tsia_movprom WHERE cdconc = 'SA' AND femovimi < TRUNC(TO_DATE('{hasta}','DD-MM-YYYY'),'YYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'V',NVL(a.ptimport,0),0),0)),0)) IVA, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'R',NVL(a.ptimport,0),0),0) * F_TIPO_CAMBIO(A.FEMOVIMI, A.CDPROMOT, 'P', (SELECT MAX(femovimi) FROM tsia_movprom WHERE cdconc = 'SA' AND femovimi < TRUNC(TO_DATE('{hasta}','DD-MM-YYYY'),'YYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'R',NVL(a.ptimport,0),0),0)),0)) IVA_RETENIDO, (NVL(SUM(DECODE(a.cdmoneda,'USD'," \
						f"DECODE(b.swconinm,'I',NVL(a.ptimport,0),0),0) * F_TIPO_CAMBIO(A.FEMOVIMI, A.CDPROMOT, 'P', (SELECT MAX(femovimi) FROM tsia_movprom WHERE cdconc = 'SA' AND femovimi < TRUNC(TO_DATE('{hasta}','DD-MM-YYYY'),'YYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'I',NVL(a.ptimport,0),0),0)),0)) ISR, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'C',NVL(a.ptimport,0),0),0) * F_TIPO_CAMBIO(A.FEMOVIMI, A.CDPROMOT, 'P', (SELECT MAX(femovimi) FROM tsia_movprom WHERE cdconc = 'SA' AND femovimi < TRUNC(TO_DATE('{hasta}','DD-MM-YYYY'),'YYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'C',NVL(a.ptimport,0),0),0)),0)) IMPUESTO_CEDULAR, (NVL(SUM(DECODE(a.cdmoneda,'USD',DECODE(b.swconinm,'T',NVL(a.ptimport,0),0),0) * F_TIPO_CAMBIO(A.FEMOVIMI, A.CDPROMOT, 'P', (SELECT MAX(femovimi) FROM tsia_movprom WHERE cdconc = 'SA' AND femovimi < TRUNC(TO_DATE('{hasta}','DD-MM-YYYY'),'YYYY')))),0) + NVL(SUM(DECODE(a.cdmoneda,'MXP',DECODE(b.swconinm,'T',NVL(a.ptimport,0),0),0)),0)) HONORARIOS FROM tsia_movprom a, tsia_tipmovi b WHERE a.cdcvemov >= 1000 AND " \
						f"a.cdconc = b.cdconc AND a.cdcvemov = b.cdcvemov AND b.swconinm IN ('C','I','R','V','T') AND TRUNC(a.femovimi) BETWEEN (SELECT MAX(femovimi) FROM tsia_movprom WHERE cdconc = 'SA' AND femovimi < TRUNC(TO_DATE('{hasta}','DD-MM-YYYY'),'YYYY')) AND to_date('{hasta}','dd/mm/yyyy') AND A.CDPROMOT='{codigo}' GROUP BY a.cdpromot )"
			if tabla == 3:
				query = f"select CONCEPTO,SUM(ROUND(IMPORTE,2)) IMPORTE from(SELECT a.cdpromot, a.cdmoneda moneda, decode(a.cdconc,'QD',22, 'V',2 , 'Q',22, 'VD',2, decode(a.cdconc, 'I',2, 'ID',2 , 'PC',1999, 'B',2, 'IC',607, decode(b.dsconc,'Imp Cedular Derecho',607,a.cdcvemov) ) ) movimiento, decode(a.cdconc,'QD','IVA Retenido' , 'V','IVA' , 'Q','IVA Retenido', 'VD','IVA', decode(a.cdconc, 'I','ISR', 'ID','ISR' , 'PC','Pago de Comisión', 'IC','Impuesto Cedular', decode(b.dsconc,'Imp Cedular Derecho','Impuesto Cedular', b.dsconc) ) ) concepto, decode(a.cdconc,'QD',2.1 , 'V',2 , 'Q',2.1, 'VD',2, decode(a.cdconc, 'I',4, 'ID',4 , 'PC',1999, 'B',1.2, 'ND',1,decode(b.dsconc,'Imp Cedular Derecho',6, b.numprior) ) ) numprior , decode(a.cdconc,'QD','Q' , 'V','V' , 'Q','Q', 'VD','V', decode(a.cdconc, 'I','I', 'ID','I' ,decode(b.dsconc,'Imp Cedular Derecho','IC', a.cdconc) )) cdconc, SUM(a.ptimport) importe, SUM(a.ptimport) SBimporte FROM tsia_movprom a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov " \
						f"AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND a.cdcvemov < 1000 AND A.CDCONC NOT IN ('PC') AND A.CDPROMOT='{codigo}' AND a.CDMONEDA='MXP' GROUP BY a.cdpromot, a.cdmoneda,decode(a.cdconc,'QD',22, 'V',2 , 'Q',22, 'VD',2, decode(a.cdconc, 'I',2, 'ID',2 , 'PC',1999, 'B',2, 'IC',607, decode(b.dsconc,'Imp Cedular Derecho',607,a.cdcvemov) ) ) , decode(a.cdconc,'QD','IVA Retenido' , 'V','IVA' , 'Q','IVA Retenido', 'VD','IVA', decode(a.cdconc, 'I','ISR', 'ID','ISR' , 'PC','Pago de Comisión', 'IC','Impuesto Cedular', decode(b.dsconc,'Imp Cedular Derecho','Impuesto Cedular', b.dsconc) ) ) , decode(a.cdconc,'QD',2.1 , 'V',2 , 'Q',2.1, 'VD',2, decode(a.cdconc, 'I',4, 'ID',4 , 'PC',1999, 'B',1.2, 'ND',1,decode(b.dsconc,'Imp Cedular Derecho',6, b.numprior) ) ) , decode(a.cdconc,'QD','Q' , 'V','V' , 'Q','Q', 'VD','V', decode(a.cdconc, 'I','I', 'ID','I' ,decode(b.dsconc,'Imp Cedular Derecho','IC', a.cdconc) )) union SELECT a.cdpromot, a.cdmoneda moneda, 19 movimiento, 'Subtotal' concepto, 1.9 numprior , 'ST' cdconc, SUM ( DECODE (a.cdconc, 'N', a.ptimport, 'ND', a.ptimport, 'GA', " \
						f"a.ptimport, 'B', a.ptimport, 'CD', a.ptimport, 'DC', a.ptimport, 'BC', a.ptimport, 'PI', a.ptimport, decode(b.dsconc,'Descuento de Comisión', a.ptimport, 0)))Importe, 0 SBIMPorte FROM tsia_movprom a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND a.cdcvemov < 1000 /* Para daños*/ AND A.CDPROMOT = '{codigo}' AND a.CDMONEDA='MXP' HAVING SUM ( DECODE (a.cdconc, 'N', a.ptimport, 'ND', a.ptimport, 'GA', a.ptimport, 'B', a.ptimport, 'DC', a.ptimport, 'BC', a.ptimport, 'PI', a.ptimport,decode(b.dsconc,'Descuento de Comisión', a.ptimport, 0)) )<> 0 GROUP BY a.cdpromot, a.cdmoneda UNION SELECT T.CDPROMOT, T.MONEDA, T.MOVIMIENTO, T.CONCEPTO, T.NUMPRIOR, T.CDCONC, SUM(T.IMPORTE), SUM(T.SBIMPORTE) FROM (SELECT A.CDPROMOT, A.CDMONEDA MONEDA, 1999 MOVIMIENTO, 'Pago de Comisión' CONCEPTO, 1999 numprior, 'PC' CDCONC, A.PTIMPORT IMPORTE, A.PTIMPORT SBIMPorte FROM tsia_movprom a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') " \
						f"AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND a.cdcvemov = 600 AND A.CDPROMOT = '{codigo}' AND a.CDMONEDA = 'MXP') T GROUP BY T.CDPROMOT, T.MONEDA, T.MOVIMIENTO, T.CONCEPTO, T.NUMPRIOR, T.CDCONC UNION SELECT A.CDPROMOT, A.CDMONEDA MONEDA, DECODE(a.cdconc, 'I',2, 'V',2, 'Q',22, 'IC',607, 'PC',1999) MOVIMIENTO, DECODE(a.cdconc, 'I','ISR', 'V','IVA', 'Q','IVA Retenido', 'IC','Impuesto Cedular', 'PC','Pago de Comisión' ) CONCEPTO, DECODE(a.cdconc, 'I',4, 'V',2, 'Q',2.1, 'IC',6, 'PC',1999) numprior, A.CDCONC CDCONC, A.PTIMPORT*-1 IMPORTE, A.PTIMPORT*-1 SBIMPorte FROM tsia_movprom a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND a.cdcvemov = 600 AND a.CDCONC IN ('I','V','Q','IC') AND A.CDPROMOT = '{codigo}' AND a.CDMONEDA='MXP' ) GROUP BY NUMPRIOR,MOVIMIENTO,CDCONC,CONCEPTO order by NUMPRIOR,MOVIMIENTO,CDCONC"
			if tabla == 4:
				query = f"select CONCEPTO,SUM(ROUND(IMPORTE,2)) IMPORTE from(SELECT a.cdpromot, a.cdmoneda moneda, decode(a.cdconc,'QD',22, 'V',2 , 'Q',22, 'VD',2, decode(a.cdconc, 'I',2, 'ID',2 , 'PC',1999, 'B',2, 'IC',607, decode(b.dsconc,'Imp Cedular Derecho',607,a.cdcvemov) ) ) movimiento, decode(a.cdconc,'QD','IVA Retenido' , 'V','IVA' , 'Q','IVA Retenido', 'VD','IVA', decode(a.cdconc, 'I','ISR', 'ID','ISR' , 'PC','Pago de Comisión', 'IC','Impuesto Cedular', decode(b.dsconc,'Imp Cedular Derecho','Impuesto Cedular', b.dsconc) ) ) concepto, decode(a.cdconc,'QD',2.1 , 'V',2 , 'Q',2.1, 'VD',2, decode(a.cdconc, 'I',4, 'ID',4 , 'PC',1999, 'B',1.2, 'ND',1,decode(b.dsconc,'Imp Cedular Derecho',6, b.numprior) ) ) numprior , decode(a.cdconc,'QD','Q' , 'V','V' , 'Q','Q', 'VD','V', decode(a.cdconc, 'I','I', 'ID','I' ,decode(b.dsconc,'Imp Cedular Derecho','IC', a.cdconc) )) cdconc, SUM(a.ptimport) importe, SUM(a.ptimport) SBimporte FROM tsia_movprom a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov AND a.cdconc = b.cdconc AND " \
						f"b.afedocta = 'S' AND a.cdcvemov < 1000 AND A.CDCONC NOT IN ('PC') AND A.CDPROMOT='{codigo}' AND a.CDMONEDA='USD' GROUP BY a.cdpromot, a.cdmoneda,decode(a.cdconc,'QD',22, 'V',2 , 'Q',22, 'VD',2, decode(a.cdconc, 'I',2, 'ID',2 , 'PC',1999, 'B',2, 'IC',607, decode(b.dsconc,'Imp Cedular Derecho',607,a.cdcvemov) ) ) , decode(a.cdconc,'QD','IVA Retenido' , 'V','IVA' , 'Q','IVA Retenido', 'VD','IVA', decode(a.cdconc, 'I','ISR', 'ID','ISR' , 'PC','Pago de Comisión', 'IC','Impuesto Cedular', decode(b.dsconc,'Imp Cedular Derecho','Impuesto Cedular', b.dsconc) ) ) , decode(a.cdconc,'QD',2.1 , 'V',2 , 'Q',2.1, 'VD',2, decode(a.cdconc, 'I',4, 'ID',4 , 'PC',1999, 'B',1.2, 'ND',1,decode(b.dsconc,'Imp Cedular Derecho',6, b.numprior) ) ) , decode(a.cdconc,'QD','Q' , 'V','V' , 'Q','Q', 'VD','V', decode(a.cdconc, 'I','I', 'ID','I' ,decode(b.dsconc,'Imp Cedular Derecho','IC', a.cdconc) )) union SELECT a.cdpromot, a.cdmoneda moneda, 19 movimiento, 'Subtotal' concepto, 1.9 numprior , 'ST' cdconc, SUM ( DECODE (a.cdconc, 'N', a.ptimport, 'ND', a.ptimport, 'GA', a.ptimport, 'B', a.ptimport, 'CD', a.ptimport, 'DC'," \
						f" a.ptimport, 'BC', a.ptimport, 'PI', a.ptimport, decode(b.dsconc,'Descuento de Comisión', a.ptimport, 0)))Importe, 0 SBIMPorte FROM tsia_movprom a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND a.cdcvemov < 1000 /* Para daños*/ AND A.CDPROMOT = '{clave}' AND a.CDMONEDA='USD' HAVING SUM ( DECODE (a.cdconc, 'N', a.ptimport, 'ND', a.ptimport, 'GA', a.ptimport, 'B', a.ptimport, 'DC', a.ptimport, 'BC', a.ptimport, 'PI', a.ptimport,decode(b.dsconc,'Descuento de Comisión', a.ptimport, 0)) )<> 0 GROUP BY a.cdpromot, a.cdmoneda UNION SELECT T.CDPROMOT, T.MONEDA, T.MOVIMIENTO, T.CONCEPTO, T.NUMPRIOR, T.CDCONC, SUM(T.IMPORTE), SUM(T.SBIMPORTE) FROM (SELECT A.CDPROMOT, A.CDMONEDA MONEDA, 1999 MOVIMIENTO, 'Pago de Comisión' CONCEPTO, 1999 numprior, 'PC' CDCONC, A.PTIMPORT IMPORTE, A.PTIMPORT SBIMPorte FROM tsia_movprom a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov " \
						f"AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND a.cdcvemov = 600 AND A.CDPROMOT = '{codigo}' AND a.CDMONEDA = 'USD') T GROUP BY T.CDPROMOT, T.MONEDA, T.MOVIMIENTO, T.CONCEPTO, T.NUMPRIOR, T.CDCONC UNION SELECT A.CDPROMOT, A.CDMONEDA MONEDA, DECODE(a.cdconc, 'I',2, 'V',2, 'Q',22, 'IC',607, 'PC',1999) MOVIMIENTO, DECODE(a.cdconc, 'I','ISR', 'V','IVA', 'Q','IVA Retenido', 'IC','Impuesto Cedular', 'PC','Pago de Comisión' ) CONCEPTO, DECODE(a.cdconc, 'I',4, 'V',2, 'Q',2.1, 'IC',6, 'PC',1999) numprior, A.CDCONC CDCONC, A.PTIMPORT*-1 IMPORTE, A.PTIMPORT*-1 SBIMPorte FROM tsia_movprom a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND a.cdcvemov = 600 AND a.CDCONC IN ('I','V','Q','IC') AND A.CDPROMOT = '{codigo}' AND a.CDMONEDA='USD' ) GROUP BY NUMPRIOR,MOVIMIENTO,CDCONC,CONCEPTO order by NUMPRIOR,MOVIMIENTO,CDCONC"
			if tabla == 5:
				query = f"SELECT decode(cdtipram,3,'VIDA','DAÑO') TIPO, (SELECT OTVALOR05 FROM TVALOPOL A WHERE A.CDUNIECO=MAIN.CDUNIECO AND A.CDRAMO=MAIN.CDRAMO AND A.ESTADO='M' AND A.NMPOLIZA=MAIN.NMPOLIZA AND A.NMSUPLEM=(SELECT MAX(NMSUPLEM) FROM TVALOPOL B WHERE A.CDUNIECO=B.CDUNIECO AND A.CDRAMO=B.CDRAMO AND A.ESTADO=B.ESTADO AND A.NMPOLIZA=B.NMPOLIZA)) grupo, MAIN.CDUNIECO OFICINA, MAIN.CDRAMO RAMO, MAIN.NMPOLIZA POLIZA, nomtomador(MAIN.CDUNIECO, MAIN.CDRAMO, 'M', MAIN.NMPOLIZA, 9999999999999999999) CONTRATANTE, MAIN.CDAGENTE, MAIN.CDMONEDA, MAIN.NMRECIBO RECIBO, MAIN.SERIERECIBO SERIE_REC, nvl(MAIN.PRIMA_TOTAL,0) PRIMA_TOTAL, nvl(MAIN.PrimaNeta,0) PRIMA_NETA, nvl(MAIN.PORAGEN,0) PORC_COMISION, nvl(MAIN.COM_DERECHO,0) COMISION_DERECHO, nvl(MAIN.COM_NETA,0) COMISION_NETA, (nvl(MAIN.ComisionNeta,0) + nvl(MAIN.IvaComision,0) + nvl(MAIN.IvaRetenido,0)+ nvl(MAIN.IsrComision,0) + nvl(MAIN.ComisionDerecho,0) + nvl(MAIN.IsrComDerecho,0) + nvl(MAIN.IvaRetDerecho,0) + nvl(MAIN.IvaDerecho,0)) TOTAL_COMISION, MAIN.NUMPRELI,MAIN.NMTRANSF, TO_CHAR(MAIN.FEMOVIMI,'DD/MM/YYYY') " \
						f"FECHACORTE FROM ( SELECT G.NUMPRELI,G.NMTRANSF,A.CDMONEDA,A.CDAGENTE,c.cdtipram,A.FEMOVIMI, A.CDUNIECO,A.CDRAMO,A.NMPOLIZA,A.NMRECIBO,NVL(A.SERIERECIBO,0) SERIERECIBO, A.FEINICIOREC, A.FEMOVIMI FEAPLIC,A.DSNOMASEG, NVL(A.PMATOTAL,0) PRIMA_TOTAL,a.PMANETAP PrimaNeta, a.IMPORAGTN COM_NETA,a.PORAGEN, NVL(decode(a.SWAPDER,'A',IMPORAGTND,NULL),0) COM_DERECHO,ROUND(NVL(a.imporagtn,0),2) ComisionNeta, ROUND(NVL(a.imporagtv,0),2) IvaComision, ROUND(NVL(a.imporagtq,0),2) IvaRetenido, ROUND(NVL(a.imporagti,0),2) IsrComision, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtnd,0),0),2) ComisionDerecho, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtid,0),0),2) IsrComDerecho, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtqd,0),0),2) IvaRetDerecho, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtvd,0),0),2) IvaDerecho from tsia_detcom a, TRAMOS C,MINTPSOFT G where TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdramo = c.cdramo AND c.cdtipram IN (1,2) AND A.CDMONEDA='MXP' AND A.CDPROMOT='{codigo}' AND A.NMPOLIZA=G.NMPOLIZA (+) AND " \
						f"A.NMRECIBO=G.NMRECIBO (+) ) MAIN ORDER BY FEMOVIMI, CDUNIECO,CDRAMO,NMPOLIZA,NMRECIBO"
			if tabla == 6:
				query = f"SELECT decode(cdtipram,3,'VIDA','DAÑO') TIPO, (SELECT OTVALOR05 FROM TVALOPOL A WHERE A.CDUNIECO=MAIN.CDUNIECO AND A.CDRAMO=MAIN.CDRAMO AND A.ESTADO='M' AND A.NMPOLIZA=MAIN.NMPOLIZA AND A.NMSUPLEM=(SELECT MAX(NMSUPLEM) FROM TVALOPOL B WHERE A.CDUNIECO=B.CDUNIECO AND A.CDRAMO=B.CDRAMO AND A.ESTADO=B.ESTADO AND A.NMPOLIZA=B.NMPOLIZA)) grupo, MAIN.CDUNIECO OFICINA, MAIN.CDRAMO RAMO, MAIN.NMPOLIZA POLIZA, nomtomador(MAIN.CDUNIECO, MAIN.CDRAMO, 'M', MAIN.NMPOLIZA, 9999999999999999999) CONTRATANTE, MAIN.CDAGENTE, MAIN.CDMONEDA, MAIN.NMRECIBO RECIBO, MAIN.SERIERECIBO SERIE_REC, nvl(MAIN.PRIMA_TOTAL,0) PRIMA_TOTAL, nvl(MAIN.PrimaNeta,0) PRIMA_NETA, nvl(MAIN.PORAGEN,0) PORC_COMISION, nvl(MAIN.COM_DERECHO,0) COMISION_DERECHO, nvl(MAIN.COM_NETA,0) COMISION_NETA, (nvl(MAIN.ComisionNeta,0) + nvl(MAIN.IvaComision,0) + nvl(MAIN.IvaRetenido,0)+ nvl(MAIN.IsrComision,0) + nvl(MAIN.ComisionDerecho,0) + nvl(MAIN.IsrComDerecho,0) + nvl(MAIN.IvaRetDerecho,0) + nvl(MAIN.IvaDerecho,0)) TOTAL_COMISION, MAIN.NUMPRELI,MAIN.NMTRANSF, TO_CHAR(MAIN.FEMOVIMI,'DD/MM/YYYY') " \
						f"FECHACORTE FROM ( SELECT G.NUMPRELI,G.NMTRANSF,A.CDMONEDA,A.CDAGENTE,c.cdtipram,A.FEMOVIMI, A.CDUNIECO,A.CDRAMO,A.NMPOLIZA,A.NMRECIBO,NVL(A.SERIERECIBO,0) SERIERECIBO, A.FEINICIOREC, A.FEMOVIMI FEAPLIC,A.DSNOMASEG, NVL(A.PMATOTAL,0) PRIMA_TOTAL,a.PMANETAP PrimaNeta, a.IMPORAGTN COM_NETA,a.PORAGEN, NVL(decode(a.SWAPDER,'A',IMPORAGTND,NULL),0) COM_DERECHO,ROUND(NVL(a.imporagtn,0),2) ComisionNeta, ROUND(NVL(a.imporagtv,0),2) IvaComision, ROUND(NVL(a.imporagtq,0),2) IvaRetenido, ROUND(NVL(a.imporagti,0),2) IsrComision, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtnd,0),0),2) ComisionDerecho, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtid,0),0),2) IsrComDerecho, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtqd,0),0),2) IvaRetDerecho, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtvd,0),0),2) IvaDerecho from tsia_detcom a, TRAMOS C,MINTPSOFT G where TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdramo = c.cdramo AND c.cdtipram IN (1,2) AND A.CDMONEDA='USD' AND A.CDPROMOT='{codigo}' AND A.NMPOLIZA=G.NMPOLIZA (+) AND " \
						f"A.NMRECIBO=G.NMRECIBO (+) ) MAIN ORDER BY FEMOVIMI, CDUNIECO,CDRAMO,NMPOLIZA,NMRECIBO"
			if tabla == 7:
				query = f"SELECT decode(cdtipram,3,'VIDA','DAÑO') TIPO, (SELECT OTVALOR05 FROM TVALOPOL A WHERE A.CDUNIECO=MAIN.CDUNIECO AND A.CDRAMO=MAIN.CDRAMO AND A.ESTADO='M' AND A.NMPOLIZA=MAIN.NMPOLIZA AND A.NMSUPLEM=(SELECT MAX(NMSUPLEM) FROM TVALOPOL B WHERE A.CDUNIECO=B.CDUNIECO AND A.CDRAMO=B.CDRAMO AND A.ESTADO=B.ESTADO AND A.NMPOLIZA=B.NMPOLIZA)) grupo, MAIN.CDUNIECO OFICINA, MAIN.CDRAMO RAMO, MAIN.NMPOLIZA POLIZA, nomtomador(MAIN.CDUNIECO, MAIN.CDRAMO, 'M', MAIN.NMPOLIZA, 9999999999999999999) CONTRATANTE, MAIN.CDAGENTE, MAIN.CDMONEDA, MAIN.NMRECIBO RECIBO, MAIN.SERIERECIBO SERIE_REC, nvl(MAIN.PRIMA_TOTAL,0) PRIMA_TOTAL, nvl(MAIN.PrimaNeta,0) PRIMA_NETA, nvl(MAIN.PORAGEN,0) PORC_COMISION, nvl(MAIN.COM_DERECHO,0) COMISION_DERECHO, nvl(MAIN.COM_NETA,0) COMISION_NETA, (nvl(MAIN.ComisionNeta,0) + nvl(MAIN.IvaComision,0) + nvl(MAIN.IvaRetenido,0)+ nvl(MAIN.IsrComision,0) + nvl(MAIN.ComisionDerecho,0) + nvl(MAIN.IsrComDerecho,0) + nvl(MAIN.IvaRetDerecho,0) + nvl(MAIN.IvaDerecho,0)) TOTAL_COMISION, MAIN.NUMPRELI,MAIN.NMTRANSF, TO_CHAR(MAIN.FEMOVIMI,'DD/MM/YYYY') " \
						f"FECHACORTE FROM ( SELECT G.NUMPRELI,G.NMTRANSF,A.CDMONEDA,A.CDAGENTE,c.cdtipram,A.FEMOVIMI, A.CDUNIECO,A.CDRAMO,A.NMPOLIZA,A.NMRECIBO,NVL(A.SERIERECIBO,0) SERIERECIBO, A.FEINICIOREC, A.FEMOVIMI FEAPLIC,A.DSNOMASEG, NVL(A.PMATOTAL,0) PRIMA_TOTAL,a.PMANETAP PrimaNeta, a.IMPORAGTN COM_NETA,a.PORAGEN, NVL(decode(a.SWAPDER,'A',IMPORAGTND,NULL),0) COM_DERECHO,ROUND(NVL(a.imporagtn,0),2) ComisionNeta, ROUND(NVL(a.imporagtv,0),2) IvaComision, ROUND(NVL(a.imporagtq,0),2) IvaRetenido, ROUND(NVL(a.imporagti,0),2) IsrComision, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtnd,0),0),2) ComisionDerecho, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtid,0),0),2) IsrComDerecho, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtqd,0),0),2) IvaRetDerecho, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtvd,0),0),2) IvaDerecho from tsia_detcom a, TRAMOS C,MINTPSOFT G where TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdramo = c.cdramo AND c.cdtipram IN (3) AND A.CDMONEDA='MXP' AND A.CDPROMOT='{codigo}' AND A.NMPOLIZA=G.NMPOLIZA (+) AND " \
						f"A.NMRECIBO=G.NMRECIBO (+) ) MAIN ORDER BY FEMOVIMI, CDUNIECO,CDRAMO,NMPOLIZA,NMRECIBO"
			if tabla == 8:
				query = f"SELECT decode(cdtipram,3,'VIDA','DAÑO') TIPO, (SELECT OTVALOR05 FROM TVALOPOL A WHERE A.CDUNIECO=MAIN.CDUNIECO AND A.CDRAMO=MAIN.CDRAMO AND A.ESTADO='M' AND A.NMPOLIZA=MAIN.NMPOLIZA AND A.NMSUPLEM=(SELECT MAX(NMSUPLEM) FROM TVALOPOL B WHERE A.CDUNIECO=B.CDUNIECO AND A.CDRAMO=B.CDRAMO AND A.ESTADO=B.ESTADO AND A.NMPOLIZA=B.NMPOLIZA)) grupo, MAIN.CDUNIECO OFICINA, MAIN.CDRAMO RAMO, MAIN.NMPOLIZA POLIZA, nomtomador(MAIN.CDUNIECO, MAIN.CDRAMO, 'M', MAIN.NMPOLIZA, 9999999999999999999) CONTRATANTE, MAIN.CDAGENTE, MAIN.CDMONEDA, MAIN.NMRECIBO RECIBO, MAIN.SERIERECIBO SERIE_REC, nvl(MAIN.PRIMA_TOTAL,0) PRIMA_TOTAL, nvl(MAIN.PrimaNeta,0) PRIMA_NETA, nvl(MAIN.PORAGEN,0) PORC_COMISION, nvl(MAIN.COM_DERECHO,0) COMISION_DERECHO, nvl(MAIN.COM_NETA,0) COMISION_NETA, (nvl(MAIN.ComisionNeta,0) + nvl(MAIN.IvaComision,0) + nvl(MAIN.IvaRetenido,0)+ nvl(MAIN.IsrComision,0) + nvl(MAIN.ComisionDerecho,0) + nvl(MAIN.IsrComDerecho,0) + nvl(MAIN.IvaRetDerecho,0) + nvl(MAIN.IvaDerecho,0)) TOTAL_COMISION, MAIN.NUMPRELI,MAIN.NMTRANSF, TO_CHAR(MAIN.FEMOVIMI,'DD/MM/YYYY') " \
						f"FECHACORTE FROM ( SELECT G.NUMPRELI,G.NMTRANSF,A.CDMONEDA,A.CDAGENTE,c.cdtipram,A.FEMOVIMI, A.CDUNIECO,A.CDRAMO,A.NMPOLIZA,A.NMRECIBO,NVL(A.SERIERECIBO,0) SERIERECIBO, A.FEINICIOREC, A.FEMOVIMI FEAPLIC,A.DSNOMASEG, NVL(A.PMATOTAL,0) PRIMA_TOTAL,a.PMANETAP PrimaNeta, a.IMPORAGTN COM_NETA,a.PORAGEN, NVL(decode(a.SWAPDER,'A',IMPORAGTND,NULL),0) COM_DERECHO,ROUND(NVL(a.imporagtn,0),2) ComisionNeta, ROUND(NVL(a.imporagtv,0),2) IvaComision, ROUND(NVL(a.imporagtq,0),2) IvaRetenido, ROUND(NVL(a.imporagti,0),2) IsrComision, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtnd,0),0),2) ComisionDerecho, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtid,0),0),2) IsrComDerecho, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtqd,0),0),2) IvaRetDerecho, ROUND(DECODE(a.swapder,'A',NVL(a.imporagtvd,0),0),2) IvaDerecho from tsia_detcom a, TRAMOS C,MINTPSOFT G where TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdramo = c.cdramo AND c.cdtipram IN (3) AND A.CDMONEDA='USD' AND A.CDPROMOT='{codigo}' AND A.NMPOLIZA=G.NMPOLIZA (+) AND " \
						f"A.NMRECIBO=G.NMRECIBO (+) ) MAIN ORDER BY FEMOVIMI, CDUNIECO,CDRAMO,NMPOLIZA,NMRECIBO"
			if tabla == 9:
				query = f"select CONCEPTO,SUM(ROUND(IMPORTE,2)) IMPORTE from(SELECT a.cdpromot, a.cdmoneda moneda, decode(a.cdconc,'QD',22, 'V',2 , 'Q',22, 'VD',2, decode(a.cdconc, 'I',2, 'ID',2 , 'PC',1999, 'B',2, 'IC',607, decode(b.dsconc,'Imp Cedular Derecho',607,a.cdcvemov) ) ) movimiento, decode(a.cdconc,'QD','IVA Retenido' , 'V','IVA' , 'Q','IVA Retenido', 'VD','IVA', decode(a.cdconc, 'I','ISR', 'ID','ISR' , 'PC','Pago de Comisión', 'IC','Impuesto Cedular', decode(b.dsconc,'Imp Cedular Derecho','Impuesto Cedular', b.dsconc) ) ) concepto, decode(a.cdconc,'QD',2.1 , 'V',2 , 'Q',2.1, 'VD',2, decode(a.cdconc, 'I',4, 'ID',4 , 'PC',1999, 'B',1.2, 'ND',1,decode(b.dsconc,'Imp Cedular Derecho',6, b.numprior) ) ) numprior , decode(a.cdconc,'QD','Q' , 'V','V' , 'Q','Q', 'VD','V', decode(a.cdconc, 'I','I', 'ID','I' ,decode(b.dsconc,'Imp Cedular Derecho','IC', a.cdconc) )) cdconc, SUM(a.ptimport) importe, SUM(a.ptimport) SBimporte FROM tsia_movprom a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov " \
						f"AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND a.cdcvemov >= 1000 AND A.CDCONC NOT IN ('PC') AND A.CDPROMOT='{hasta}' AND a.CDMONEDA='MXP' GROUP BY a.cdpromot, a.cdmoneda,decode(a.cdconc,'QD',22, 'V',2 , 'Q',22, 'VD',2, decode(a.cdconc, 'I',2, 'ID',2 , 'PC',1999, 'B',2, 'IC',607, decode(b.dsconc,'Imp Cedular Derecho',607,a.cdcvemov) ) ) , decode(a.cdconc,'QD','IVA Retenido' , 'V','IVA' , 'Q','IVA Retenido', 'VD','IVA', decode(a.cdconc, 'I','ISR', 'ID','ISR' , 'PC','Pago de Comisión', 'IC','Impuesto Cedular', decode(b.dsconc,'Imp Cedular Derecho','Impuesto Cedular', b.dsconc) ) ) , decode(a.cdconc,'QD',2.1 , 'V',2 , 'Q',2.1, 'VD',2, decode(a.cdconc, 'I',4, 'ID',4 , 'PC',1999, 'B',1.2, 'ND',1,decode(b.dsconc,'Imp Cedular Derecho',6, b.numprior) ) ) , decode(a.cdconc,'QD','Q' , 'V','V' , 'Q','Q', 'VD','V', decode(a.cdconc, 'I','I', 'ID','I' ,decode(b.dsconc,'Imp Cedular Derecho','IC', a.cdconc) )) union SELECT a.cdpromot, a.cdmoneda moneda, 19 movimiento, 'Subtotal' concepto, 1.9 numprior , 'ST' cdconc, " \
						f"SUM ( DECODE (a.cdconc, 'N', a.ptimport, 'ND', a.ptimport, 'GA', a.ptimport, 'B', a.ptimport, 'CD', a.ptimport, 'DC', a.ptimport, 'BC', a.ptimport, 'PI', a.ptimport, decode(b.dsconc,'Descuento de Comisión', a.ptimport, 0)))Importe, 0 SBIMPorte FROM tsia_movprom a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND a.cdcvemov >= 1000  AND A.CDPROMOT = '{codigo}' AND a.CDMONEDA='MXP' HAVING SUM ( DECODE (a.cdconc, 'N', a.ptimport, 'ND', a.ptimport, 'GA', a.ptimport, 'B', a.ptimport, 'DC', a.ptimport, 'BC', a.ptimport, 'PI', a.ptimport,decode(b.dsconc,'Descuento de Comisión', a.ptimport, 0)) )<> 0 GROUP BY a.cdpromot, a.cdmoneda UNION SELECT T.CDPROMOT, T.MONEDA, T.MOVIMIENTO, T.CONCEPTO, T.NUMPRIOR, T.CDCONC, SUM(T.IMPORTE), SUM(T.SBIMPORTE) FROM (SELECT A.CDPROMOT, A.CDMONEDA MONEDA, 1999 MOVIMIENTO, 'Pago de Comisión' CONCEPTO, 1999 numprior, 'PC' CDCONC, A.PTIMPORT IMPORTE, A.PTIMPORT SBIMPorte FROM tsia_movprom a, tsia_tipmovi b " \
						f"WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND a.cdcvemov = 1600 AND A.CDPROMOT = '{codigo}' AND a.CDMONEDA = 'MXP') T GROUP BY T.CDPROMOT, T.MONEDA, T.MOVIMIENTO, T.CONCEPTO, T.NUMPRIOR, T.CDCONC UNION SELECT A.CDPROMOT, A.CDMONEDA MONEDA, DECODE(a.cdconc, 'I',2, 'V',2, 'Q',22, 'IC',607, 'PC',1999) MOVIMIENTO, DECODE(a.cdconc, 'I','ISR', 'V','IVA', 'Q','IVA Retenido', 'IC','Impuesto Cedular', 'PC','Pago de Comisión' ) CONCEPTO, DECODE(a.cdconc, 'I',4, 'V',2, 'Q',2.1, 'IC',6, 'PC',1999) numprior, A.CDCONC CDCONC, A.PTIMPORT*-1 IMPORTE, A.PTIMPORT*-1 SBIMPorte FROM tsia_movprom a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND a.cdcvemov = 1600 AND a.CDCONC IN ('I','V','Q','IC') AND A.CDPROMOT = '{codigo}' AND a.CDMONEDA='MXP' ) GROUP BY NUMPRIOR,MOVIMIENTO,CDCONC,CONCEPTO order by NUMPRIOR,MOVIMIENTO,CDCONC"
			if tabla == 10:
				query = f"select CONCEPTO,SUM(ROUND(IMPORTE,2)) IMPORTE from(SELECT a.cdpromot, a.cdmoneda moneda, decode(a.cdconc,'QD',22, 'V',2 , 'Q',22, 'VD',2, decode(a.cdconc, 'I',2, 'ID',2 , 'PC',1999, 'B',2, 'IC',607, decode(b.dsconc,'Imp Cedular Derecho',607,a.cdcvemov) ) ) movimiento, decode(a.cdconc,'QD','IVA Retenido' , 'V','IVA' , 'Q','IVA Retenido', 'VD','IVA', decode(a.cdconc, 'I','ISR', 'ID','ISR' , 'PC','Pago de Comisión', 'IC','Impuesto Cedular', decode(b.dsconc,'Imp Cedular Derecho','Impuesto Cedular', b.dsconc) ) ) concepto, decode(a.cdconc,'QD',2.1 , 'V',2 , 'Q',2.1, 'VD',2, decode(a.cdconc, 'I',4, 'ID',4 , 'PC',1999, 'B',1.2, 'ND',1,decode(b.dsconc,'Imp Cedular Derecho',6, b.numprior) ) ) numprior , decode(a.cdconc,'QD','Q' , 'V','V' , 'Q','Q', 'VD','V', decode(a.cdconc, 'I','I', 'ID','I' ,decode(b.dsconc,'Imp Cedular Derecho','IC', a.cdconc) )) cdconc, SUM(a.ptimport) importe, SUM(a.ptimport) SBimporte FROM tsia_movprom a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov AND a.cdconc = b.cdconc " \
						f"AND b.afedocta = 'S' AND a.cdcvemov >= 1000 AND A.CDCONC NOT IN ('PC') AND A.CDPROMOT='{codigo}' AND a.CDMONEDA='USD' GROUP BY a.cdpromot, a.cdmoneda,decode(a.cdconc,'QD',22, 'V',2 , 'Q',22, 'VD',2, decode(a.cdconc, 'I',2, 'ID',2 , 'PC',1999, 'B',2, 'IC',607, decode(b.dsconc,'Imp Cedular Derecho',607,a.cdcvemov) ) ) , decode(a.cdconc,'QD','IVA Retenido' , 'V','IVA' , 'Q','IVA Retenido', 'VD','IVA', decode(a.cdconc, 'I','ISR', 'ID','ISR' , 'PC','Pago de Comisión', 'IC','Impuesto Cedular', decode(b.dsconc,'Imp Cedular Derecho','Impuesto Cedular', b.dsconc) ) ) , decode(a.cdconc,'QD',2.1 , 'V',2 , 'Q',2.1, 'VD',2, decode(a.cdconc, 'I',4, 'ID',4 , 'PC',1999, 'B',1.2, 'ND',1,decode(b.dsconc,'Imp Cedular Derecho',6, b.numprior) ) ) , decode(a.cdconc,'QD','Q' , 'V','V' , 'Q','Q', 'VD','V', decode(a.cdconc, 'I','I', 'ID','I' ,decode(b.dsconc,'Imp Cedular Derecho','IC', a.cdconc) )) union SELECT a.cdpromot, a.cdmoneda moneda, 19 movimiento, 'Subtotal' concepto, 1.9 numprior , 'ST' cdconc, SUM ( DECODE (a.cdconc, 'N', a.ptimport, 'ND', a.ptimport, 'GA', a.ptimport, 'B', a.ptimport, 'CD', " \
						f"a.ptimport, 'DC', a.ptimport, 'BC', a.ptimport, 'PI', a.ptimport, decode(b.dsconc,'Descuento de Comisión', a.ptimport, 0)))Importe, 0 SBIMPorte FROM tsia_movprom a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND a.cdcvemov >= 1000 AND A.CDPROMOT = '{codigo}' AND a.CDMONEDA='USD' HAVING SUM ( DECODE (a.cdconc, 'N', a.ptimport, 'ND', a.ptimport, 'GA', a.ptimport, 'B', a.ptimport, 'DC', a.ptimport, 'BC', a.ptimport, 'PI', a.ptimport,decode(b.dsconc,'Descuento de Comisión', a.ptimport, 0)) )<> 0 GROUP BY a.cdpromot, a.cdmoneda UNION SELECT T.CDPROMOT, T.MONEDA, T.MOVIMIENTO, T.CONCEPTO, T.NUMPRIOR, T.CDCONC, SUM(T.IMPORTE), SUM(T.SBIMPORTE) FROM (SELECT A.CDPROMOT, A.CDMONEDA MONEDA, 1999 MOVIMIENTO, 'Pago de Comisión' CONCEPTO, 1999 numprior, 'PC' CDCONC, A.PTIMPORT IMPORTE, A.PTIMPORT SBIMPorte FROM tsia_movprom a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov " \
						f"AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND a.cdcvemov = 1600 AND A.CDPROMOT = '{codigo}' AND a.CDMONEDA = 'USD') T GROUP BY T.CDPROMOT, T.MONEDA, T.MOVIMIENTO, T.CONCEPTO, T.NUMPRIOR, T.CDCONC UNION SELECT A.CDPROMOT, A.CDMONEDA MONEDA, DECODE(a.cdconc, 'I',2, 'V',2, 'Q',22, 'IC',607, 'PC',1999) MOVIMIENTO, DECODE(a.cdconc, 'I','ISR', 'V','IVA', 'Q','IVA Retenido', 'IC','Impuesto Cedular', 'PC','Pago de Comisión' ) CONCEPTO, DECODE(a.cdconc, 'I',4, 'V',2, 'Q',2.1, 'IC',6, 'PC',1999) numprior, A.CDCONC CDCONC, A.PTIMPORT*-1 IMPORTE, A.PTIMPORT*-1 SBIMPorte FROM tsia_movprom a, tsia_tipmovi b WHERE TRUNC(a.femovimi) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') AND a.cdcvemov = b.cdcvemov AND a.cdconc = b.cdconc AND b.afedocta = 'S' AND a.cdcvemov = 1600 AND a.CDCONC IN ('I','V','Q','IC') AND A.CDPROMOT = '{codigo}' AND a.CDMONEDA='USD' ) GROUP BY NUMPRIOR,MOVIMIENTO,CDCONC,CONCEPTO order by NUMPRIOR,MOVIMIENTO,CDCONC"
			if tabla == 11:
				query = f"SELECT T.FECHA_MOVIMIENTO FECHA_MOVIMIENTO , SUM(T.IMPORTE_DANOS) IMPORTE_DANOS , SUM(T.IMPORTE_VIDA) IMPORTE_VIDA , SUM(T.TOTAL) TOTAL , T.FECHA_PAGO FECHA_PAGO , T.NUM_COMPROBANTE NUM_COMPROBANTE , SUM(T.IMPORTE_PAG_DANOS) IMPORTE_PAG_DANOS, SUM(T.IMPORTE_PAG_VIDA) IMPORTE_PAG_VIDA , SUM(T.TOTAL_PAGADO) TOTAL_PAGADO FROM ( SELECT TO_CHAR(A.FEMOVIMI,'DD/MM/YYYY') FECHA_MOVIMIENTO , DECODE(A.CDCVEMOV,600,A.PTIMPORT,0) IMPORTE_DANOS , DECODE(A.CDCVEMOV,1600,A.PTIMPORT,0) IMPORTE_VIDA , A.PTIMPORT TOTAL , DECODE(NVL(B.CDCOMPRO,0),0,'',TO_CHAR(B.FECHAENV,'DD/MM/YYYY')) FECHA_PAGO , DECODE(NVL(B.CDCOMPRO,0),0,'',DECODE(B.CDMETPAG,'EFT',B.NMTRANSF,'CHK',B.NMCHEQUE,B.CDCOMPRO)) NUM_COMPROBANTE , DECODE(NVL(B.CDCOMPRO,0),0,0,DECODE(B.CDCVEMOV,600,B.PTIMPPAG,0)) IMPORTE_PAG_DANOS, DECODE(NVL(B.CDCOMPRO,0),0,0,DECODE(B.CDCVEMOV,1600,B.PTIMPPAG,0)) IMPORTE_PAG_VIDA , DECODE(NVL(B.CDCOMPRO,0),0,0,B.PTIMPPAG) TOTAL_PAGADO , A.CDPROMOT PROMOTOR , B.NUMPRELI NUMPRELI FROM TSIA_MOVPROM A, MINTPSOFT B WHERE A.CDCVEMOV IN (600,1600) AND TRUNC(A.FEMOVIMI) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') " \
						f"AND A.NUMPRELIQ = B.NUMPRELI(+) AND A.CDCVEMOV = B.CDCVEMOV(+) AND A.CDCONC = B.CDCONC (+) AND A.CDPROMOT = '{codigo}' AND A.CDMONEDA = 'MXP') T GROUP BY T.PROMOTOR,T.NUMPRELI,T.FECHA_MOVIMIENTO,T.FECHA_PAGO,T.NUM_COMPROBANTE ORDER BY T.FECHA_MOVIMIENTO ASC"
			if tabla == 12:
				query = f"SELECT T.FECHA_MOVIMIENTO FECHA_MOVIMIENTO , SUM(T.IMPORTE_DANOS) IMPORTE_DANOS , SUM(T.IMPORTE_VIDA) IMPORTE_VIDA , SUM(T.TOTAL) TOTAL , T.FECHA_PAGO FECHA_PAGO , T.NUM_COMPROBANTE NUM_COMPROBANTE , SUM(T.IMPORTE_PAG_DANOS) IMPORTE_PAG_DANOS, SUM(T.IMPORTE_PAG_VIDA) IMPORTE_PAG_VIDA , SUM(T.TOTAL_PAGADO) TOTAL_PAGADO FROM ( SELECT TO_CHAR(A.FEMOVIMI,'DD/MM/YYYY') FECHA_MOVIMIENTO , DECODE(A.CDCVEMOV,600,A.PTIMPORT,0) IMPORTE_DANOS , DECODE(A.CDCVEMOV,1600,A.PTIMPORT,0) IMPORTE_VIDA , A.PTIMPORT TOTAL , DECODE(NVL(B.CDCOMPRO,0),0,'',TO_CHAR(B.FECHAENV,'DD/MM/YYYY')) FECHA_PAGO , DECODE(NVL(B.CDCOMPRO,0),0,'',DECODE(B.CDMETPAG,'EFT',B.NMTRANSF,'CHK',B.NMCHEQUE,B.CDCOMPRO)) NUM_COMPROBANTE , DECODE(NVL(B.CDCOMPRO,0),0,0,DECODE(B.CDCVEMOV,600,B.PTIMPPAG,0)) IMPORTE_PAG_DANOS, DECODE(NVL(B.CDCOMPRO,0),0,0,DECODE(B.CDCVEMOV,1600,B.PTIMPPAG,0)) IMPORTE_PAG_VIDA , DECODE(NVL(B.CDCOMPRO,0),0,0,B.PTIMPPAG) TOTAL_PAGADO , A.CDPROMOT PROMOTOR , B.NUMPRELI NUMPRELI FROM TSIA_MOVPROM A, MINTPSOFT B WHERE A.CDCVEMOV IN (600,1600) AND TRUNC(A.FEMOVIMI) BETWEEN to_date('{desde}','dd/mm/yyyy') AND to_date('{hasta}','dd/mm/yyyy') " \
						f"AND A.NUMPRELIQ = B.NUMPRELI(+) AND A.CDCVEMOV = B.CDCVEMOV(+) AND A.CDCONC = B.CDCONC (+) AND A.CDPROMOT = '{codigo}' AND A.CDMONEDA = 'USD') T GROUP BY T.PROMOTOR,T.NUMPRELI,T.FECHA_MOVIMIENTO,T.FECHA_PAGO,T.NUM_COMPROBANTE ORDER BY T.FECHA_MOVIMIENTO ASC"
	return query


def getheaderquery(clave,codigo,desde,hasta):
	str= ""
	if clave == 'A':
		str=f"SELECT DISTINCT d.dsnombre Nombre, a.cdagente Clave, d.cdideper RFC, e.dsdomici||', '||h.dsprovin||', '||f.dscoloni||', '||j.dsmunici||DECODE(e.cdpostal,NULL,NULL,', C.P. '||e.cdpostal) DIRECCION, 'Agente' Tipo_Productor, b.cdpromot||' - '||c.dspromot PROMOTORIA, a.cdagente Clave_Agente, PKG_Alea_Algoritmos_Cob.fr_Recup_CLABE(a.cdagente) CUENTA_CLABE, TO_CHAR(SYSDATE,'DD/MM/YYYY') PERIODO_CORTE, '{desde}' FEINICIO, '{hasta}' FEFIN, TO_CHAR((SELECT Min(feenvio) FROM tsia_movagen A ,BON_TPAGLIQ B WHERE B.NUMPRELIQ =A.NUMPRELIQ AND b.FEENVIO BETWEEN TO_DATE('{desde}','dd/mm/yyyy') AND TO_DATE('{hasta}','dd/mm/yyyy') AND a.cdagente= '{codigo}' AND B.CDAGPROU = A.CDAGENTE AND B.CDCVEMOV = A.CDCVEMOV AND B.CDCONC = A.CDCONC ),'DD/MM/YYYY') FECHA_PRELIQ FROM tsia_movagen a, tsia_catagt b, tsia_promotor c, mpersona d,mdomicil e,tcolonia f,tmanteni g,tprovin h,tcodipos i,tmunici j WHERE TRUNC(a.femovimi) BETWEEN TO_DATE('{desde}','DD/MM/YYYY') AND TO_DATE('{hasta}','DD/MM/YYYY') AND a.cdagente = '{codigo}' AND a.cdagente = b.cdagente AND b.cdpromot = c.cdpromot AND b.cdperson = d.cdperson AND d.cdperson = e.cdperson AND e.nmorddom = 1 AND e.cdcoloni = f.cdcoloni AND g.cdtabla = 'TPAISES' AND g.codigo = e.cdpais AND h.cdprovin = j.cdprovin AND a.ptimport <> 0 AND i.cdcodpos = e.cdpostal and j.cdprovin = i.cdprovin and j.cdmunici = i.cdmunici"
	if clave == 'P':
		str=f"SELECT DISTINCT c.dspromot Nombre, a.cdpromot Clave, c.rfc RFC, c.dspoblac||', '||c.dscoloni||', '||c.dscalle||', '||c.nmtelef1||DECODE(c.nmcodpos,NULL,NULL,', C.P. '||c.nmcodpos) DIRECCION, 'Promotor' Tipo_Productor, ' ' PROMOTORIA, c.CDEJECUT Clave_Agente, (SELECT clabe FROM rsa_mcuentas WHERE cdperson=c.cdperson and NMORDCTA=1 UNION SELECT cuenta FROM rsa_mctaint WHERE cdperson = c.cdperson AND NMORDCTA=1) cuenta_bancaria, TO_CHAR(SYSDATE,'DD/MM/YYYY') PERIODO_CORTE, '{desde}' FEINICIO, '{hasta}' FEFIN, TO_CHAR((SELECT Min(feenvio) FROM tsia_movprom A ,BON_TPAGLIQ B WHERE B.NUMPRELIQ =A.NUMPRELIQ AND b.FEENVIO BETWEEN TO_DATE('{desde}','dd/mm/yyyy') AND TO_DATE('{hasta}','dd/mm/yyyy') AND a.cdpromot= '{codigo}' AND B.CDAGPROU = A.cdpromot AND B.CDCVEMOV = A.CDCVEMOV AND B.CDCONC = A.CDCONC ),'DD/MM/YYYY') FECHA_PRELIQ FROM tsia_movprom a, tsia_promotor c, mpersona d, mdomicil e WHERE TRUNC(a.femovimi) BETWEEN TO_DATE('{desde}','DD/MM/YYYY') AND TO_DATE('{hasta}','DD/MM/YYYY') AND a.cdpromot = '{codigo}' AND a.cdpromot = c.cdpromot AND c.cdperson = d.cdperson AND d.cdperson = e.cdperson AND e.nmorddom = 1 AND a.ptimport <> 0"
	return str

def getheaderpdf(tipo,lista_aux,reporte):
	lista=[]
	if tipo=='P':
		lista= [("   ", "   ", "ESTADO DE CUENTA DE "+reporte, "   ", "   "),
				("Nombre del SAT para Sura:", "Seguros SURA S.A. de C.V", "     ", "Nombre del Promotor:", lista_aux[0]),
				("RFC de Sura:", "R.F.C R&S-811221KR6", "   ", "Clave del Promotor:", lista_aux[1]),
				("Domicilio Sura:", "Blvd. Adolfo López Mateos No. 2448", "   ", "RFC del Agente:", lista_aux[2]),
				("Col. Altavista C.P. 01060 Ciudad de México.", "Tel 5723-7999", "   ", "Domicilio del Promotor:", lista_aux[3]),
				("Fecha de generación:", lista_aux[8], "   ", "Tipo de productor:", lista_aux[4]),
				("Desde:", lista_aux[9], "   ", "Promotoría a la que pertenece:", lista_aux[5]),
				("Hasta:", lista_aux[10], "   ", "Clave de productor:", lista_aux[6]),
				("Fecha de Preliquidación:", lista_aux[11], "   ", "Cuenta bancaria dada de alta:", lista_aux[7]),
				("   ", "   ", "   ", "   ", "   ")]
	if tipo=='A':
		lista= [("   ","   ","ESTADO DE CUENTA DE "+reporte,"   ","   "),
				("Nombre del SAT para Sura:", "Seguros SURA S.A. de C.V","     ","Nombre del Agente:",lista_aux[0]),
				("RFC de Sura:", "R.F.C R&S-811221KR6","   ","Clave del Agente:",lista_aux[1]),
				("Domicilio Sura:", "Blvd. Adolfo López Mateos No. 2448","   ","RFC del Agente:",lista_aux[2]),
				("Col. Altavista C.P. 01060 Ciudad de México.", "Tel 5723-7999","   ","Domicilio del Agente:",lista_aux[3]),
				("Fecha de generación:",lista_aux[8],"   ","Tipo de productor:",lista_aux[4]),
				("Desde:",lista_aux[9],"   ","Promotoría a la que pertenece:",lista_aux[5]),
				("Hasta:",lista_aux[10],"   ","Clave de productor:",lista_aux[6]),
				("Fecha de Preliquidación:",lista_aux[11],"   ","Cuenta bancaria dada de alta:",lista_aux[7]),
				("   ","   ","   ","   ","   ")]
	return  lista

if __name__ == '__main__':
    app.run(debug=True, port=SERVER_PORT)
